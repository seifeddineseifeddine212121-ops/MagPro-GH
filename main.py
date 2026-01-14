import arabic_reshaper
import hashlib
import json
import logging
import math
import openpyxl
import os
import random
import re
import shutil
import sqlite3
import sys
import textwrap
import threading
import time
import traceback
import webbrowser
import zipfile
# ==========================================
DEBUG = True
if DEBUG:
    os.environ['KIVY_LOG_LEVEL'] = 'info'
    os.environ['KIVY_NO_CONSOLELOG'] = '0'
else:
    os.environ['KIVY_LOG_LEVEL'] = 'error'
    os.environ['KIVY_NO_CONSOLELOG'] = '1'
# ==========================================
from PIL import Image, ImageDraw, ImageFont
from bidi.algorithm import get_display
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from fpdf import FPDF, XPos, YPos
from kivy.clock import Clock, mainthread
from kivy.config import Config
from kivy.core.text import LabelBase
from kivy.core.window import Window
from kivy.graphics.context_instructions import PushMatrix, PopMatrix, Rotate
from kivy.lang import Builder
from kivy.metrics import dp
from kivy.properties import StringProperty, NumericProperty, ObjectProperty, ListProperty, BooleanProperty, ColorProperty
from kivy.uix.camera import Camera
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.modalview import ModalView
from kivy.uix.recycleboxlayout import RecycleBoxLayout
from kivy.uix.recycleview import RecycleView
from kivy.uix.recycleview.views import RecycleDataViewBehavior
from kivy.uix.spinner import Spinner
from kivy.utils import platform
from kivymd import fonts_path
from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDRaisedButton, MDIconButton, MDFillRoundFlatButton, MDFlatButton
from kivymd.uix.card import MDCard, MDSeparator
from kivymd.uix.dialog import MDDialog
from kivymd.uix.floatlayout import MDFloatLayout
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.label import MDLabel, MDIcon
from kivymd.uix.list import MDList, OneLineListItem, TwoLineAvatarIconListItem, ThreeLineAvatarIconListItem, IconLeftWidget, IconRightWidget, IRightBodyTouch, ILeftBody, OneLineAvatarIconListItem
from kivymd.uix.pickers import MDDatePicker
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.selectioncontrol import MDCheckbox
from kivymd.uix.snackbar import Snackbar
from kivymd.uix.spinner import MDSpinner
from kivymd.uix.textfield import MDTextField
from kivymd.uix.toolbar import MDTopAppBar
from urllib.parse import quote
# ==========================================
if DEBUG:
    Config.set('kivy', 'log_level', 'info')
    Config.set('kivy', 'log_enable', 1)
else:
    Config.set('kivy', 'log_level', 'error')
    Config.set('kivy', 'log_enable', 0)
Config.write()
try:
    from pyzbar.pyzbar import decode
    from PIL import Image as PILImage
except ImportError:
    decode = None
    if DEBUG:
        print('[WARNING] pyzbar library not found. Barcode scanning will be disabled.')
    else:
        pass
# ==========================================
if platform == 'android':
    try:
        from jnius import autoclass
        BluetoothAdapter = autoclass('android.bluetooth.BluetoothAdapter')
        BluetoothDevice = autoclass('android.bluetooth.BluetoothDevice')
        UUID = autoclass('java.util.UUID')
        AudioManager = autoclass('android.media.AudioManager')
        ToneGenerator = autoclass('android.media.ToneGenerator')
    except Exception as e:
        if DEBUG:
            print(f'[ERROR] Android libraries failed to load: {e}')
        else:
            pass
app_dir = os.path.dirname(os.path.abspath(__file__))
FONT_FILE = os.path.join(app_dir, 'font.ttf')
custom_font_loaded = False
try:
    if os.path.exists(FONT_FILE) and os.path.isfile(FONT_FILE):
        if DEBUG:
            print(f'[INFO] Found custom font at: {FONT_FILE}')
        LabelBase.register(name='ArabicFont', fn_regular=FONT_FILE, fn_bold=FONT_FILE)
        LabelBase.register(name='Roboto', fn_regular=FONT_FILE, fn_bold=FONT_FILE)
        LabelBase.register(name='RobotoMedium', fn_regular=FONT_FILE, fn_bold=FONT_FILE)
        LabelBase.register(name='RobotoBold', fn_regular=FONT_FILE, fn_bold=FONT_FILE)
        custom_font_loaded = True
    elif DEBUG:
        print('[WARNING] Custom font file NOT found. Using fallback.')
except Exception as e:
    print(f'[ERROR] Critical error loading custom font: {e}')
if not custom_font_loaded:
    fallback_regular = os.path.join(fonts_path, 'Roboto-Regular.ttf')
    fallback_bold = os.path.join(fonts_path, 'Roboto-Bold.ttf')
    try:
        LabelBase.register(name='ArabicFont', fn_regular=fallback_regular, fn_bold=fallback_bold)
    except Exception:
        LabelBase.register(name='ArabicFont', fn_regular=None, fn_bold=None)
# ==========================================
reshaper = arabic_reshaper.ArabicReshaper(configuration={'delete_harakat': True, 'support_ligatures': True, 'use_unshaped_instead_of_isolated': True})
# ==========================================
KV_BUILDER = '\n<LeftButtonsContainer>:\n    adaptive_width: True\n    spacing: "4dp"\n    padding: "4dp"\n    pos_hint: {"center_y": .5}\n\n<RightButtonsContainer>:\n    adaptive_width: True\n    spacing: "8dp"\n    pos_hint: {"center_y": .5}\n\n<CustomHistoryItem>:\n    orientation: "horizontal"\n    size_hint_y: None\n    height: dp(80)\n    padding: dp(10)\n    spacing: dp(5)\n    radius: [10]\n    elevation: 1\n    ripple_behavior: True\n    md_bg_color: root.bg_color\n    on_release: root.on_tap_action()\n    \n    MDIcon:\n        icon: root.icon\n        theme_text_color: "Custom"\n        text_color: root.icon_color\n        pos_hint: {"center_y": .5}\n        font_size: "32sp"\n        size_hint_x: None\n        width: dp(40)\n        \n    MDBoxLayout:\n        orientation: "vertical"\n        pos_hint: {"center_y": .5}\n        spacing: dp(4)\n        size_hint_x: 0.5\n        \n        MDLabel:\n            text: root.text\n            bold: True\n            font_style: "Subtitle1"\n            font_size: "16sp"\n            theme_text_color: "Primary"\n            shorten: True\n            shorten_from: \'right\'\n            font_name: \'ArabicFont\'\n            markup: True\n            \n        MDLabel:\n            text: root.secondary_text\n            font_style: "Caption"\n            theme_text_color: "Secondary"\n            font_name: \'ArabicFont\'\n            \n    MDLabel:\n        text: root.right_text\n        halign: "right"\n        pos_hint: {"center_y": .5}\n        font_style: "Subtitle2"\n        bold: True\n        theme_text_color: "Custom"\n        text_color: root.icon_color\n        size_hint_x: 0.3\n        font_name: \'ArabicFont\'\n\n    MDIconButton:\n        icon: "pencil"\n        theme_text_color: "Custom"\n        text_color: (0, 0.5, 0.8, 1)\n        pos_hint: {"center_y": .5}\n        on_release: root.on_edit_action()\n\n<ProductRecycleItem>:\n    orientation: \'vertical\'\n    size_hint_y: None\n    height: dp(90)\n    padding: 0\n    spacing: 0\n    \n    MDCard:\n        orientation: \'horizontal\'\n        padding: dp(10)\n        spacing: dp(10)\n        radius: [8]\n        elevation: 1\n        ripple_behavior: True\n        on_release: root.on_tap()\n        md_bg_color: (1, 1, 1, 1)\n        \n        # --- مربع الصورة ---\n        MDCard:\n            size_hint: None, None\n            size: dp(55), dp(55)\n            radius: [8]\n            elevation: 1\n            padding: [0, 0, 0, 0]\n            spacing: 0\n            pos_hint: {\'center_y\': .5}\n            md_bg_color: (0.95, 0.95, 0.95, 1)\n            ripple_behavior: True\n            on_release: root.on_zoom()\n\n            MDFloatLayout:\n                size_hint: 1, 1\n                \n                MDIcon:\n                    icon: root.icon_name\n                    theme_text_color: "Custom"\n                    text_color: root.icon_color\n                    pos_hint: {\'center_x\': .5, \'center_y\': .5}\n                    font_size: \'30sp\'\n                    opacity: 0 if root.image_path else 1\n\n                FitImage:\n                    source: root.image_path if root.image_path else \'\'\n                    radius: [8]\n                    mipmap: True\n                    opacity: 1 if root.image_path else 0\n                    size_hint: 1, 1\n                    pos_hint: {\'center_x\': .5, \'center_y\': .5}\n                    fit_mode: "cover"\n\n        MDBoxLayout:\n            orientation: \'vertical\'\n            pos_hint: {\'center_y\': .5}\n            spacing: dp(5)\n            \n            # --- اسم المنتج ---\n            MDLabel:\n                text: root.text_name\n                font_style: "Subtitle1"\n                bold: True\n                text_size: self.width, None\n                max_lines: 2\n                halign: \'left\'\n                valign: \'top\'\n                shorten: False\n                font_size: \'16sp\'\n                theme_text_color: "Custom"\n                text_color: (0.1, 0.1, 0.1, 1)\n                font_name: \'ArabicFont\'\n            \n            MDBoxLayout:\n                orientation: \'horizontal\'\n                spacing: dp(5)\n                \n                # --- السعر (معدل ليظهر كاملاً) ---\n                MDLabel:\n                    text: root.text_price\n                    font_style: "H6"\n                    theme_text_color: "Custom"\n                    text_color: root.price_color\n                    bold: True\n                    size_hint_x: None\n                    adaptive_width: True\n                    font_size: \'19sp\'\n                    font_name: \'ArabicFont\'\n                    max_lines: 1\n                    shorten: False\n                \n                # --- المخزون ---\n                MDLabel:\n                    text: root.text_stock\n                    theme_text_color: "Custom"\n                    text_color: (0.1, 0.1, 0.1, 1)\n                    halign: \'right\'\n                    size_hint_x: 1\n                    bold: True\n                    font_size: \'13sp\'\n                    max_lines: 1\n                    shorten: True\n                    font_name: \'ArabicFont\'\n\n<ProductRecycleView>:\n    viewclass: \'ProductRecycleItem\'\n    RecycleBoxLayout:\n        default_size: None, dp(95)\n        default_size_hint: 1, None\n        size_hint_y: None\n        height: self.minimum_height\n        orientation: \'vertical\'\n        spacing: dp(4)\n        padding: dp(5)\n\n<HistoryRecycleItem>:\n    orientation: "horizontal"\n    size_hint_y: None\n    height: dp(90)\n    padding: [dp(8), dp(5), dp(8), dp(5)]\n    spacing: dp(5)\n    radius: [10]\n    elevation: 1\n    ripple_behavior: True\n    md_bg_color: root.bg_color\n    on_release: root.on_tap()\n\n    MDIcon:\n        icon: root.icon_name\n        theme_text_color: "Custom"\n        text_color: root.icon_color\n        pos_hint: {"center_y": .5}\n        font_size: "30sp"\n        size_hint_x: None\n        width: dp(35)\n\n    MDBoxLayout:\n        orientation: "vertical"\n        pos_hint: {"center_y": .5}\n        adaptive_height: True\n        spacing: dp(3)\n        size_hint_x: 1\n\n        MDLabel:\n            text: root.text_primary\n            bold: True\n            font_style: "Subtitle1"\n            font_size: "15sp"\n            theme_text_color: "Primary"\n            size_hint_y: None\n            adaptive_height: True \n            text_size: self.width, None\n            halign: \'left\'\n            shorten: False\n            max_lines: 2\n            font_name: \'ArabicFont\'\n            markup: True\n\n        MDLabel:\n            text: root.text_secondary\n            font_style: "Caption"\n            font_size: "12sp"\n            theme_text_color: "Secondary"\n            font_name: \'ArabicFont\'\n            size_hint_y: None\n            adaptive_height: True\n            text_size: self.width, None\n            halign: \'left\'\n            shorten: True\n            shorten_from: \'right\'\n\n    MDLabel:\n        text: root.text_amount\n        halign: "right"\n        pos_hint: {"center_y": .5}\n        font_style: "Subtitle2"\n        bold: True\n        theme_text_color: "Custom"\n        text_color: root.icon_color\n        size_hint_x: None\n        adaptive_width: True\n        width: self.texture_size[0]\n        padding_x: dp(5)\n        font_name: \'ArabicFont\'\n\n<HistoryRecycleView>:\n    viewclass: \'HistoryRecycleItem\'\n    RecycleBoxLayout:\n        default_size: None, dp(95)\n        default_size_hint: 1, None\n        size_hint_y: None\n        height: self.minimum_height\n        orientation: \'vertical\'\n        spacing: dp(5)\n        padding: dp(5)\n\n<EntityRecycleItem>:\n    orientation: "horizontal"\n    size_hint_y: None\n    height: dp(80)\n    padding: dp(10)\n    spacing: dp(15)\n    ripple_behavior: True\n    md_bg_color: (1, 1, 1, 1)\n    radius: [0]\n    on_release: root.on_tap()\n\n    MDIcon:\n        icon: root.icon_name\n        theme_text_color: "Custom"\n        text_color: root.icon_color\n        pos_hint: {"center_y": .5}\n        font_size: "32sp"\n        size_hint_x: None\n        width: dp(40)\n\n    MDBoxLayout:\n        orientation: "vertical"\n        pos_hint: {"center_y": .5}\n        size_hint_x: 1\n        spacing: dp(4)\n\n        MDLabel:\n            text: root.text_name\n            bold: True\n            font_style: "Subtitle1"\n            font_name: \'ArabicFont\'\n            theme_text_color: "Custom"\n            text_color: (0.1, 0.1, 0.1, 1)\n            shorten: True\n            shorten_from: \'right\'\n            valign: \'center\'\n\n        MDLabel:\n            text: root.text_balance\n            font_style: "Caption"\n            font_name: \'ArabicFont\'\n            markup: True\n            theme_text_color: "Secondary"\n            valign: \'top\'\n\n<EntityRecycleView>:\n    viewclass: \'EntityRecycleItem\'\n    RecycleBoxLayout:\n        default_size: None, dp(80)\n        default_size_hint: 1, None\n        size_hint_y: None\n        height: self.minimum_height\n        orientation: \'vertical\'\n        spacing: dp(2)\n        padding: dp(0)\n\n<MgmtEntityRecycleItem>:\n    orientation: "horizontal"\n    size_hint_y: None\n    height: dp(80)\n    padding: dp(10)\n    spacing: dp(5)\n    ripple_behavior: True\n    md_bg_color: (1, 1, 1, 1)\n    on_release: root.on_pay()\n\n    MDIcon:\n        icon: "account-circle"\n        theme_text_color: "Custom"\n        text_color: (0.5, 0.5, 0.5, 1)\n        pos_hint: {"center_y": .5}\n        font_size: "32sp"\n        size_hint_x: None\n        width: dp(40)\n\n    MDBoxLayout:\n        orientation: "vertical"\n        pos_hint: {"center_y": .5}\n        size_hint_x: 1\n        spacing: dp(2)\n        padding: [dp(10), 0, 0, 0]\n\n        MDLabel:\n            text: root.text_name\n            bold: True\n            font_style: "Subtitle1"\n            font_name: \'ArabicFont\'\n            theme_text_color: "Custom"\n            text_color: (0.1, 0.1, 0.1, 1)\n            shorten: True\n            shorten_from: \'right\'\n            halign: "left"\n\n        MDLabel:\n            text: root.text_balance\n            font_style: "Caption"\n            font_name: \'ArabicFont\'\n            markup: True\n            theme_text_color: "Secondary"\n            halign: "left"\n\n    MDIconButton:\n        icon: "clock-time-eight-outline"\n        theme_text_color: "Custom"\n        text_color: (0, 0.5, 0.5, 1)\n        pos_hint: {"center_y": .5}\n        on_release: root.on_history()\n\n<MgmtEntityRecycleView>:\n    viewclass: \'MgmtEntityRecycleItem\'\n    RecycleBoxLayout:\n        default_size: None, dp(80)\n        default_size_hint: 1, None\n        size_hint_y: None\n        height: self.minimum_height\n        orientation: \'vertical\'\n        spacing: dp(2)\n        padding: dp(0)\n\n<CartRecycleItem>:\n    orientation: "horizontal"\n    size_hint_y: None\n    height: dp(85)\n    padding: [dp(15), 0, 0, 0]\n    md_bg_color: 1, 1, 1, 1\n    radius: [0]\n    ripple_behavior: True\n    on_release: root.on_tap()\n\n    MDBoxLayout:\n        orientation: "vertical"\n        pos_hint: {"center_y": .5}\n        adaptive_height: True\n        spacing: dp(4)\n\n        MDLabel:\n            text: root.text_name\n            font_style: "Subtitle1"\n            bold: True\n            theme_text_color: "Primary"\n            adaptive_height: True\n            font_name: \'ArabicFont\'\n\n        MDLabel:\n            text: root.text_details\n            font_size: "16sp"\n            theme_text_color: "Custom"\n            text_color: root.details_color\n            bold: True\n            adaptive_height: True\n            font_name: \'ArabicFont\'\n\n    MDIconButton:\n        icon: "delete"\n        theme_text_color: "Custom"\n        text_color: (0.9, 0, 0, 1)\n        pos_hint: {"center_y": .5}\n        icon_size: "24sp"\n        on_release: root.on_delete()\n\n<CartRecycleView>:\n    viewclass: \'CartRecycleItem\'\n    RecycleBoxLayout:\n        default_size: None, dp(85)\n        default_size_hint: 1, None\n        size_hint_y: None\n        height: self.minimum_height\n        orientation: \'vertical\'\n        spacing: dp(1)\n'
# ==========================================
class AppConstants:
    DEBUG = False
    DB_NAME = 'magpro_local.db'
    FONT_ARABIC = 'ArabicFont'
    DEFAULT_CLIENT_NAME = 'COMPTOIR'
    DEFAULT_SUPPLIER_NAME = 'COMPTOIR'
    DOC_TRANSLATIONS = {'BV': 'BON DE VENTE', 'BA': "BON D'ACHAT", 'FC': 'FACTURE', 'FF': 'FACTURE ACHAT', 'RC': 'RETOUR CLIENT', 'RF': 'RETOUR FOURNISSEUR', 'TR': 'BON DE TRANSFERT', 'FP': 'FACTURE PROFORMA', 'DP': 'BON DE COMMANDE', 'BI': 'BON INITIAL', 'CLIENT_PAY': 'VERSEMENT', 'SUPPLIER_PAY': 'REGLEMENT'}
    STOCK_MOVEMENTS = {'BV': -1, 'SALE': -1, 'FC': -1, 'INVOICE_SALE': -1, 'RF': -1, 'RETURN_PURCHASE': -1, 'BA': 1, 'PURCHASE': 1, 'FF': 1, 'INVOICE_PURCHASE': 1, 'RC': 1, 'RETURN_SALE': 1, 'BI': 1, 'INITIAL': 1, 'FP': 0, 'PROFORMA': 0, 'DP': 0, 'ORDER': 0, 'CLIENT_PAY': 0, 'SUPPLIER_PAY': 0, 'TR': 0, 'TRANSFER': 0}
    FINANCIAL_FACTORS = {'BV': 1, 'SALE': 1, 'FC': 1, 'INVOICE_SALE': 1, 'BA': 1, 'PURCHASE': 1, 'FF': 1, 'INVOICE_PURCHASE': 1, 'BI': 1, 'RC': -1, 'RETURN_SALE': -1, 'RF': -1, 'RETURN_PURCHASE': -1, 'CLIENT_PAY': -1, 'SUPPLIER_PAY': -1, 'VERSEMENT': -1, 'REGLEMENT': -1, 'FP': 0, 'PROFORMA': 0, 'DP': 0, 'ORDER': 0, 'TR': 0, 'TRANSFER': 0}
    DOC_VISUALS = {'BV': {'name': 'Vente', 'icon': 'cart', 'color': (0, 0.5, 0.8, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'SALE': {'name': 'Vente', 'icon': 'cart', 'color': (0, 0.5, 0.8, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'BA': {'name': 'Achat', 'icon': 'truck', 'color': (1, 0.6, 0, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'PURCHASE': {'name': 'Achat', 'icon': 'truck', 'color': (1, 0.6, 0, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'RC': {'name': 'Retour CL.', 'icon': 'keyboard-return', 'color': (0.8, 0, 0, 1), 'bg': (1, 0.95, 0.95, 1)}, 'RF': {'name': 'Retour FR.', 'icon': 'undo', 'color': (0, 0.6, 0.6, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'FC': {'name': 'Facture', 'icon': 'file-document', 'color': (0, 0, 0.8, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'FF': {'name': 'Facture Achat', 'icon': 'file-document-edit', 'color': (1, 0.3, 0, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'TR': {'name': 'Transfert', 'icon': 'swap-horizontal', 'color': (0.5, 0, 0.5, 1), 'bg': (0.95, 0.9, 1, 1)}, 'TRANSFER': {'name': 'Transfert', 'icon': 'swap-horizontal', 'color': (0.5, 0, 0.5, 1), 'bg': (0.95, 0.9, 1, 1)}, 'CLIENT_PAY': {'name': 'Versement', 'icon': 'cash-plus', 'color': (0, 0.7, 0, 1), 'bg': (0.9, 1, 0.9, 1)}, 'SUPPLIER_PAY': {'name': 'Règlement', 'icon': 'cash-minus', 'color': (0.8, 0, 0, 1), 'bg': (1, 0.9, 0.9, 1)}, 'BI': {'name': 'Bon Initial', 'icon': 'database-plus', 'color': (0, 0.5, 0.5, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'FP': {'name': 'Proforma', 'icon': 'file-eye', 'color': (0.5, 0, 0.5, 1), 'bg': (0.98, 0.98, 0.98, 1)}, 'DP': {'name': 'Commande', 'icon': 'clipboard-list', 'color': (0, 0.6, 0.6, 1), 'bg': (0.98, 0.98, 0.98, 1)}}
    YEARLY_RESET_SEQUENCES = ['FC', 'FF', 'FP', 'DP']
    APPLY_STAMP_DUTY = ['FC']
    MODE_TITLES = {'sale': 'Vente (BV)', 'purchase': 'Achat (BA)', 'return_sale': 'Retour (RC)', 'return_purchase': 'Retour (RF)', 'transfer': 'Transfert (TR)', 'manage_products': 'Produits', 'invoice_sale': 'Facture (FC)', 'invoice_purchase': 'Facture (FF)', 'proforma': 'Proforma (FP)', 'order_purchase': 'Commande (DP)', 'client_payment': 'Versement', 'supplier_payment': 'Règlement'}
    MODE_COLORS = {'sale': 'Green', 'purchase': 'Orange', 'return_sale': 'Red', 'return_purchase': 'Teal', 'transfer': 'Purple', 'manage_products': 'Blue', 'invoice_sale': 'Blue', 'invoice_purchase': 'DeepOrange', 'proforma': 'Purple', 'order_purchase': 'Teal', 'client_payment': 'Teal', 'supplier_payment': 'Brown'}

    @staticmethod
    def get_entity_type(mode):
        supplier_modes = ['purchase', 'return_purchase', 'invoice_purchase', 'order_purchase', 'bi', 'supplier_payment']
        if mode in supplier_modes:
            return 'supplier'
        return 'account'

    @staticmethod
    def calculate_stamp_duty(amount):
        try:
            val = float(amount)
        except:
            return 0.0
        if val <= 300:
            return 0.0
        units = math.ceil(val / 100.0)
        if val <= 30000:
            duty = units * 1.0
        elif val <= 100000:
            duty = units * 1.5
        else:
            duty = units * 2.0
        return float(max(5.0, math.ceil(duty)))

def to_decimal(value, default='0.00'):
    if value is None or value == '':
        return Decimal(default)
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(default)

def quantize_decimal(d, precision='0.01'):
    if not isinstance(d, Decimal):
        d = to_decimal(d)
    return d.quantize(Decimal(precision), rounding=ROUND_HALF_UP)

def format_number_simple(num):
    try:
        d = to_decimal(num)
        return '{:,.2f}'.format(d).replace(',', ' ').replace('.', ',')
    except:
        return '0,00'

def format_warranty_days_fr(days):
    if not isinstance(days, int) or days <= 0:
        return 'Aucune'
    if days == 1:
        return '1 jour'
    parts = []
    years = days // 365
    if years > 0:
        parts.append(f"{years} an{('s' if years > 1 else '')}")
        days %= 365
    months = days // 30
    if months > 0:
        parts.append(f'{months} mois')
        days %= 30
    if days > 0:
        parts.append(f"{days} jour{('s' if days > 1 else '')}")
    return ' et '.join(parts) if parts else 'Aucune'

def number_to_words_fr(n_in):
    try:
        n_decimal = to_decimal(n_in)
    except (ValueError, TypeError):
        return ''
    if n_decimal.is_zero():
        return 'zéro'
    is_negative = n_decimal < 0
    if is_negative:
        n_decimal = abs(n_decimal)
    integer_part = int(n_decimal)
    decimal_part = int((n_decimal - integer_part) * 100)
    units = ['', 'un', 'deux', 'trois', 'quatre', 'cinq', 'six', 'sept', 'huit', 'neuf']
    teens = ['dix', 'onze', 'douze', 'treize', 'quatorze', 'quinze', 'seize', 'dix-sept', 'dix-huit', 'dix-neuf']
    tens = ['', 'dix', 'vingt', 'trente', 'quarante', 'cinquante', 'soixante', 'soixante-dix', 'quatre-vingt', 'quatre-vingt-dix']

    def get_words(num):
        if num == 0:
            return []
        if num < 10:
            return [units[num]]
        if num < 20:
            return [teens[num - 10]]
        if num < 70:
            word_parts = [tens[num // 10]]
            rem = num % 10
            if rem > 0:
                if rem == 1 and num // 10 not in [7, 9]:
                    word_parts.append('-et-un')
                else:
                    word_parts.append(f'-{units[rem]}')
            return [''.join(word_parts)]
        if num < 80:
            return [f'soixante-{get_words(num - 60)[0]}']
        if num < 100:
            base = 'quatre-vingt'
            rem = num - 80
            if rem == 0:
                return [base + 's']
            return [f'{base}-{get_words(rem)[0]}']
        if num < 1000:
            prefix_val = num // 100
            prefix = [] if prefix_val == 1 else get_words(prefix_val)
            rest = get_words(num % 100)
            base = 'cent'
            if prefix_val > 1 and (not rest):
                base += 's'
            return prefix + [base] + rest
        if num < 1000000:
            prefix_val = num // 1000
            prefix = [] if prefix_val == 1 else get_words(prefix_val)
            rest = get_words(num % 1000)
            return prefix + ['mille'] + rest
        if num < 2000000:
            return ['un', 'million'] + get_words(num % 1000000)
        if num < 1000000000:
            return get_words(num // 1000000) + ['millions'] + get_words(num % 1000000)
        return ['nombre', 'trop', 'grand']
    integer_words_list = get_words(integer_part) if integer_part > 0 else []
    decimal_words_list = get_words(decimal_part) if decimal_part > 0 else []
    result = []
    if is_negative:
        result.append('moins')
    if integer_words_list:
        result.extend(integer_words_list)
    if decimal_words_list:
        if integer_words_list:
            result.append('et')
        result.extend(decimal_words_list)
        result.append('centime' + ('s' if decimal_part > 1 else ''))
    return ' '.join(result).strip().replace(' -', '-')

def _rows_to_dicts(cursor, rows):
    if rows is None:
        return None
    if isinstance(rows, list):
        return [dict(row) for row in rows]
    return dict(rows)

class PDF(FPDF):

    def __init__(self, *args, **kwargs):
        logging.getLogger('fontTools').setLevel(logging.ERROR)
        logging.getLogger('fontTools.subset').setLevel(logging.ERROR)
        logging.getLogger('fontTools.ttLib').setLevel(logging.ERROR)
        super().__init__(*args, **kwargs)
        self.store_info = {}
        self.entity_info = {}
        self.doc_info = {}
        self.totals = {}
        self.payment_info = {}
        self.amount_in_words = ''
        self.balance_data = None
        self._load_fonts()

    def _load_fonts(self):
        try:
            if os.path.exists(FONT_FILE):
                self.add_font('arabic', '', FONT_FILE)
                self.add_font('arabic', 'B', FONT_FILE)
            else:
                self.add_font('arabic', '')
                self.add_font('arabic', 'B', '')
        except Exception:
            pass

    def is_arabic(self, text):
        return bool(re.search('[\u0600-ۿ]', str(text)))

    def smart_text(self, text):
        safe_text = str(text) if text is not None else ''
        if self.is_arabic(safe_text):
            try:
                return get_display(arabic_reshaper.reshape(safe_text))
            except:
                return safe_text
        return safe_text

    def header(self):
        if self.page_no() == 1:
            logo_path_str = self.store_info.get('store_logo')
            if logo_path_str:
                app_dir = os.path.dirname(os.path.abspath(__file__))
                IMAGES_DIR = os.path.join(app_dir, 'images')
                logo_full_path = os.path.join(IMAGES_DIR, logo_path_str)
                if os.path.exists(logo_full_path):
                    self.image(logo_full_path, 10, 8, 33)
            store_name = self.smart_text(self.store_info.get('store_name', ''))
            info_lines = []
            if (activity := self.smart_text(self.store_info.get('store_activity'))):
                info_lines.append(activity)
            if (address := self.smart_text(self.store_info.get('store_address'))):
                info_lines.append(address)
            phone_email = ' / '.join(filter(None, [f"Tél: {self.store_info.get('store_phone')}" if self.store_info.get('store_phone') else None, f"Email: {self.store_info.get('store_email')}" if self.store_info.get('store_email') else None]))
            if phone_email:
                info_lines.append(phone_email)
            doc_type = self.doc_info.get('doc_type')
            if doc_type not in ('BV', 'BA'):
                rc_nif = ' / '.join(filter(None, [f"RC: {self.store_info.get('store_rc')}" if self.store_info.get('store_rc') else None, f"NIF: {self.store_info.get('store_nif')}" if self.store_info.get('store_nif') else None]))
                if rc_nif:
                    info_lines.append(rc_nif)
                nai_nis = ' / '.join(filter(None, [f"N.Art: {self.store_info.get('store_nai')}" if self.store_info.get('store_nai') else None, f"NIS: {self.store_info.get('store_nis')}" if self.store_info.get('store_nis') else None]))
                if nai_nis:
                    info_lines.append(nai_nis)
            self.set_y(8)
            block_width = 115
            x_pos_store_info = self.w - self.r_margin - block_width - 25 if logo_path_str else (self.w - block_width) / 2
            self.set_x(x_pos_store_info)
            self.set_font('arabic', 'B', 24)
            self.cell(block_width, 10, store_name, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            self.set_x(x_pos_store_info)
            self.set_font('arabic', '', 9)
            self.multi_cell(block_width, 5, '\n'.join(info_lines), align='C')
            self.ln(10)
            y_start_boxes = self.get_y()
            self.set_y(y_start_boxes)
            self.set_x(-100)
            self.set_font('arabic', 'B', 10)
            info_text = f"Date: {self.doc_info.get('date', '')}\n{self.doc_info.get('doc_name_fr', '')}\nN°: {self.doc_info.get('doc_number', '')}"
            order_number = self.doc_info.get('order_number')
            if doc_type in ('FC', 'FF') and order_number:
                info_text += f'\nN° Commande: {order_number}'
            self.multi_cell(90, 7, info_text, border=1, align='L')
            y_invoice_end = self.get_y()
            self.set_y(y_start_boxes)
            self.set_x(10)
            display_entity_name = self.entity_info.get('name', '')
            client_lines = []
            if (addr := self.smart_text(self.entity_info.get('address'))):
                client_lines.append(f'Adresse: {addr}')
            phone_email_line = ' / '.join(filter(None, [f"Tél: {self.entity_info.get('phone')}" if self.entity_info.get('phone') else None, f"Email: {self.entity_info.get('email')}" if self.entity_info.get('email') else None]))
            if phone_email_line:
                client_lines.append(phone_email_line)
            rc_nif_line = ' / '.join(filter(None, [f"RC: {self.entity_info.get('rc')}" if self.entity_info.get('rc') else None, f"NIF: {self.entity_info.get('nif')}" if self.entity_info.get('nif') else None]))
            if rc_nif_line:
                client_lines.append(rc_nif_line)
            nai_nis_line = ' / '.join(filter(None, [f"N.Art: {self.entity_info.get('nai')}" if self.entity_info.get('nai') else None, f"NIS: {self.entity_info.get('nis')}" if self.entity_info.get('nis') else None]))
            if nai_nis_line:
                client_lines.append(nai_nis_line)
            if (payment := self.doc_info.get('payment_method')):
                client_lines.append(f'Mode de Paiement: {payment}')
            client_details_rest = '\n'.join(client_lines)
            full_client_text = f"{self.entity_info.get('label', 'Client')}: {self.smart_text(display_entity_name)}\n{client_details_rest}"
            self.set_font('arabic', 'B', 9)
            self.multi_cell(95, 5, full_client_text, border=1, align='L')
            y_client_end = self.get_y()
            self.set_y(max(y_invoice_end, y_client_end))
            self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')

    def _draw_final_page_footer_content(self):
        self.ln(2)
        if self.amount_in_words:
            self.set_x(10)
            self.set_font('arabic', '', 10)
            self.multi_cell(0, 5, self.smart_text(f'Arrêté la présente facture à la somme de : {self.amount_in_words}'), align='L')
            self.ln(2)
        current_y_pos = self.get_y()
        y_after_totals = y_after_left_content = current_y_pos
        pdf_w = self.w
        doc_type = self.doc_info.get('doc_type')
        total_labels = [('Total HT:', 'total_ht'), ('Total TVA:', 'total_tva'), ('Total Remises:', 'total_discount'), ('Droit de Timbre:', 'stamp_duty'), ('Net à Payer:', 'final_total')]
        if doc_type in ('BV', 'BA'):
            total_labels = [item for item in total_labels if item[1] != 'total_tva']
        start_x_position = pdf_w - self.r_margin - 70
        self.set_y(current_y_pos)
        for label_text, key in total_labels:
            if (value := self.totals.get(key)) is not None:
                d_val = to_decimal(value)
                if d_val == 0 and key != 'final_total' and (key != 'total_ht'):
                    continue
                is_final = key == 'final_total'
                self.set_font('arabic', 'B' if is_final else '', 10)
                self.set_x(start_x_position)
                self.cell(30, 6, label_text, border=1, align='C')
                self.cell(40, 6, f'{format_number_simple(d_val)} DA', border=1, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        entity_name = self.entity_info.get('name', '')
        defaults = ['COMPTOIR']
        is_default_entity = any((x in str(entity_name).lower() for x in defaults))
        excluded_doc_types = ['BI', 'DP', 'FP', 'TR']
        should_show_payment = doc_type not in excluded_doc_types and (not is_default_entity)
        if self.payment_info and should_show_payment:
            paid_amount = to_decimal(self.payment_info.get('amount', '0'))
            self.set_font('arabic', 'B', 10)
            self.set_x(start_x_position)
            self.cell(30, 6, 'Montant Payé:', border=1, align='C')
            self.cell(40, 6, f'{format_number_simple(paid_amount)} DA', border=1, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            final_total = to_decimal(self.totals.get('final_total', 0))
            if paid_amount < final_total:
                remaining = final_total - paid_amount
                self.set_font('arabic', 'B', 10)
                self.set_x(start_x_position)
                self.cell(30, 6, 'Reste à Payer:', border=1, align='C')
                self.set_text_color(0, 0, 0)
                self.cell(40, 6, f'{format_number_simple(remaining)} DA', border=1, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        y_after_totals = self.get_y()
        self.set_y(current_y_pos)
        self.set_x(10)
        if self.balance_data and should_show_payment:
            self.set_font('arabic', 'B', 10)
            self.cell(60, 6, 'Situation du Solde:', border='B', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.set_font('arabic', '', 10)
            self.ln(2)
            old_balance_str = f"Ancien Solde:   {format_number_simple(self.balance_data['old_balance'])} DA"
            trans_amount_str = f"Opération:      {format_number_simple(self.balance_data['transaction_amount'])} DA"
            new_balance_str = f"Nouveau Solde: {format_number_simple(self.balance_data['new_balance'])} DA"
            self.cell(0, 5, old_balance_str, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.cell(0, 5, trans_amount_str, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.ln(1)
            self.set_font('arabic', 'B', 10)
            self.cell(0, 5, new_balance_str, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            y_after_left_content = self.get_y()
        self.set_y(max(y_after_totals, y_after_left_content) + 5)

    def draw_table_with_fill(self, headers, data, col_widths):
        full_footer_margin = 85
        simple_footer_margin = 20

        def draw_vertical_lines(start_y, end_y):
            x_pos = self.l_margin
            self.line(x_pos, start_y, x_pos, end_y)
            for width in col_widths:
                x_pos += width
                self.line(x_pos, start_y, x_pos, end_y)
        self.set_font('arabic', 'B', 9)
        self.set_fill_color(230, 230, 230)
        for i, header in enumerate(headers):
            self.cell(col_widths[i], 8, self.smart_text(header), border=1, align='C', fill=True)
        self.ln()
        table_body_start_y = self.get_y()
        self.set_font('arabic', '', 9)
        for row_index, row in enumerate(data):
            is_last_item = row_index == len(data) - 1
            row_height = 7
            designation_col_index = 2
            if len(row) > designation_col_index:
                lines = self.multi_cell(col_widths[designation_col_index], 7, self.smart_text(row[designation_col_index]), dry_run=True, output='LINES')
                row_height = max(7, len(lines) * 7)
            required_margin = full_footer_margin if is_last_item else simple_footer_margin
            if self.get_y() + row_height > self.h - required_margin:
                draw_vertical_lines(table_body_start_y, self.get_y())
                self.cell(sum(col_widths), 0, '', border='T')
                self.add_page()
                self.set_font('arabic', 'B', 9)
                for i, header in enumerate(headers):
                    self.cell(col_widths[i], 8, self.smart_text(header), border=1, align='C', fill=True)
                self.ln()
                self.set_font('arabic', '', 9)
                table_body_start_y = self.get_y()
            y_before_row = self.get_y()
            x_before_row = self.get_x()
            for i, (cell_data, width) in enumerate(zip(row, col_widths)):
                self.set_xy(x_before_row + sum(col_widths[:i]), y_before_row)
                align = 'L' if i == designation_col_index else 'C'
                self.multi_cell(width, 7, self.smart_text(str(cell_data)), border=0, align=align)
            self.set_y(y_before_row + row_height)
        final_table_bottom_y = self.h - full_footer_margin
        current_y = self.get_y()
        if current_y < final_table_bottom_y:
            draw_vertical_lines(table_body_start_y, final_table_bottom_y)
            self.set_y(final_table_bottom_y)
        else:
            draw_vertical_lines(table_body_start_y, current_y)
        self.cell(sum(col_widths), 0, '', border='T')
        self._draw_final_page_footer_content()

class DatabaseManager:

    def __init__(self, db_name='magpro_local.db'):
        if platform == 'android':
            try:
                from jnius import autoclass
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                files_dir = PythonActivity.mActivity.getFilesDir().getAbsolutePath()
                self.db_name = os.path.join(files_dir, db_name)
            except:
                self.db_name = db_name
        else:
            app_dir = os.path.dirname(os.path.abspath(__file__))
            self.db_name = os.path.join(app_dir, db_name)
        self.conn = None
        self.force_clean_startup()
        self.create_tables()
        self._migrate_db_for_gps()
        self.seed_data()

    def force_clean_startup(self):
        wal_path = self.db_name + '-wal'
        shm_path = self.db_name + '-shm'
        if os.path.exists(wal_path) or os.path.exists(shm_path):
            print('[INIT] Cleaning up leftover DB files...')
            try:
                temp_conn = sqlite3.connect(self.db_name)
                temp_conn.execute('PRAGMA wal_checkpoint(TRUNCATE);')
                temp_conn.execute('PRAGMA journal_mode=DELETE;')
                temp_conn.commit()
                temp_conn.close()
                time.sleep(0.1)
            except Exception as e:
                print(f'[INIT] DB Merge Warning: {e}')
            try:
                if os.path.exists(wal_path):
                    os.remove(wal_path)
                if os.path.exists(shm_path):
                    os.remove(shm_path)
            except Exception as e:
                print(f'[INIT] Force Delete Warning: {e}')

    def _migrate_db_for_gps(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            try:
                cursor.execute("ALTER TABLE clients ADD COLUMN gps_location TEXT DEFAULT ''")
            except:
                pass
            try:
                cursor.execute("ALTER TABLE suppliers ADD COLUMN gps_location TEXT DEFAULT ''")
            except:
                pass
            conn.commit()
        finally:
            conn.close()

    def save_entity(self, entity_data):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            e_type = entity_data.get('type', 'account')
            table = 'clients' if e_type == 'account' else 'suppliers'
            action = entity_data.get('action', 'add')
            name = entity_data.get('name', '')
            phone = entity_data.get('phone', '')
            address = entity_data.get('address', '')
            email = entity_data.get('email', '')
            activity = entity_data.get('activity', '')
            rc = entity_data.get('rc', '')
            nif = entity_data.get('nif', '')
            nis = entity_data.get('nis', '')
            nai = entity_data.get('nai', '')
            p_cat = entity_data.get('price_category', '')
            gps = entity_data.get('gps_location', '')
            if action == 'add':
                cursor.execute(f'\n                    INSERT INTO {table} \n                    (name, phone, balance, address, email, activity, rc, nif, nis, nai, price_category, gps_location) \n                    VALUES (?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?)\n                ', (name, phone, address, email, activity, rc, nif, nis, nai, p_cat, gps))
            else:
                e_id = entity_data.get('id')
                cursor.execute(f'\n                    UPDATE {table} \n                    SET name=?, phone=?, address=?, email=?, activity=?, rc=?, nif=?, nis=?, nai=?, price_category=?, gps_location=? \n                    WHERE id=?\n                ', (name, phone, address, email, activity, rc, nif, nis, nai, p_cat, gps, e_id))
            conn.commit()
        finally:
            conn.close()

    def clean_up_wal(self):
        try:
            if self.db_name and os.path.exists(self.db_name):
                try:
                    temp_conn = sqlite3.connect(self.db_name)
                    temp_conn.execute('PRAGMA wal_checkpoint(TRUNCATE);')
                    temp_conn.execute('PRAGMA journal_mode=DELETE;')
                    temp_conn.commit()
                    temp_conn.close()
                except Exception as e:
                    print(f'SQL Cleanup Warning: {e}')
            if platform != 'android':
                time.sleep(0.1)
            wal_path = self.db_name + '-wal'
            shm_path = self.db_name + '-shm'
            if os.path.exists(wal_path):
                try:
                    os.remove(wal_path)
                except:
                    pass
            if os.path.exists(shm_path):
                try:
                    os.remove(shm_path)
                except:
                    pass
        except Exception as e:
            print(f'General Cleanup error: {e}')

    def get_connection(self):
        db_dir = os.path.dirname(self.db_name)
        if db_dir and (not os.path.exists(db_dir)):
            try:
                os.makedirs(db_dir)
            except:
                pass
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = sqlite3.Row
        try:
            conn.execute('PRAGMA journal_mode=WAL;')
            conn.execute('PRAGMA synchronous=NORMAL;')
            conn.execute('PRAGMA foreign_keys=ON;')
        except Exception as e:
            pass
        return conn

    def connect(self):
        self.conn = sqlite3.connect(self.db_name)
        self.conn.row_factory = sqlite3.Row
        try:
            self.conn.execute('PRAGMA journal_mode=WAL;')
            self.conn.execute('PRAGMA synchronous=NORMAL;')
            self.conn.execute('PRAGMA foreign_keys=ON;')
        except Exception as e:
            print(f'Warning: DB connection optimization failed: {e}')

    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None

    def create_tables(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS users (\n                    id INTEGER PRIMARY KEY AUTOINCREMENT,\n                    username TEXT UNIQUE,\n                    password_hash TEXT,\n                    role TEXT\n                )\n            ')
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS product_families (\n                    id INTEGER PRIMARY KEY AUTOINCREMENT,\n                    name TEXT UNIQUE\n                )\n            ')
            cursor.execute("INSERT OR IGNORE INTO product_families (name) VALUES ('TOUS')")
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS products (\n                    id INTEGER PRIMARY KEY AUTOINCREMENT,\n                    name TEXT,\n                    barcode TEXT,\n                    price REAL,\n                    stock REAL,\n                    category TEXT,\n                    family TEXT DEFAULT "TOUS",\n                    product_ref TEXT,\n                    image_path TEXT,\n                    reference TEXT,\n                    purchase_price REAL DEFAULT 0,\n                    price_semi REAL DEFAULT 0,\n                    price_wholesale REAL DEFAULT 0,\n                    is_used INTEGER DEFAULT 0,\n                    stock_warehouse REAL DEFAULT 0,\n                    is_promo_active INTEGER DEFAULT 0,\n                    promo_type TEXT DEFAULT "fixed",\n                    promo_value REAL DEFAULT 0,\n                    promo_qty_limit REAL DEFAULT 0,\n                    promo_expiry TEXT DEFAULT ""\n                )\n            ')
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS suppliers (\n                    id INTEGER PRIMARY KEY AUTOINCREMENT,\n                    name TEXT,\n                    phone TEXT,\n                    balance REAL DEFAULT 0,\n                    address TEXT,\n                    email TEXT,\n                    activity TEXT,\n                    rc TEXT,\n                    nif TEXT,\n                    nis TEXT,\n                    nai TEXT,\n                    price_category TEXT\n                )\n            ')
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS clients (\n                    id INTEGER PRIMARY KEY AUTOINCREMENT,\n                    name TEXT,\n                    phone TEXT,\n                    balance REAL DEFAULT 0,\n                    address TEXT,\n                    email TEXT,\n                    activity TEXT,\n                    rc TEXT,\n                    nif TEXT,\n                    nis TEXT,\n                    nai TEXT,\n                    price_category TEXT\n                )\n            ')
            cursor.execute("\n                CREATE TABLE IF NOT EXISTS transactions (\n                    id INTEGER PRIMARY KEY AUTOINCREMENT,\n                    transaction_type TEXT,\n                    entity_category TEXT, \n                    client_name TEXT,\n                    total_amount REAL,\n                    discount REAL DEFAULT 0,\n                    date TIMESTAMP,\n                    entity_id INTEGER,\n                    custom_label TEXT,\n                    user_name TEXT,\n                    note TEXT,\n                    payment_details TEXT,\n                    location TEXT DEFAULT 'store' \n                )\n            ")
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS transaction_items (\n                    id INTEGER PRIMARY KEY AUTOINCREMENT,\n                    transaction_id INTEGER,\n                    product_id INTEGER,\n                    product_name TEXT,\n                    qty REAL,\n                    price REAL,\n                    cost_price REAL DEFAULT 0,\n                    tva REAL DEFAULT 0,\n                    is_return INTEGER DEFAULT 0,\n                    FOREIGN KEY(transaction_id) REFERENCES transactions(id)\n                )\n            ')
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS app_settings (\n                    key TEXT PRIMARY KEY,\n                    value TEXT\n                )\n            ')
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS local_stats (\n                    date TEXT PRIMARY KEY,\n                    sales REAL DEFAULT 0,\n                    purchases REAL DEFAULT 0,\n                    c_pay REAL DEFAULT 0,\n                    s_pay REAL DEFAULT 0\n                )\n            ')
            cursor.execute('\n                CREATE TABLE IF NOT EXISTS document_sequences (\n                    name TEXT PRIMARY KEY, \n                    current_value INTEGER DEFAULT 0\n                )\n            ')
            conn.commit()
        finally:
            conn.close()

    def get_active_entity_ids_today(self, entity_category='client'):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            today_str = datetime.now().strftime('%Y-%m-%d')
            query = "\n                SELECT DISTINCT entity_id \n                FROM transactions \n                WHERE date(date) = date(?) \n                AND entity_category = ?\n                AND transaction_type NOT IN ('CLIENT_PAY', 'SUPPLIER_PAY', 'VERSEMENT', 'REGLEMENT')\n            "
            cat_val = 'client' if entity_category == 'account' else 'supplier'
            cursor.execute(query, (today_str, cat_val))
            rows = cursor.fetchall()
            return [row[0] for row in rows if row[0] is not None]
        finally:
            conn.close()

    def get_comprehensive_stats(self, target_date=None):
        if target_date:
            date_str = str(target_date)
        else:
            date_str = datetime.now().strftime('%Y-%m-%d')
        conn = self.get_connection()
        stats = {'sales': Decimal('0.00'), 'purchases': Decimal('0.00'), 'profit': Decimal('0.00'), 'cash_in': Decimal('0.00'), 'cash_out': Decimal('0.00'), 'stock_value': Decimal('0.00')}
        try:
            cursor = conn.cursor()
            cursor.execute('\n                SELECT transaction_type, total_amount, payment_details \n                FROM transactions \n                WHERE date(date) = date(?)\n            ', (date_str,))
            rows = cursor.fetchall()
            for row in rows:
                t_type = str(row[0]).upper()
                total = to_decimal(row[1])
                paid = Decimal('0.00')
                try:
                    pd = json.loads(row[2]) if row[2] else {}
                    paid = to_decimal(pd.get('amount', 0))
                except:
                    paid = Decimal('0.00')
                s_factor = AppConstants.STOCK_MOVEMENTS.get(t_type, 0)
                f_factor = AppConstants.FINANCIAL_FACTORS.get(t_type, 0)
                if s_factor == -1 and f_factor == 1:
                    stats['sales'] += total
                    if paid > 0:
                        stats['cash_in'] += paid
                elif s_factor == 1 and f_factor == -1:
                    stats['sales'] -= total
                    if paid > 0:
                        stats['cash_out'] += paid
                elif s_factor == 1 and f_factor == 1:
                    stats['purchases'] += total
                    if paid > 0:
                        stats['cash_out'] += paid
                elif s_factor == -1 and f_factor == -1:
                    stats['purchases'] -= total
                    if paid > 0:
                        stats['cash_in'] += paid
                elif t_type in ['CLIENT_PAY', 'VERSEMENT']:
                    stats['cash_in'] += total
                elif t_type in ['SUPPLIER_PAY', 'REGLEMENT']:
                    stats['cash_out'] += total
            cursor.execute('\n                SELECT t.transaction_type, ti.qty, ti.price, ti.cost_price\n                FROM transaction_items ti\n                JOIN transactions t ON ti.transaction_id = t.id\n                WHERE date(t.date) = date(?)\n            ', (date_str,))
            item_rows = cursor.fetchall()
            for i_row in item_rows:
                t_type = str(i_row[0]).upper()
                qty = to_decimal(i_row[1])
                sell_price = to_decimal(i_row[2])
                hist_cost = to_decimal(i_row[3])
                s_factor = AppConstants.STOCK_MOVEMENTS.get(t_type, 0)
                f_factor = AppConstants.FINANCIAL_FACTORS.get(t_type, 0)
                margin = (sell_price - hist_cost) * qty
                if s_factor == -1 and f_factor == 1:
                    stats['profit'] += margin
                elif s_factor == 1 and f_factor == -1:
                    stats['profit'] -= margin
            cursor.execute('SELECT stock, stock_warehouse, purchase_price FROM products')
            stock_rows = cursor.fetchall()
            for p in stock_rows:
                s_store = to_decimal(p[0])
                s_wh = to_decimal(p[1])
                cost = to_decimal(p[2])
                total_qty = s_store + s_wh
                if total_qty < 0:
                    total_qty = Decimal('0')
                stats['stock_value'] += total_qty * cost
        except Exception as e:
            print(f'Stats Calculation Error: {e}')
        finally:
            conn.close()
        return {k: float(quantize_decimal(v)) for k, v in stats.items()}

    def delete_family(self, name):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT 1 FROM products WHERE family = ? LIMIT 1', (name,))
            row = cursor.fetchone()
            if row:
                return False
            cursor.execute('DELETE FROM product_families WHERE name = ?', (name,))
            conn.commit()
            return True
        except Exception as e:
            print(f'Error deleting family: {e}')
            return False
        finally:
            conn.close()

    def get_products(self, limit=50, offset=0, search_query=None, family_filter=None):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            base_query = '\n                SELECT id, name, barcode, price, stock, category, family, \n                       product_ref, image_path, reference, \n                       purchase_price, price_semi, price_wholesale, \n                       is_used, stock_warehouse, \n                       is_promo_active, promo_type, promo_value, \n                       promo_qty_limit, promo_expiry \n                FROM products \n            '
            conditions = []
            params = []
            if search_query:
                if search_query.strip().lower() == 'promo':
                    conditions.append('is_promo_active = 1')
                else:
                    conditions.append('(name LIKE ? OR barcode LIKE ? OR product_ref LIKE ? OR reference LIKE ?)')
                    search_param = f'%{search_query}%'
                    params.extend([search_param, search_param, search_param, search_param])
            if family_filter and family_filter != 'TOUS':
                conditions.append('family = ?')
                params.append(family_filter)
            where_clause = ''
            if conditions:
                where_clause = ' WHERE ' + ' AND '.join(conditions)
            query = f'{base_query} {where_clause} ORDER BY name COLLATE NOCASE ASC LIMIT ? OFFSET ?'
            params.extend([limit, offset])
            cursor.execute(query, params)
            rows = cursor.fetchall()
            results = []
            for row in rows:
                results.append(dict(row))
            return results
        finally:
            conn.close()

    def get_families(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT name FROM product_families ORDER BY name ASC')
            rows = cursor.fetchall()
            return [row[0] for row in rows]
        finally:
            conn.close()

    def add_family(self, name):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('INSERT INTO product_families (name) VALUES (?)', (name,))
            conn.commit()
            return True
        except:
            return False
        finally:
            conn.close()

    def get_next_sequence_value(self, seq_name):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT current_value FROM document_sequences WHERE name = ?', (seq_name,))
            row = cursor.fetchone()
            if row:
                next_val = row[0] + 1
                cursor.execute('UPDATE document_sequences SET current_value = ? WHERE name = ?', (next_val, seq_name))
            else:
                next_val = 1
                cursor.execute('INSERT INTO document_sequences (name, current_value) VALUES (?, ?)', (seq_name, next_val))
            conn.commit()
            return next_val
        finally:
            conn.close()

    def get_invoice_number(self, doc_type):
        is_yearly = doc_type in AppConstants.YEARLY_RESET_SEQUENCES
        year = datetime.now().year
        if is_yearly:
            seq_name = f'SEQ_{doc_type}_{year}'
            date_part = str(year)
        else:
            seq_name = f'SEQ_{doc_type}'
            date_part = datetime.now().strftime('%d%m')
        next_val = self.get_next_sequence_value(seq_name)
        return f'{doc_type}{next_val:05d}/{date_part}'

    def check_product_has_movements(self, product_id):
        if not product_id:
            return False
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            query = "\n                SELECT COUNT(*) \n                FROM transaction_items ti \n                JOIN transactions t ON ti.transaction_id = t.id \n                WHERE ti.product_id = ? AND t.transaction_type != 'BI'\n            "
            cursor.execute(query, (product_id,))
            count = cursor.fetchone()[0]
            return count > 0
        finally:
            conn.close()

    def get_product_bi_transaction(self, product_id):
        if not product_id:
            return None
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            query = "\n                SELECT t.id \n                FROM transaction_items ti \n                JOIN transactions t ON ti.transaction_id = t.id \n                WHERE ti.product_id = ? AND t.transaction_type = 'BI' \n                LIMIT 1\n            "
            cursor.execute(query, (product_id,))
            row = cursor.fetchone()
            return row[0] if row else None
        finally:
            conn.close()

    def update_bi_transaction_qty(self, trans_id, product_id, new_qty, new_price):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('\n                UPDATE transaction_items \n                SET qty = ?, price = ?, cost_price = ? \n                WHERE transaction_id = ? AND product_id = ?\n            ', (new_qty, new_price, new_price, trans_id, product_id))
            new_total = new_qty * new_price
            cursor.execute('\n                UPDATE transactions \n                SET total_amount = ? \n                WHERE id = ?\n            ', (new_total, trans_id))
            conn.commit()
        finally:
            conn.close()

    def save_product(self, product_data):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            action = product_data.get('action', 'add')
            name = product_data.get('name', '')
            barcode = product_data.get('barcode', '')
            ref_val = product_data.get('reference', '')
            category = product_data.get('category', 'General')
            family = product_data.get('family', 'Générale')
            price = float(product_data.get('price', 0))
            stock = float(product_data.get('stock', 0))
            p_semi = float(product_data.get('price_semi', 0))
            p_whole = float(product_data.get('price_wholesale', 0))
            is_used = 1 if product_data.get('is_used', False) else 0
            product_ref = product_data.get('product_ref', '')
            p_price = float(product_data.get('purchase_price', 0))
            image_path = product_data.get('image_path', '')
            is_promo = 1 if product_data.get('is_promo_active', False) else 0
            promo_type = product_data.get('promo_type', 'fixed')
            try:
                promo_val = float(product_data.get('promo_value', 0))
            except:
                promo_val = 0.0
            try:
                promo_limit = float(product_data.get('promo_qty_limit', 0))
            except:
                promo_limit = 0.0
            promo_exp = str(product_data.get('promo_expiry', ''))
            p_id = product_data.get('id')
            if action == 'add':
                cursor.execute('\n                    INSERT INTO products \n                    (name, barcode, price, stock, category, family, product_ref, reference, purchase_price, price_semi, price_wholesale, is_used, image_path, is_promo_active, promo_type, promo_value, promo_qty_limit, promo_expiry) \n                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)\n                ', (name, barcode, price, stock, category, family, product_ref, ref_val, p_price, p_semi, p_whole, is_used, image_path, is_promo, promo_type, promo_val, promo_limit, promo_exp))
                p_id = cursor.lastrowid
            else:
                cursor.execute('\n                    UPDATE products \n                    SET name=?, barcode=?, price=?, stock=?, category=?, family=?, product_ref=?, reference=?, purchase_price=?, price_semi=?, price_wholesale=?, is_used=?, image_path=?, is_promo_active=?, promo_type=?, promo_value=?, promo_qty_limit=?, promo_expiry=? \n                    WHERE id=?\n                ', (name, barcode, price, stock, category, family, product_ref, ref_val, p_price, p_semi, p_whole, is_used, image_path, is_promo, promo_type, promo_val, promo_limit, promo_exp, p_id))
            conn.commit()
            return p_id
        finally:
            conn.close()

    def delete_product(self, product_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM products WHERE id=?', (product_id,))
            conn.commit()
        finally:
            conn.close()

    def seed_data(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT count(*) FROM users')
            if cursor.fetchone()[0] == 0:
                pw_hash = hashlib.sha256(''.encode()).hexdigest()
                cursor.execute('INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)', ('ADMIN', pw_hash, 'admin'))
            cursor.execute('SELECT count(*) FROM clients WHERE name=?', (AppConstants.DEFAULT_CLIENT_NAME,))
            if cursor.fetchone()[0] == 0:
                cursor.execute("INSERT INTO clients (name, price_category, balance) VALUES (?, 'Détail', 0)", (AppConstants.DEFAULT_CLIENT_NAME,))
            cursor.execute('SELECT count(*) FROM suppliers WHERE name=?', (AppConstants.DEFAULT_SUPPLIER_NAME,))
            if cursor.fetchone()[0] == 0:
                cursor.execute("INSERT INTO suppliers (name, price_category, balance) VALUES (?, 'Gros', 0)", (AppConstants.DEFAULT_SUPPLIER_NAME,))
            conn.commit()
        finally:
            conn.close()

    def login(self, password):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            pwd_str = password if password else ''
            pw_hash = hashlib.sha256(pwd_str.encode()).hexdigest()
            cursor.execute('SELECT * FROM users WHERE username = ? AND password_hash = ?', ('ADMIN', pw_hash))
            user = cursor.fetchone()
            return dict(user) if user else None
        finally:
            conn.close()

    def update_admin_password(self, new_password):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            pwd_str = new_password if new_password else ''
            pw_hash = hashlib.sha256(pwd_str.encode()).hexdigest()
            cursor.execute("UPDATE users SET password_hash = ? WHERE username = 'ADMIN'", (pw_hash,))
            conn.commit()
            return True
        except Exception as e:
            print(f'Error updating password: {e}')
            return False
        finally:
            conn.close()

    def update_stock(self, product_id, quantity_change, location='store'):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            column = 'stock_warehouse' if location == 'warehouse' else 'stock'
            sql = f'UPDATE products SET {column} = {column} + ? WHERE id = ?'
            cursor.execute(sql, (quantity_change, product_id))
            conn.commit()
        finally:
            conn.close()

    def get_entities(self, entity_type='account', search_query=None, limit=50, offset=0, sort_by='name', active_ids=None):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            table = 'clients' if entity_type == 'account' else 'suppliers'
            sql = f'SELECT * FROM {table} WHERE 1=1'
            params = []
            if search_query:
                sql += ' AND (name LIKE ? OR phone LIKE ?)'
                params.extend([f'%{search_query}%', f'%{search_query}%'])
            if sort_by == 'balance':
                sql += ' ORDER BY balance DESC'
            elif sort_by == 'active' and active_ids:
                ids_str = ','.join(map(str, active_ids))
                sql += f' ORDER BY CASE WHEN id IN ({ids_str}) THEN 0 ELSE 1 END, name COLLATE NOCASE ASC'
            else:
                sql += ' ORDER BY name COLLATE NOCASE ASC'
            sql += ' LIMIT ? OFFSET ?'
            params.append(limit)
            params.append(offset)
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

    def update_entity_balance(self, entity_id, amount, entity_type='account'):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            table = 'clients' if entity_type == 'account' else 'suppliers'
            cursor.execute(f'UPDATE {table} SET balance = balance + ? WHERE id = ?', (amount, entity_id))
            conn.commit()
        finally:
            conn.close()

    def delete_entity(self, entity_id, entity_type='account'):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            table = 'clients' if entity_type == 'account' else 'suppliers'
            cursor.execute(f'DELETE FROM {table} WHERE id = ?', (entity_id,))
            conn.commit()
        finally:
            conn.close()

    def save_transaction(self, transaction_data):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('BEGIN')
            t_id = transaction_data.get('id')
            t_type = transaction_data.get('doc_type', 'SALE').upper()
            stock_factor = AppConstants.STOCK_MOVEMENTS.get(t_type, 0)
            fin_factor = AppConstants.FINANCIAL_FACTORS.get(t_type, 0)
            is_supplier_op = stock_factor == 1 or 'SUPPLIER' in t_type or 'PURCHASE' in t_type
            entity_category_val = 'supplier' if is_supplier_op else 'client'
            user = transaction_data.get('user_name', '')
            total = float(transaction_data.get('amount', 0))
            items_list = transaction_data.get('items', [])
            date_str = transaction_data.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            ent_id = transaction_data.get('entity_id')
            note = transaction_data.get('note', '')
            new_location = transaction_data.get('purchase_location', 'store')
            is_simple = transaction_data.get('is_simple_payment', False)
            payment_info = transaction_data.get('payment_info', {})
            if is_simple:
                final_ref = transaction_data.get('custom_label', '')
                paid_amount = total
            else:
                if t_id:
                    cursor.execute('SELECT custom_label FROM transactions WHERE id=?', (t_id,))
                    row = cursor.fetchone()
                    final_ref = row[0] if row else self.get_invoice_number(t_type)
                else:
                    final_ref = self.get_invoice_number(t_type)
                paid_amount = float(payment_info.get('amount', 0)) if payment_info else 0.0
            payment_json = json.dumps(payment_info, ensure_ascii=False)
            if t_id:
                cursor.execute('SELECT * FROM transactions WHERE id=?', (t_id,))
                old_trans = dict(cursor.fetchone())
                old_type = old_trans.get('transaction_type')
                old_loc = old_trans.get('location', 'store')
                old_ent_id = old_trans.get('entity_id')
                old_total = float(old_trans.get('total_amount', 0))
                try:
                    old_payment_json = json.loads(old_trans.get('payment_details', '{}'))
                    old_paid = float(old_payment_json.get('amount', 0))
                except:
                    old_paid = 0.0
                if AppConstants.FINANCIAL_FACTORS.get(old_type) == -1 and (not items_list):
                    old_paid = old_total
                old_s_factor = AppConstants.STOCK_MOVEMENTS.get(old_type, 0)
                old_f_factor = AppConstants.FINANCIAL_FACTORS.get(old_type, 0)
                cursor.execute('SELECT product_id, qty FROM transaction_items WHERE transaction_id=?', (t_id,))
                old_items = cursor.fetchall()
                for item in old_items:
                    p_id = item['product_id']
                    if not p_id or p_id == -999:
                        continue
                    qty = float(item['qty'])
                    if old_type in ['TR', 'TRANSFER']:
                        col_src = 'stock_warehouse' if old_loc == 'warehouse' else 'stock'
                        col_dst = 'stock' if old_loc == 'warehouse' else 'stock_warehouse'
                        cursor.execute(f'UPDATE products SET {col_src} = ROUND({col_src} + ?, 3), {col_dst} = ROUND({col_dst} - ?, 3) WHERE id = ?', (qty, qty, p_id))
                    elif old_s_factor != 0:
                        revert_qty = qty * (old_s_factor * -1)
                        col = 'stock_warehouse' if old_loc == 'warehouse' else 'stock'
                        cursor.execute(f'UPDATE products SET {col} = ROUND({col} + ?, 3) WHERE id = ?', (revert_qty, p_id))
                if old_ent_id:
                    bal_revert = 0.0
                    prev_impact = old_total * old_f_factor - old_paid
                    bal_revert = -1 * prev_impact
                    if old_f_factor == -1 and old_s_factor == 0:
                        bal_revert = old_total
                    if bal_revert != 0:
                        cursor.execute('SELECT id FROM clients WHERE id=?', (old_ent_id,))
                        target_tbl = 'clients' if cursor.fetchone() else 'suppliers'
                        cursor.execute(f'UPDATE {target_tbl} SET balance = ROUND(balance + ?, 2) WHERE id = ?', (bal_revert, old_ent_id))
                cursor.execute('DELETE FROM transaction_items WHERE transaction_id=?', (t_id,))
                cursor.execute('\n                    UPDATE transactions \n                    SET total_amount=?, entity_id=?, entity_category=?, user_name=?, note=?, payment_details=?, location=?, transaction_type=?\n                    WHERE id=?\n                ', (total, ent_id, entity_category_val, user, note, payment_json, new_location, t_type, t_id))
            else:
                cursor.execute('\n                    INSERT INTO transactions \n                    (transaction_type, entity_category, total_amount, date, entity_id, custom_label, user_name, note, payment_details, location) \n                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)\n                ', (t_type, entity_category_val, total, date_str, ent_id, final_ref, user, note, payment_json, new_location))
                t_id = cursor.lastrowid
            for item in items_list:
                p_id = item.get('id')
                qty = float(item.get('qty', 0))
                price = float(item.get('price', 0))
                tva = float(item.get('tva', 0))
                is_virtual = p_id == -999 or item.get('is_virtual') or (not p_id)
                current_cost = 0.0
                if not is_virtual:
                    cursor.execute('SELECT purchase_price FROM products WHERE id=?', (p_id,))
                    res = cursor.fetchone()
                    if res:
                        current_cost = float(res[0] or 0)
                if not is_virtual:
                    if t_type in ['TR', 'TRANSFER']:
                        col_src = 'stock_warehouse' if new_location == 'warehouse' else 'stock'
                        col_dst = 'stock' if new_location == 'warehouse' else 'stock_warehouse'
                        cursor.execute(f'UPDATE products SET {col_src} = ROUND({col_src} - ?, 3) WHERE id = ?', (qty, p_id))
                        cursor.execute(f'UPDATE products SET {col_dst} = ROUND({col_dst} + ?, 3) WHERE id = ?', (qty, p_id))
                    elif stock_factor != 0:
                        change = qty * stock_factor
                        col = 'stock_warehouse' if new_location == 'warehouse' else 'stock'
                        cursor.execute(f'UPDATE products SET {col} = ROUND({col} + ?, 3) WHERE id = ?', (change, p_id))
                safe_pid = p_id if p_id and p_id != -999 else 0
                cursor.execute('\n                    INSERT INTO transaction_items \n                    (transaction_id, product_id, product_name, qty, price, tva, is_return, cost_price) \n                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)\n                ', (t_id, safe_pid, item.get('name', 'Product'), qty, price, tva, 1 if item.get('is_return') else 0, current_cost))
            if ent_id:
                balance_impact = 0.0
                if fin_factor == -1 and stock_factor == 0:
                    balance_impact = -total
                else:
                    balance_impact = total * fin_factor - paid_amount
                if balance_impact != 0:
                    target_tbl = 'suppliers' if entity_category_val == 'supplier' else 'clients'
                    cursor.execute(f'UPDATE {target_tbl} SET balance = ROUND(balance + ?, 2) WHERE id = ?', (balance_impact, ent_id))
                    if cursor.rowcount == 0:
                        alt_tbl = 'clients' if target_tbl == 'suppliers' else 'suppliers'
                        cursor.execute(f'UPDATE {alt_tbl} SET balance = ROUND(balance + ?, 2) WHERE id = ?', (balance_impact, ent_id))
            conn.commit()
            return t_id
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            if conn:
                conn.close()

    def delete_transaction(self, t_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM transactions WHERE id=?', (t_id,))
            trans = cursor.fetchone()
            if not trans:
                return
            trans = dict(trans)
            t_type = trans.get('transaction_type', '').upper()
            ent_id = trans.get('entity_id')
            total = to_decimal(trans.get('total_amount', 0))
            loc = trans.get('location', 'store')
            paid = Decimal('0.00')
            try:
                details = json.loads(trans.get('payment_details', '{}'))
                paid = to_decimal(details.get('amount', 0))
            except:
                paid = Decimal('0.00')
            if t_type in ['CLIENT_PAY', 'SUPPLIER_PAY']:
                paid = total
            s_factor = AppConstants.STOCK_MOVEMENTS.get(t_type, 0)
            f_factor = AppConstants.FINANCIAL_FACTORS.get(t_type, 0)
            cursor.execute('SELECT product_id, qty FROM transaction_items WHERE transaction_id=?', (t_id,))
            for item in cursor.fetchall():
                p_id = item['product_id']
                if not p_id:
                    continue
                qty = to_decimal(item['qty'])
                if t_type in ['TR', 'TRANSFER']:
                    col_src = 'stock_warehouse' if loc == 'warehouse' else 'stock'
                    col_dst = 'stock' if loc == 'warehouse' else 'stock_warehouse'
                    cursor.execute(f'UPDATE products SET {col_src} = {col_src} + ?, {col_dst} = {col_dst} - ? WHERE id = ?', (float(qty), float(qty), p_id))
                elif s_factor != 0:
                    revert = qty * Decimal(s_factor * -1)
                    col = 'stock_warehouse' if loc == 'warehouse' else 'stock'
                    cursor.execute(f'UPDATE products SET {col} = {col} + ? WHERE id = ?', (float(revert), p_id))
            if ent_id:
                bal_revert = Decimal('0.00')
                if t_type in ['CLIENT_PAY', 'SUPPLIER_PAY']:
                    bal_revert = total
                else:
                    original_impact = total * Decimal(f_factor) - paid
                    bal_revert = original_impact * Decimal('-1')
                bal_revert = quantize_decimal(bal_revert)
                if bal_revert != 0:
                    cursor.execute('SELECT id FROM clients WHERE id=?', (ent_id,))
                    tbl = 'clients' if cursor.fetchone() else 'suppliers'
                    cursor.execute(f'UPDATE {tbl} SET balance = balance + ? WHERE id = ?', (float(bal_revert), ent_id))
            cursor.execute('DELETE FROM transaction_items WHERE transaction_id=?', (t_id,))
            cursor.execute('DELETE FROM transactions WHERE id=?', (t_id,))
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f'Delete Error: {e}')
        finally:
            conn.close()

    def get_transactions(self, target_date=None, entity_id=None, entity_category=None, limit=50, offset=0):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            query = 'SELECT * FROM transactions WHERE 1=1'
            params = []
            if target_date:
                query += ' AND date(date) = date(?)'
                params.append(str(target_date))
            if entity_id is not None:
                query += ' AND entity_id = ?'
                params.append(entity_id)
            if entity_category:
                query += ' AND entity_category = ?'
                params.append(entity_category)
            query += ' ORDER BY id DESC LIMIT ? OFFSET ?'
            params.append(limit)
            params.append(offset)
            cursor.execute(query, params)
            rows = cursor.fetchall()
            result = []
            for row in rows:
                t_data = dict(row)
                t_data['items'] = []
                result.append(t_data)
            return result
        finally:
            conn.close()

    def get_transaction_full_details(self, t_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM transactions WHERE id = ?', (t_id,))
            trans_row = cursor.fetchone()
            if not trans_row:
                return None
            trans = dict(trans_row)
            entity_id = trans.get('entity_id')
            entity = {}
            if entity_id:
                cursor.execute('SELECT * FROM clients WHERE id = ?', (entity_id,))
                ent_row = cursor.fetchone()
                if not ent_row:
                    cursor.execute('SELECT * FROM suppliers WHERE id = ?', (entity_id,))
                    ent_row = cursor.fetchone()
                if ent_row:
                    entity = dict(ent_row)
            cursor.execute('\n                SELECT ti.*, ti.product_name as name, p.product_ref, p.reference\n                FROM transaction_items ti\n                LEFT JOIN products p ON ti.product_id = p.id\n                WHERE ti.transaction_id = ?\n            ', (t_id,))
            items_rows = cursor.fetchall()
            items = [dict(r) for r in items_rows]
            return {'transaction': trans, 'entity': entity, 'items': items}
        finally:
            conn.close()

    def set_setting(self, key, value):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('REPLACE INTO app_settings (key, value) VALUES (?, ?)', (str(key), str(value)))
            conn.commit()
        finally:
            conn.close()

    def get_setting(self, key, default=None):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT value FROM app_settings WHERE key = ?', (str(key),))
            row = cursor.fetchone()
            return row['value'] if row else default
        finally:
            conn.close()

    def get_store_info(self):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT key, value FROM app_settings WHERE key LIKE 'store_%'")
            rows = cursor.fetchall()
            return {row['key']: row['value'] for row in rows}
        finally:
            conn.close()

    def setting_exists(self, key):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT 1 FROM app_settings WHERE key = ?', (str(key),))
            return cursor.fetchone() is not None
        finally:
            conn.close()

    def save_stats_data(self, date_str, sales, purchases, c_pay, s_pay):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('\n                REPLACE INTO local_stats (date, sales, purchases, c_pay, s_pay) \n                VALUES (?, ?, ?, ?, ?)\n            ', (date_str, sales, purchases, c_pay, s_pay))
            conn.commit()
        finally:
            conn.close()

    def get_stats_data(self, date_str):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM local_stats WHERE date = ?', (date_str,))
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

    def get_product_by_barcode(self, barcode):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM products WHERE barcode = ?', (barcode,))
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

    def get_product_by_id(self, p_id):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM products WHERE id = ?', (p_id,))
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

    def get_product_by_name(self, name):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM products WHERE name = ?', (name,))
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

    def get_entity_by_id(self, e_id, e_type='account'):
        conn = self.get_connection()
        try:
            cursor = conn.cursor()
            table = 'clients' if e_type == 'account' else 'suppliers'
            cursor.execute(f'SELECT * FROM {table} WHERE id = ?', (e_id,))
            row = cursor.fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

class CartRecycleItem(RecycleDataViewBehavior, MDCard):
    text_name = StringProperty('')
    text_details = StringProperty('')
    details_color = ListProperty([0.4, 0.4, 0.4, 1])
    raw_data = ObjectProperty(None)

    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        app = MDApp.get_running_app()
        self.text_name = app.fix_text(data.get('name', ''))
        self.text_details = data.get('details', '')
        self.details_color = data.get('d_color', [0.4, 0.4, 0.4, 1])
        self.raw_data = data.get('raw_item')
        return super().refresh_view_attrs(rv, index, data)

    def on_tap(self):
        MDApp.get_running_app().edit_cart_item(self.raw_data)

    def on_delete(self):
        MDApp.get_running_app().remove_from_cart(self.raw_data)

class CartRecycleView(RecycleView):

    def __init__(self, **kwargs):
        super(CartRecycleView, self).__init__(**kwargs)
        self.data = []

class NoMenuTextField(MDTextField):

    def _show_cut_copy_paste(self, pos, selection, mode=None):
        pass

    def on_double_tap(self):
        pass

class SmartTextField(MDTextField):

    def __init__(self, **kwargs):
        self._raw_text = kwargs.get('text', '')
        self._programmatic_change = False
        self.base_direction = 'ltr'
        self.halign = 'left'
        self._input_reshaper = arabic_reshaper.ArabicReshaper(configuration={'delete_harakat': True, 'support_ligatures': False, 'use_unshaped_instead_of_isolated': True})
        if self._raw_text:
            try:
                reshaped = self._input_reshaper.reshape(self._raw_text)
                kwargs['text'] = get_display(reshaped)
            except:
                pass
        super().__init__(**kwargs)
        self.font_name = 'ArabicFont'
        self.font_name_hint_text = 'ArabicFont'
        if self._raw_text:
            self._update_alignment(self._raw_text)
        self.bind(text=self.on_text_change)

    def on_text_change(self, instance, value):
        if self._programmatic_change:
            self._programmatic_change = False
            return
        if value and (not self._raw_text):
            self._raw_text = value
            self._update_alignment(value)
        elif value != get_display(self._input_reshaper.reshape(self._raw_text)):
            self._raw_text = value
            self._update_alignment(value)

    def insert_text(self, substring, from_undo=False):
        self._programmatic_change = True
        self._raw_text += substring
        reshaped = self._input_reshaper.reshape(self._raw_text)
        bidi_text = get_display(reshaped)
        self.text = bidi_text
        self._update_alignment(self._raw_text)

    def do_backspace(self, from_undo=False, mode='bkspc'):
        if not self._raw_text:
            return
        self._programmatic_change = True
        self._raw_text = self._raw_text[:-1]
        reshaped = self._input_reshaper.reshape(self._raw_text)
        bidi_text = get_display(reshaped)
        self.text = bidi_text
        self._update_alignment(self._raw_text)

    def _update_alignment(self, text):
        if not text:
            self.halign = 'left'
            self.base_direction = 'ltr'
            return
        has_arabic = any(('\u0600' <= c <= 'ۿ' for c in text))
        if has_arabic:
            self.halign = 'right'
            self.base_direction = 'rtl'
        else:
            self.halign = 'left'
            self.base_direction = 'ltr'

    def get_value(self):
        if not self._raw_text and self.text:
            return self.text
        return self._raw_text if self._raw_text is not None else ''

class LeftButtonsContainer(ILeftBody, MDBoxLayout):
    adaptive_width = True

class RightButtonsContainer(IRightBodyTouch, MDBoxLayout):
    adaptive_width = True

class ProductRecycleItem(RecycleDataViewBehavior, MDBoxLayout):
    index = None
    text_name = StringProperty('')
    text_price = StringProperty('')
    text_stock = StringProperty('')
    icon_name = StringProperty('package-variant')
    icon_color = ListProperty([0, 0, 0, 1])
    price_color = ListProperty([0, 0, 0, 1])
    image_path = StringProperty('')
    product_data = ObjectProperty(None)

    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        self.text_name = data.get('name', '')
        self.text_price = data.get('price_text', '')
        self.text_stock = data.get('stock_text', '')
        self.icon_name = data.get('icon', 'package-variant')
        self.icon_color = data.get('icon_color', [0, 0, 0, 1])
        self.price_color = data.get('price_color', [0, 0, 0, 1])
        self.image_path = data.get('image_path', '')
        self.product_data = data.get('raw_data')
        return super().refresh_view_attrs(rv, index, data)

    def on_tap(self):
        app = MDApp.get_running_app()
        if self.product_data:
            app.open_add_to_cart_dialog(self.product_data, app.current_mode)

    def on_zoom(self):
        if self.image_path:
            MDApp.get_running_app().show_zoomed_image(self.image_path, self.text_name)
        else:
            self.on_tap()

class HistoryRecycleItem(RecycleDataViewBehavior, MDCard):
    index = None
    text_primary = StringProperty('')
    text_secondary = StringProperty('')
    text_amount = StringProperty('')
    icon_name = StringProperty('file')
    icon_color = ColorProperty([0, 0, 0, 1])
    bg_color = ColorProperty([1, 1, 1, 1])
    item_data = ObjectProperty(None, allownone=True)
    key = StringProperty('')

    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        app = MDApp.get_running_app()
        self.text_primary = app.fix_text(data.get('raw_text', ''))
        self.text_secondary = app.fix_text(data.get('raw_sec', ''))
        self.text_amount = data.get('amount_text', '')
        self.icon_name = data.get('icon', 'file')
        self.icon_color = data.get('icon_color', [0, 0, 0, 1])
        self.bg_color = data.get('bg_color', [1, 1, 1, 1])
        self.item_data = data.get('raw_data')
        self.key = data.get('key', '')
        return super().refresh_view_attrs(rv, index, data)

    def on_tap(self):
        app = MDApp.get_running_app()
        if self.key:
            try:
                t_id = int(self.key)
                print(f'Opening transaction ID: {t_id}')
                app.view_local_transaction_details({'id': t_id})
            except Exception as e:
                print(f'Error opening history item: {e}')
                app.notify("Erreur lors de l'ouverture des détails", 'error')

class MgmtEntityRecycleItem(RecycleDataViewBehavior, MDCard):
    index = None
    text_name = StringProperty('')
    text_balance = StringProperty('')
    entity_data = ObjectProperty(None, allownone=True)
    _long_press_event = None
    _is_long_press = False

    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        app = MDApp.get_running_app()
        raw_name = data.get('raw_name', '')
        self.text_name = app.fix_text(raw_name) if app else raw_name
        self.text_balance = data.get('balance_text', '')
        self.entity_data = data.get('raw_data')
        self.md_bg_color = data.get('bg_color', (1, 1, 1, 1))
        return super().refresh_view_attrs(rv, index, data)

    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos):
            self._is_long_press = False
            self._long_press_event = Clock.schedule_once(lambda dt: self._trigger_long_press(), 0.35)
        return super().on_touch_down(touch)

    def on_touch_move(self, touch):
        if self._long_press_event:
            if abs(touch.dx) > 10 or abs(touch.dy) > 10:
                self._long_press_event.cancel()
                self._long_press_event = None
        return super().on_touch_move(touch)

    def on_touch_up(self, touch):
        if self._long_press_event:
            self._long_press_event.cancel()
            self._long_press_event = None
        if self._is_long_press:
            return True
        return super().on_touch_up(touch)

    def _trigger_long_press(self):
        self._is_long_press = True
        self.on_menu()

    def on_pay(self):
        if self._is_long_press:
            return
        app = MDApp.get_running_app()
        if self.entity_data:
            app.start_direct_payment_from_manager(self.entity_data)

    def on_menu(self):
        app = MDApp.get_running_app()
        if self.entity_data:
            app.open_entity_edit_menu(self.entity_data)

    def on_history(self):
        app = MDApp.get_running_app()
        if self.entity_data:
            app.open_entity_history_dialog(self.entity_data)

class EntityRecycleItem(RecycleDataViewBehavior, MDCard):
    index = None
    text_name = StringProperty('')
    text_balance = StringProperty('')
    icon_name = StringProperty('account')
    icon_color = ListProperty([0, 0, 0, 1])
    entity_data = ObjectProperty(None, allownone=True)

    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        app = MDApp.get_running_app()
        raw_name = data.get('raw_name', '')
        self.text_name = app.fix_text(raw_name) if app else raw_name
        self.text_balance = data.get('balance_text', '')
        self.icon_name = data.get('icon', 'account')
        self.icon_color = data.get('icon_color', [0, 0, 0, 1])
        self.entity_data = data.get('raw_data')
        return super().refresh_view_attrs(rv, index, data)

    def on_tap(self):
        app = MDApp.get_running_app()
        if self.entity_data:
            app.select_entity_from_rv(self.entity_data)

class HistoryRecycleView(RecycleView):

    def __init__(self, **kwargs):
        super(HistoryRecycleView, self).__init__(**kwargs)
        self.data = []
        self.loading_lock = False
        self.do_scroll_x = False
        self.do_scroll_y = True
        self.scroll_type = ['bars', 'content']
        self.bar_width = dp(4)

    def on_scroll_y(self, instance, value):
        if value <= 0.1 and (not self.loading_lock):
            app = MDApp.get_running_app()
            if app and (not app.is_loading_history) and (len(self.data) >= 50):
                self.loading_lock = True
                app.load_more_history()

class MgmtEntityRecycleView(RecycleView):

    def __init__(self, **kwargs):
        super(MgmtEntityRecycleView, self).__init__(**kwargs)
        self.data = []
        self.loading_lock = False

    def on_scroll_y(self, instance, value):
        if value <= 0.1 and (not self.loading_lock):
            app = MDApp.get_running_app()
            if app and (not app.is_loading_entities) and (len(self.data) >= 50):
                self.loading_lock = True
                app.load_more_entities(reset=False)

class EntityRecycleView(RecycleView):

    def __init__(self, **kwargs):
        super(EntityRecycleView, self).__init__(**kwargs)
        self.data = []
        self.loading_lock = False

    def on_scroll_y(self, instance, value):
        if value <= 0.1 and (not self.loading_lock):
            app = MDApp.get_running_app()
            if app and (not app.is_loading_entities) and (len(self.data) >= 50):
                self.loading_lock = True
                app.load_more_entities(reset=False)

class ProductRecycleView(RecycleView):

    def __init__(self, **kwargs):
        super(ProductRecycleView, self).__init__(**kwargs)
        self.data = []
        self.loading_lock = False
        self.do_scroll_x = False
        self.do_scroll_y = True
        self.scroll_type = ['bars', 'content']
        self.bar_width = dp(4)

    def on_scroll_y(self, instance, value):
        if value <= 0.2 and (not self.loading_lock):
            app = MDApp.get_running_app()
            if app and (not app.is_loading_more) and (len(self.data) >= 50):
                self.loading_lock = True
                app.load_more_products(reset=False)

class StockApp(MDApp):
    cart = []
    current_mode = 'sale'
    current_user_name = 'ADMIN'
    is_seller_mode = BooleanProperty(False)
    selected_location = 'store'
    selected_entity = None
    editing_transaction_key = None
    editing_payment_amount = None
    dialog = None
    status_bar_label = None
    status_bar_bg = None
    rv_products = None
    _notify_event = None
    entity_list_layout = None
    history_list_layout = None
    pending_dialog = None
    action_dialog = None
    srv_dialog = None
    stat_sales_today = NumericProperty(0)
    stat_purchases_today = NumericProperty(0)
    stat_client_payments = NumericProperty(0)
    stat_supplier_payments = NumericProperty(0)
    stat_net_total = NumericProperty(0)
    buttons_container = None
    stats_container = None
    cart_list_layout = None
    lbl_cart_count = None
    lbl_cart_total = None
    lbl_total_title = None
    current_entity_type_mgmt = 'account'
    current_product_list_source = []
    current_page_offset = 0
    batch_size = 50
    is_loading_more = False
    history_page_offset = 0
    history_batch_size = 50
    is_loading_history = False
    history_view_date = None
    entity_page_offset = 0
    is_loading_entities = False
    active_entity_rv = None

    def fix_text(self, text):
        if not text or not isinstance(text, str):
            return str(text) if text is not None else ''
        if any(('\u0600' <= c <= 'ۿ' for c in text)):
            try:
                reshaped_text = reshaper.reshape(text)
                return get_display(reshaped_text)
            except Exception:
                return text
        return text

    def on_keyboard(self, window, key, scancode, codepoint, modifier):
        if key == 27:
            if hasattr(self, 'scan_dialog') and self.scan_dialog:
                self.close_barcode_scanner()
                return True
            all_dialogs = ['dialog', 'bt_dialog', 'zoom_dialog', 'img_picker_dialog', 'family_dialog', 'add_fam_dialog', 'confirm_del_fam_dialog', 'settings_menu_dialog', 'filter_dialog', 'mgmt_dialog', 'options_dialog', 'ae_dialog', 'cat_dialog', 'del_conf_dialog', 'pass_dialog', 'logout_diag', 'store_settings_dialog', 'entity_dialog', 'simple_pay_dialog', 'debt_dialog', 'overpay_dialog', 'pay_dialog', 'srv_dialog', 'pending_dialog', 'entity_hist_dialog', 'auth_dialog', 'toggle_dialog', 'mapping_diag', 'import_diag', 'restore_dialog', 'final_restore_confirm', 'activation_dialog_ref']
            for d_name in all_dialogs:
                if hasattr(self, d_name):
                    d_instance = getattr(self, d_name)
                    if d_instance and d_instance.parent:
                        d_instance.dismiss()
                        return True
            current_screen = self.sm.current
            if current_screen == 'cart':
                self.back_to_products()
                return True
            elif current_screen == 'products':
                self.go_back()
                return True
            elif current_screen == 'dashboard':
                self.logout()
                return True
            elif current_screen == 'login':
                return False
        return False

    def open_client_location(self, location_data):
        if not location_data:
            self.notify('Aucune position GPS enregistrée', 'error')
            return
        loc = str(location_data).strip()
        if not loc:
            return
        if 'http' in loc or 'waze' in loc or 'geo:' in loc:
            url = loc
        else:
            url = f'https://www.google.com/maps/search/?api=1&query={quote(loc)}'
        try:
            webbrowser.open(url)
            self.notify('Ouverture de Google Maps...', 'info')
        except Exception as e:
            self.notify(f'Erreur: {e}', 'error')

    def get_comprehensive_stats(self, target_date=None):
        return self.db.get_comprehensive_stats(target_date)

    def get_unified_path(self, filename):
        if platform == 'android':
            try:
                from jnius import autoclass
                Environment = autoclass('android.os.Environment')
                # محاولة الحصول على مسار التنزيلات العام
                path = Environment.getExternalStoragePublicDirectory(Environment.DIRECTORY_DOWNLOADS).getAbsolutePath()
                return os.path.join(path, filename)
            except:
                try:
                    # محاولة بديلة عبر Context
                    PythonActivity = autoclass('org.kivy.android.PythonActivity')
                    context = PythonActivity.mActivity
                    file_dir = context.getExternalFilesDir(Environment.DIRECTORY_DOWNLOADS)
                    return os.path.join(file_dir.getAbsolutePath(), filename)
                except:
                    # الملاذ الأخير: مجلد بيانات التطبيق
                    return os.path.join(self.user_data_dir, filename)
        else:
            # للكمبيوتر
            return os.path.join(os.path.expanduser('~'), 'Downloads', filename)

    def open_android_native_picker(self):
        try:
            from jnius import autoclass, cast
            from android import activity
            Intent = autoclass('android.content.Intent')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            intent = Intent(Intent.ACTION_GET_CONTENT)
            intent.setType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            try:
                mimeTypes = ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']
                intent.putExtra(Intent.EXTRA_MIME_TYPES, mimeTypes)
            except:
                pass
            intent.addCategory(Intent.CATEGORY_OPENABLE)
            activity.bind(on_activity_result=self._on_android_file_chosen)
            currentActivity = cast('android.app.Activity', PythonActivity.mActivity)
            currentActivity.startActivityForResult(intent, 101)
        except Exception as e:
            self.notify(f'Erreur lancement sélecteur: {e}', 'error')

    def _on_android_file_chosen(self, requestCode, resultCode, intent):
        from android import activity
        activity.unbind(on_activity_result=self._on_android_file_chosen)
        if requestCode == 101 and resultCode == -1:
            if intent:
                uri = intent.getData()
                if uri:
                    self._copy_uri_to_temp_and_process(uri)
                else:
                    self.notify('Aucun fichier sélectionné', 'error')
        else:
            self.notify('Sélection annulée', 'info')

    def _copy_uri_to_temp_and_process(self, uri):
        self.notify('Importation en cours...', 'info')
        threading.Thread(target=self._background_copy_task, args=(uri,), daemon=True).start()

    def _background_copy_task(self, uri):
        pfd = None
        try:
            from jnius import autoclass, cast
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            currentActivity = cast('android.app.Activity', PythonActivity.mActivity)
            content_resolver = currentActivity.getContentResolver()
            pfd = content_resolver.openFileDescriptor(uri, 'r')
            fd = pfd.getFd()
            temp_file_path = os.path.join(self.user_data_dir, 'temp_import.xlsx')
            with open(temp_file_path, 'wb') as f_out:
                while True:
                    try:
                        chunk = os.read(fd, 64 * 1024)
                    except OSError:
                        break
                    if not chunk:
                        break
                    f_out.write(chunk)
            try:
                pfd.close()
            except:
                pass
            Clock.schedule_once(lambda dt: self._finish_import_process(temp_file_path), 0)
        except Exception as e:
            print(f'COPY ERROR: {e}')
            if pfd:
                try:
                    pfd.close()
                except:
                    pass
            Clock.schedule_once(lambda dt: self.notify(f'Erreur copie: {str(e)}', 'error'), 0)

    def _finish_import_process(self, file_path):
        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
            self.pre_process_import([file_path])
        else:
            self.notify('Erreur: Fichier vide ou illisible', 'error')

    def prepare_products_for_rv(self, products_list):
        self.current_product_list_source = products_list
        self.current_page_offset = 0
        self.is_loading_more = False
        if self.rv_products:
            if not self.rv_products.data:
                self.rv_products.refresh_from_data()
        self.load_more_products(reset=True)

    def load_more_entities(self, reset=False):
        if self.is_loading_entities and (not reset):
            return
        is_mgmt = hasattr(self, 'mgmt_dialog') and self.mgmt_dialog and self.mgmt_dialog.parent
        is_selection = hasattr(self, 'entity_dialog') and self.entity_dialog and self.entity_dialog.parent
        target_type = 'account'
        search_q = ''
        current_sort = getattr(self, 'current_entity_sort', 'name')
        if is_mgmt:
            target_type = self.current_entity_type_mgmt
            self.active_entity_rv = self.rv_mgmt_entity
            if hasattr(self, 'entity_search') and self.entity_search.text:
                search_q = self.entity_search.text
        elif is_selection:
            target_type = getattr(self, 'entities_source_type', 'account')
            self.active_entity_rv = self.rv_entity
            if hasattr(self, 'entity_search') and self.entity_search.text:
                search_q = self.entity_search.text
        else:
            return
        if reset:
            self.entity_page_offset = 0
            self.is_loading_entities = False
            if self.active_entity_rv:
                self.active_entity_rv.scroll_y = 1.0
                self.active_entity_rv.data = []
                self.active_entity_rv.refresh_from_data()
        self.is_loading_entities = True
        active_today_ids = self.db.get_active_entity_ids_today(target_type)
        limit = 50
        entities = self.db.get_entities(target_type, search_query=search_q, limit=limit, offset=self.entity_page_offset, sort_by=current_sort, active_ids=active_today_ids)
        rv_data = []
        is_client_mode = target_type == 'account'
        default_name = AppConstants.DEFAULT_CLIENT_NAME if is_client_mode else AppConstants.DEFAULT_SUPPLIER_NAME
        if is_selection and self.entity_page_offset == 0 and (not search_q) and (current_sort == 'name'):
            bal_markup = '[size=14sp][b][color=404040]-- Direct --[/color][/b][/size]'
            rv_data.append({'raw_name': default_name, 'balance_text': bal_markup, 'icon': 'store', 'icon_color': [0, 0.4, 0.7, 1], 'raw_data': {'id': None, 'name': default_name, 'price_category': 'Détail'}, 'bg_color': (1, 1, 1, 1)})
        for e in entities:
            name = e.get('name', '')
            if name == default_name or name == 'COMPTOIR':
                continue
            e_id = e.get('id')
            balance = float(e.get('balance', 0))
            bal_text = f'{balance:,.2f} DA'.replace(',', ' ').replace('.', ',')
            is_active_today = e_id in active_today_ids
            row_bg_color = (0.88, 1, 0.88, 1) if is_active_today else (1, 1, 1, 1)
            if is_mgmt:
                col_hex = 'D50000' if balance > 0 else '00C853'
                balance_markup = f'Solde: [color={col_hex}][b]{bal_text}[/b][/color]'
                rv_data.append({'raw_name': name, 'balance_text': balance_markup, 'raw_data': e, 'bg_color': row_bg_color})
            else:
                bal_color_hex = '00C853' if is_client_mode else 'D50000'
                balance_markup = f'Solde: [color={bal_color_hex}][b]{bal_text}[/b][/color]'
                if balance <= 0:
                    icon_name = 'account-check'
                    icon_col = [0, 0.7, 0, 1]
                else:
                    icon_name = 'account-alert'
                    icon_col = [0.9, 0, 0, 1]
                rv_data.append({'raw_name': name, 'balance_text': balance_markup, 'icon': icon_name, 'icon_color': icon_col, 'raw_data': e, 'bg_color': (1, 1, 1, 1)})
        Clock.schedule_once(lambda dt: self._append_entities_to_rv(rv_data, reset, len(entities)))

    def open_family_selector_dialog(self):
        families = self.db.get_families()
        if 'TOUS' in families:
            families.remove('TOUS')
        full_list = ['TOUS'] + families
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, adaptive_height=True, padding=dp(0))
        scroll = MDScrollView(size_hint_y=None, height=dp(300))
        list_layout = MDList()
        for fam in full_list:
            item = OneLineAvatarIconListItem(text=fam, on_release=lambda x, f=fam: self.select_family_and_close(f))
            if fam != 'TOUS':
                del_icon = IconRightWidget(icon='trash-can-outline', theme_text_color='Custom', text_color=(0.8, 0, 0, 1))
                del_icon.bind(on_release=lambda x, f=fam: self.show_delete_family_confirmation(f))
                item.add_widget(del_icon)
            list_layout.add_widget(item)
        scroll.add_widget(list_layout)
        content.add_widget(scroll)
        self.family_dialog = MDDialog(title='Sélectionner une famille', type='custom', content_cls=content, size_hint=(0.85, None))
        self.family_dialog.open()

    def confirm_delete_family(self, family_name):
        if self.db.delete_family(family_name):
            self.notify(f'Famille "{family_name}" supprimée', 'success')
            self.family_dialog.dismiss()
            self.open_family_selector_dialog()
            if self.btn_select_family.text == family_name:
                self.btn_select_family.text = ''
        else:
            self.notify(f'Impossible: Des produits sont liés à "{family_name}"', 'error')

    def select_family_and_close(self, family_name):
        self.btn_select_family.text = family_name
        self.family_dialog.dismiss()

    def show_add_family_dialog(self):
        content = MDBoxLayout(orientation='vertical', spacing=dp(10), size_hint_y=None, height=dp(100), padding=dp(10))
        field = SmartTextField(hint_text='Nom de la nouvelle famille')
        content.add_widget(field)

        def save(x):
            name = field.text.strip()
            if name:
                if self.db.add_family(name):
                    self.notify(f'Famille "{name}" ajoutée', 'success')
                    self.btn_select_family.text = name
                    self.add_fam_dialog.dismiss()
                else:
                    self.notify('Erreur: Existe déjà ?', 'error')
            else:
                self.notify('Nom vide', 'error')
        self.add_fam_dialog = MDDialog(title='Nouvelle Famille', type='custom', content_cls=content, buttons=[MDRaisedButton(text='AJOUTER', on_release=save), MDFlatButton(text='ANNULER', on_release=lambda x: self.add_fam_dialog.dismiss())])
        self.add_fam_dialog.open()

    def update_family_filter_ui(self):
        if not hasattr(self, 'family_filter_box'):
            return
        self.family_filter_box.clear_widgets()
        families = self.db.get_families()
        families.insert(0, 'TOUTES')
        for fam in families:
            is_selected = fam == self.selected_family_filter
            bg_color = (0.2, 0.2, 0.2, 1) if is_selected else (0.9, 0.9, 0.9, 1)
            text_color = (1, 1, 1, 1) if is_selected else (0, 0, 0, 1)
            btn = MDFillRoundFlatButton(text=fam, md_bg_color=bg_color, text_color=text_color, font_size='13sp')
            btn.bind(on_release=lambda x, f=fam: self.on_family_selected(f))
            self.family_filter_box.add_widget(btn)

    def on_family_selected(self, family_name):
        self.selected_family_filter = family_name
        self.update_family_filter_ui()
        self.load_more_products(reset=True)

    def show_delete_family_confirmation(self, family_name):

        def confirm_delete(x):
            self.confirm_del_fam_dialog.dismiss()
            self.perform_delete_family(family_name)
        self.confirm_del_fam_dialog = MDDialog(title='Confirmation', text=f"Voulez-vous vraiment supprimer la famille '{family_name}' ?", buttons=[MDFlatButton(text='ANNULER', on_release=lambda x: self.confirm_del_fam_dialog.dismiss()), MDRaisedButton(text='OUI, SUPPRIMER', md_bg_color=(0.8, 0, 0, 1), on_release=confirm_delete)])
        self.confirm_del_fam_dialog.open()

    def perform_delete_family(self, family_name):
        is_deleted = self.db.delete_family(family_name)
        if is_deleted:
            self.notify(f'Famille "{family_name}" supprimée', 'success')
            self.play_sound('success')
            if hasattr(self, 'family_dialog') and self.family_dialog:
                self.family_dialog.dismiss()
            if hasattr(self, 'btn_select_family') and self.btn_select_family.text == family_name:
                self.btn_select_family.text = 'TOUS'
            self.open_family_selector_dialog()
        else:
            self.play_sound('error')
            self.notify(f"Impossible : La famille '{family_name}' contient des produits !", 'error')

    def filter_entities_paginated(self, instance, text=None):
        query = instance.get_value() if hasattr(instance, 'get_value') else text
        if self._entity_search_event:
            self._entity_search_event.cancel()
        self._entity_search_event = Clock.schedule_once(lambda dt: self.load_more_entities(reset=True), 0.5)

    def _append_entities_to_rv(self, new_data, reset, db_items_count):
        if self.active_entity_rv:
            if reset:
                self.active_entity_rv.data = new_data
                self.active_entity_rv.scroll_y = 1.0
                self.entity_page_offset = db_items_count
            else:
                self.active_entity_rv.data.extend(new_data)
                self.entity_page_offset += db_items_count
            self.active_entity_rv.refresh_from_data()
            self.active_entity_rv.loading_lock = False
        self.is_loading_entities = False

    def open_add_to_cart_dialog(self, product, mode):
        if mode == 'manage_products':
            self.show_manage_product_dialog(product)
            return

        def fmt_num(value):
            if not value:
                return '0'
            try:
                val_float = float(value)
                return str(int(val_float)) if val_float.is_integer() else str(val_float)
            except:
                return str(value)
        doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
        doc_type = doc_type_map.get(mode, 'BV')
        stock_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
        is_transfer = doc_type in ['TR', 'TRANSFER']
        curr_price = 0.0
        if stock_f == -1 or doc_type in ['FP', 'FC', 'RC']:
            base_price = float(product.get('price', 0) or 0)
            curr_price = base_price
            if self.selected_entity:
                cat = str(self.selected_entity.get('category', 'Détail')).strip()
                if cat == 'Gros':
                    val = float(product.get('price_wholesale', 0) or 0)
                    if val > 0:
                        curr_price = val
                elif cat == 'Demi-Gros':
                    val = float(product.get('price_semi', 0) or 0)
                    if val > 0:
                        curr_price = val
            is_promo_valid = False
            if product.get('is_promo_active', 0) == 1:
                promo_exp = str(product.get('promo_expiry', '')).strip()
                date_valid = True
                if promo_exp and len(promo_exp) > 5:
                    try:
                        from datetime import datetime
                        exp_date = datetime.strptime(promo_exp, '%Y-%m-%d').date()
                        if datetime.now().date() > exp_date:
                            date_valid = False
                    except:
                        pass
                if date_valid:
                    is_promo_valid = True
                    p_type = product.get('promo_type', 'fixed')
                    try:
                        p_val = float(product.get('promo_value', 0))
                    except:
                        p_val = 0.0
                    if p_type == 'fixed':
                        if p_val > 0:
                            curr_price = p_val
                    else:
                        curr_price = base_price * (1 - p_val / 100)
        else:
            curr_price = float(product.get('purchase_price', product.get('price', 0)) or 0)
        prod_name = self.fix_text(product.get('name'))
        price_val_str = fmt_num(curr_price)
        self.active_input_target = 'qty'
        self.input_reset_mode = True

        def update_field_colors():
            ACTIVE_BG = (0.9, 1, 0.9, 1)
            INACTIVE_BG = (0.95, 0.95, 0.95, 1)
            if hasattr(self, 'qty_card'):
                self.qty_card.md_bg_color = ACTIVE_BG if self.active_input_target == 'qty' else INACTIVE_BG
                self.qty_card.elevation = 3 if self.active_input_target == 'qty' else 0
            if hasattr(self, 'price_card'):
                self.price_card.md_bg_color = ACTIVE_BG if self.active_input_target == 'price' else INACTIVE_BG
                self.price_card.elevation = 3 if self.active_input_target == 'price' else 0
        header_box = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing='5dp', padding=[0, 0, 0, '5dp'])
        lbl_prod = MDLabel(text=prod_name, halign='center', bold=True, font_style='Subtitle1', theme_text_color='Primary', adaptive_height=True)
        header_box.add_widget(lbl_prod)
        dialog_height = dp(420) if is_transfer else dp(500)
        content = MDBoxLayout(orientation='vertical', spacing='8dp', size_hint_y=None, height=dialog_height, padding=[0, '5dp', 0, 0])
        content.add_widget(header_box)
        if not is_transfer:
            self.price_card = MDCard(size_hint_y=None, height='70dp', radius=[10], padding=[10, 0, 10, 0], elevation=0)
            self.price_field = NoMenuTextField(text=price_val_str, hint_text='Prix Unitaire (DA)', font_size='26sp', halign='center', mode='line', readonly=True, line_color_normal=(0, 0, 0, 0), line_color_focus=(0, 0, 0, 0), pos_hint={'center_y': 0.5})
            self.price_field.theme_text_color = 'Custom'
            self.price_field.text_color_normal = (0, 0, 0, 1)
            self.price_field.text_color_focus = (0, 0, 0, 1)
            if stock_f == -1 and 'is_promo_valid' in locals() and is_promo_valid:
                self.price_field.text_color_normal = (0.8, 0, 0, 1)

            def on_price_touch(instance, touch):
                if instance.collide_point(*touch.pos):
                    if self.active_input_target != 'price':
                        self.input_reset_mode = True
                    self.active_input_target = 'price'
                    update_field_colors()
                    return True
                return False
            self.price_field.bind(on_touch_down=on_price_touch)
            self.price_card.add_widget(self.price_field)
            price_row_container = MDBoxLayout(orientation='horizontal', size_hint_y=None, height='75dp', padding=[60, 0, 60, 0])
            price_row_container.add_widget(self.price_card)
            content.add_widget(price_row_container)
        qty_row = MDBoxLayout(orientation='horizontal', spacing='10dp', size_hint_y=None, height='65dp', padding=[40, 0])
        btn_minus = MDIconButton(icon='minus', theme_text_color='Custom', text_color=(1, 1, 1, 1), md_bg_color=(0.9, 0.3, 0.3, 1), pos_hint={'center_y': 0.5}, icon_size='20sp')
        self.qty_card = MDCard(size_hint_x=1, size_hint_y=None, height='60dp', radius=[10], padding=[10, 0, 10, 0], elevation=0, pos_hint={'center_y': 0.5})
        self.qty_field = NoMenuTextField(text='1', hint_text='Qté', font_size='28sp', halign='center', readonly=True, mode='line', line_color_normal=(0, 0, 0, 0), line_color_focus=(0, 0, 0, 0), pos_hint={'center_y': 0.5})
        self.qty_field.theme_text_color = 'Custom'
        self.qty_field.text_color_normal = (0, 0, 0, 1)
        self.qty_field.text_color_focus = (0, 0, 0, 1)
        self.qty_field.get_value = lambda: self.qty_field.text

        def on_qty_touch(instance, touch):
            if instance.collide_point(*touch.pos):
                if self.active_input_target != 'qty':
                    self.input_reset_mode = True
                self.active_input_target = 'qty'
                update_field_colors()
                return True
            return False
        self.qty_field.bind(on_touch_down=on_qty_touch)
        self.qty_card.add_widget(self.qty_field)
        btn_plus = MDIconButton(icon='plus', theme_text_color='Custom', text_color=(1, 1, 1, 1), md_bg_color=(0.2, 0.7, 0.2, 1), pos_hint={'center_y': 0.5}, icon_size='20sp')
        qty_row.add_widget(btn_minus)
        qty_row.add_widget(self.qty_card)
        qty_row.add_widget(btn_plus)
        content.add_widget(qty_row)
        self.btn_add = MDRaisedButton(text='AJOUTER', md_bg_color=(0, 0.7, 0, 1), text_color=(1, 1, 1, 1), size_hint_x=0.7, size_hint_y=1, font_size='18sp', elevation=3)
        temp_product = product.copy()
        if not is_transfer:
            temp_product['price'] = float(curr_price)

        def perform_add(x):
            try:
                if not is_transfer and hasattr(self, 'price_field'):
                    p_text = self.price_field.text.replace(',', '.')
                    if not p_text:
                        p_text = '0'
                    temp_product['price'] = float(p_text)
                q_text = self.qty_field.text.replace(',', '.')
                if not q_text:
                    q_text = '1'
                self.qty_field.text = q_text
                self.add_to_cart(temp_product)
                if self.dialog:
                    self.dialog.dismiss()
            except ValueError:
                self.notify('Valeurs invalides', 'error')
        self.btn_add.bind(on_release=perform_add)

        def update_button_text():
            if is_transfer:
                self.btn_add.text = 'AJOUTER'
                return
            try:
                q = float(self.qty_field.text.replace(',', '.') or 0)
            except:
                q = 1.0
            try:
                p = float(self.price_field.text.replace(',', '.') or 0)
            except:
                p = 0.0
            total_line = q * p
            self.btn_add.text = f'AJOUTER\n{total_line:.2f} DA'

        def increase(x):
            try:
                v = float(self.qty_field.text.replace(',', '.') or 0)
                self.qty_field.text = fmt_num(v + 1)
            except:
                self.qty_field.text = '1'
            update_button_text()

        def decrease(x):
            try:
                v = float(self.qty_field.text.replace(',', '.') or 0)
                if v > 1:
                    self.qty_field.text = fmt_num(v - 1)
            except:
                self.qty_field.text = '1'
            update_button_text()
        btn_plus.bind(on_release=increase)
        btn_minus.bind(on_release=decrease)

        def get_active_field():
            if is_transfer:
                return self.qty_field
            return self.price_field if self.active_input_target == 'price' else self.qty_field

        def add_digit(digit):
            field = get_active_field()
            current = field.text
            if self.input_reset_mode:
                if digit == '.':
                    field.text = '0.'
                else:
                    field.text = str(digit)
                self.input_reset_mode = False
            elif digit == '.':
                if '.' in current:
                    return
                if not current:
                    field.text = '0.'
                else:
                    field.text = current + '.'
            elif current == '0':
                field.text = str(digit)
            else:
                field.text = current + str(digit)
            update_button_text()

        def backspace(instance=None):
            field = get_active_field()
            current = field.text
            self.input_reset_mode = False
            if len(current) > 0:
                field.text = current[:-1]
            update_button_text()
        grid = MDGridLayout(cols=3, spacing='8dp', size_hint_y=1, padding=[20, 0])
        keys = ['7', '8', '9', '4', '5', '6', '1', '2', '3', '.', '0', 'DEL']
        for key in keys:
            if key == 'DEL':
                btn = MDIconButton(icon='backspace-outline', theme_text_color='Custom', text_color=(0, 0, 0, 1), md_bg_color=(0.8, 0.8, 0.8, 1), size_hint=(1, 1), icon_size='20sp', on_release=backspace)
            else:
                btn = MDRaisedButton(text=key, md_bg_color=(0.95, 0.95, 0.95, 1), theme_text_color='Custom', text_color=(0, 0, 0, 1), font_size='22sp', size_hint=(1, 1), elevation=1, on_release=lambda x, k=key: add_digit(k))
            grid.add_widget(btn)
        content.add_widget(grid)
        content.add_widget(MDLabel(text='', size_hint_y=None, height='10dp'))
        buttons_box = MDBoxLayout(orientation='horizontal', spacing='10dp', size_hint_y=None, height='60dp')
        btn_cancel = MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=(0.5, 0.5, 0.5, 1), size_hint_x=0.3, on_release=lambda x: self.dialog.dismiss())
        buttons_box.add_widget(btn_cancel)
        buttons_box.add_widget(self.btn_add)
        content.add_widget(buttons_box)
        update_field_colors()
        update_button_text()
        self.dialog = MDDialog(title='', type='custom', content_cls=content, buttons=[], size_hint=(0.85, None))
        self.dialog.open()

    @mainthread
    def _append_to_rv(self, new_data, reset=False):
        if self.rv_products:
            if reset:
                self.rv_products.data = new_data
                self.rv_products.scroll_y = 1.0
                self.current_page_offset = len(new_data)
            else:
                self.rv_products.data.extend(new_data)
                self.current_page_offset += len(new_data)
            self.rv_products.refresh_from_data()
            self.rv_products.loading_lock = False
        self.is_loading_more = False

    def filter_products(self, instance, text):
        query = instance.get_value() if hasattr(instance, 'get_value') else text
        if not hasattr(self, 'last_search_id'):
            self.last_search_id = 0
        self.last_search_id += 1
        current_search_id = self.last_search_id
        if self._search_event:
            self._search_event.cancel()
        self._search_event = Clock.schedule_once(lambda dt: self._start_background_search(query, current_search_id), 0.3)

    def _start_background_search(self, query, search_id):
        threading.Thread(target=self._search_worker, args=(query, search_id), daemon=True).start()

    def _search_worker(self, query, search_id):
        family = getattr(self, 'selected_family_filter', 'TOUS')
        if not query:
            results = self.db.get_products(limit=50, offset=0, family_filter=family)
        else:
            results = self.db.get_products(limit=100, offset=0, search_query=query, family_filter=family)
        if search_id == self.last_search_id:
            Clock.schedule_once(lambda dt: self._prepare_and_send_data(results), 0)
        else:
            print(f'[DEBUG] Discarding old search results for ID {search_id}')

    def _prepare_and_send_data(self, products_list):
        rv_data = []
        current_family = getattr(self, 'selected_family_filter', 'TOUS')
        allowed_autre_modes = ['sale', 'invoice_sale', 'proforma', 'order_purchase']
        if self.current_mode in allowed_autre_modes and current_family == 'TOUS':
            virtual_item = {'id': -999, 'name': 'Autre Article', 'price': 0, 'purchase_price': 0, 'stock': -1000000, 'stock_warehouse': 0, 'barcode': '', 'reference': '', 'is_virtual': True}
            if not products_list or products_list[0].get('id') != -999:
                products_list.insert(0, virtual_item)
        purchase_modes = ['purchase', 'invoice_purchase', 'return_purchase', 'order_purchase', 'bi']
        is_purchase_view = self.current_mode in purchase_modes
        is_transfer = self.current_mode == 'transfer'
        modes_showing_promo = ['sale', 'invoice_sale', 'proforma', 'client_payment', 'manage_products']
        should_show_promo = self.current_mode in modes_showing_promo
        customer_category = 'Détail'
        if self.selected_entity:
            customer_category = str(self.selected_entity.get('category', 'Détail')).strip()

        def fmt_qty(val):
            try:
                val = float(val)
                return str(int(val)) if val.is_integer() else str(val)
            except:
                return '0'
        img_dir = os.path.join(self.user_data_dir, 'product_images')
        try:
            from datetime import datetime
            for p in products_list:
                try:
                    s_store = float(p.get('stock', 0) or 0)
                    is_unlimited = s_store <= -900000
                    if p.get('id') == -999 and is_transfer:
                        continue
                    if is_transfer and is_unlimited:
                        continue
                    s_wh = float(p.get('stock_warehouse', 0) or 0)
                    total_stock = s_store + s_wh
                    price_fmt = ''
                    price_color = [0, 0, 0, 1]
                    stock_text = ''
                    if is_transfer:
                        price_fmt = f'Tot: {fmt_qty(total_stock)}'
                        price_color = [0.2, 0.2, 0.8, 1]
                        stock_text = f'Mag: {fmt_qty(s_store)} | Dép: {fmt_qty(s_wh)}'
                    elif is_purchase_view:
                        price = float(p.get('purchase_price', p.get('price', 0)) or 0)
                        price_fmt = f'{price:.2f} DA'
                        price_color = [0.9, 0.5, 0, 1]
                    else:
                        base_price = float(p.get('price', 0) or 0)
                        final_display_price = base_price
                        if customer_category == 'Gros':
                            wh_price = float(p.get('price_wholesale', 0) or 0)
                            if wh_price > 0:
                                final_display_price = wh_price
                        elif customer_category == 'Demi-Gros':
                            semi_price = float(p.get('price_semi', 0) or 0)
                            if semi_price > 0:
                                final_display_price = semi_price
                        has_promo = False
                        if should_show_promo:
                            raw_active = p.get('is_promo_active', 0)
                            is_active = str(raw_active) == '1' or raw_active == 1
                            if is_active:
                                promo_exp = str(p.get('promo_expiry', '')).strip()
                                date_valid = True
                                if promo_exp and len(promo_exp) > 5:
                                    try:
                                        exp_date = datetime.strptime(promo_exp, '%Y-%m-%d').date()
                                        if datetime.now().date() > exp_date:
                                            date_valid = False
                                    except:
                                        pass
                                if date_valid:
                                    has_promo = True
                                    p_type = str(p.get('promo_type', 'fixed'))
                                    try:
                                        p_val = float(p.get('promo_value', 0))
                                    except:
                                        p_val = 0.0
                                    if p_type == 'fixed':
                                        if p_val > 0:
                                            final_display_price = p_val
                                    else:
                                        final_display_price = base_price * (1 - p_val / 100)
                        price_fmt = f'{final_display_price:.2f} DA'
                        if has_promo:
                            price_color = [0, 0.2, 0.8, 1]
                            icon = 'sale'
                            icon_col = [0.9, 0.1, 0.1, 1]
                        else:
                            price_color = [0, 0.6, 0, 1]
                        if is_unlimited:
                            stock_text = 'Illimité'
                            if p.get('id') == -999:
                                price_fmt = 'Prix Libre'
                                price_color = [0, 0.4, 0.8, 1]
                                icon = 'pencil-plus'
                                icon_col = [0, 0.4, 0.8, 1]
                            elif not has_promo:
                                icon = 'package-variant'
                        elif s_wh == 0:
                            stock_text = f'Qté: {fmt_qty(s_store)}'
                            if not has_promo:
                                icon = 'package-variant' if s_store > 0 else 'package-variant-closed'
                                icon_col = [0, 0.6, 0, 1] if s_store > 0 else [0.8, 0, 0, 1]
                        else:
                            stock_text = f'Qté: {fmt_qty(s_store)} | Dép: {fmt_qty(s_wh)}'
                            if not has_promo:
                                icon = 'package-variant' if total_stock > 0 else 'package-variant-closed'
                                icon_col = [0, 0.6, 0, 1] if total_stock > 0 else [0.8, 0, 0, 1]
                    icon = icon if 'icon' in locals() else 'package-variant'
                    icon_col = icon_col if 'icon_col' in locals() else [0, 0.6, 0, 1]
                    raw_name = str(p.get('name', 'Inconnu'))
                    display_name = self.fix_text(raw_name)
                    raw_img_path = p.get('image_path', '')
                    final_img_path = ''
                    if raw_img_path:
                        if os.path.exists(raw_img_path):
                            final_img_path = raw_img_path
                        else:
                            filename = os.path.basename(raw_img_path)
                            potential_path = os.path.join(img_dir, filename)
                            if os.path.exists(potential_path):
                                final_img_path = potential_path
                    rv_data.append({'name': display_name, 'price_text': price_fmt, 'stock_text': stock_text, 'icon': icon, 'icon_color': icon_col, 'price_color': price_color, 'image_path': final_img_path, 'raw_data': p})
                except Exception as ex:
                    print(f'Error preparing item: {ex}')
                    continue
        except Exception as e:
            print(f'Data Prep Error: {e}')
        self._apply_search_results(rv_data)

    def play_sound(self, type_):
        if platform == 'android' and hasattr(self, 'tone_gen') and self.tone_gen:
            try:
                if type_ == 'success':
                    self.tone_gen.startTone(24, 150)
                elif type_ == 'error':
                    self.tone_gen.startTone(97, 300)
                elif type_ == 'duplicate':
                    self.tone_gen.startTone(29, 150)
            except:
                pass

    def play_beep(self):
        if platform == 'android' and hasattr(self, 'tone_gen') and self.tone_gen:
            try:
                self.tone_gen.startTone(24, 150)
            except:
                pass

    @mainthread
    def _apply_search_results(self, rv_data):
        if self.rv_products:
            self.rv_products.data = rv_data
            self.rv_products.refresh_from_data()

    def open_bluetooth_selector(self, instance):
        if platform != 'android':
            self.notify('Fonction disponible uniquement sur Android', 'error')
            return
        try:
            adapter = BluetoothAdapter.getDefaultAdapter()
            if not adapter or not adapter.isEnabled():
                self.notify('Veuillez activer le Bluetooth', 'error')
                return
            paired_devices = adapter.getBondedDevices().toArray()
            content = MDBoxLayout(orientation='vertical', size_hint_y=None, height=dp(400))
            scroll = MDScrollView()
            list_layout = MDList()
            if not paired_devices:
                list_layout.add_widget(OneLineListItem(text='Aucun appareil associé (Paired)'))
            else:
                for device in paired_devices:
                    d_name = device.getName()
                    d_mac = device.getAddress()
                    item = TwoLineAvatarIconListItem(text=d_name, secondary_text=d_mac, on_release=lambda x, name=d_name, mac=d_mac: self.select_printer(name, mac))
                    item.add_widget(IconLeftWidget(icon='printer-wireless'))
                    list_layout.add_widget(item)
            scroll.add_widget(list_layout)
            content.add_widget(scroll)
            self.bt_dialog = MDDialog(title='Choisir Imprimante', type='custom', content_cls=content, buttons=[MDFlatButton(text='ANNULER', on_release=lambda x: self.bt_dialog.dismiss())])
            self.bt_dialog.open()
        except Exception as e:
            self.notify(f'Erreur Bluetooth: {e}', 'error')

    def select_printer(self, name, mac):
        self.db.set_setting('printer_mac', mac)
        self.db.set_setting('printer_name', name)
        if hasattr(self, 'bt_dialog') and self.bt_dialog:
            self.bt_dialog.dismiss()
        self.notify(f'Sélectionné: {name}', 'success')
        if hasattr(self, 'settings_menu_dialog') and self.settings_menu_dialog:
            self.settings_menu_dialog.dismiss()
            self.open_settings_menu()

    def clear_printer_selection(self, instance):
        self.db.set_setting('printer_mac', '')
        self.db.set_setting('printer_name', '')
        self.notify('Imprimante effacée', 'info')
        if hasattr(self, 'settings_menu_dialog') and self.settings_menu_dialog:
            self.settings_menu_dialog.dismiss()
            self.open_settings_menu()

    def get_wrapped_text(self, text, font, max_width):
        lines = []
        if not text:
            return lines
        words = text.split(' ')
        current_line = []
        for word in words:
            current_line.append(word)
            line_str = ' '.join(current_line)
            bbox = font.getbbox(line_str)
            w = bbox[2] - bbox[0]
            if w > max_width:
                if len(current_line) == 1:
                    lines.append(current_line[0])
                    current_line = []
                else:
                    current_line.pop()
                    lines.append(' '.join(current_line))
                    current_line = [word]
        if current_line:
            lines.append(' '.join(current_line))
        return lines

    def get_image_raster_data(self, image):
        max_width = 576
        if image.width > max_width:
            ratio = max_width / float(image.width)
            new_height = int(image.height * ratio)
            image = image.resize((max_width, new_height), Image.Resampling.LANCZOS)
        if image.width % 8 != 0:
            target_width = image.width // 8 * 8
            image = image.crop((0, 0, target_width, image.height))
        image = image.convert('1')
        width, height = image.size
        xL = width // 8 % 256
        xH = width // 8 // 256
        yL = height % 256
        yH = height // 256
        cmd = b'\x1dv0\x00' + bytes([xL, xH, yL, yH])
        raw_bytes = image.tobytes()
        inverted_bytes = bytearray([b ^ 255 for b in raw_bytes])
        return cmd + inverted_bytes

    def create_receipt_image(self, transaction_data):
        PAPER_WIDTH = 576
        margin = 10
        img_height = 4500
        image = Image.new('RGB', (PAPER_WIDTH, img_height), (255, 255, 255))
        draw = ImageDraw.Draw(image)
        try:
            font_size_reg = 22
            font_size_large = 38
            font_size_med = 26
            font_reg = ImageFont.truetype(FONT_FILE, font_size_reg)
            font_bold = ImageFont.truetype(FONT_FILE, font_size_reg)
            font_large = ImageFont.truetype(FONT_FILE, font_size_large)
            font_med = ImageFont.truetype(FONT_FILE, font_size_med)
        except:
            font_reg = ImageFont.load_default()
            font_bold = font_reg
            font_large = font_reg
            font_med = font_reg

        def proc_ar(text):
            if not text:
                return ''
            try:
                text = str(text)
                reshaped_text = reshaper.reshape(text)
                bidi_text = get_display(reshaped_text)
                return bidi_text
            except:
                return str(text)

        def draw_text_line(text, y_pos, font_obj, align='left', color=(0, 0, 0)):
            if not text:
                return y_pos
            bidi_text = proc_ar(text)
            try:
                bbox = font_obj.getbbox(bidi_text)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            except:
                text_width = 0
                text_height = 20
            x_pos = margin
            if align == 'center':
                x_pos = (PAPER_WIDTH - text_width) // 2
            elif align == 'right':
                x_pos = PAPER_WIDTH - text_width - margin
            draw.text((x_pos, y_pos), bidi_text, font=font_obj, fill=color)
            return y_pos + text_height + 8

        def draw_separator(curr_y):
            draw.line([(margin, curr_y), (PAPER_WIDTH - margin, curr_y)], fill=(0, 0, 0), width=2)
            return curr_y + 10

        def draw_lr(left, right, font, y_pos, is_bold=False):
            l = proc_ar(left)
            r = proc_ar(right)
            try:
                bbox_r = font.getbbox(r)
                bbox_l = font.getbbox(l)
                w_r = bbox_r[2] - bbox_r[0]
                h_r = bbox_r[3] - bbox_r[1]
                h_l = bbox_l[3] - bbox_l[1]
            except:
                w_r = 0
                h_r = 20
                h_l = 20
            x_r = PAPER_WIDTH - w_r - margin
            draw.text((margin, y_pos), l, font=font, fill=(0, 0, 0))
            draw.text((x_r, y_pos), r, font=font, fill=(0, 0, 0))
            if is_bold:
                draw.text((margin + 1, y_pos), l, font=font, fill=(0, 0, 0))
                draw.text((x_r + 1, y_pos), r, font=font, fill=(0, 0, 0))
            return max(h_r, h_l, 30) + 8
        doc_type_raw = str(transaction_data.get('doc_type', transaction_data.get('transaction_type', ''))).strip().upper()
        ref_text = str(transaction_data.get('custom_label') or transaction_data.get('invoice_number') or '')
        stock_factor = AppConstants.STOCK_MOVEMENTS.get(doc_type_raw, 0)
        fin_factor = AppConstants.FINANCIAL_FACTORS.get(doc_type_raw, 0)
        is_transfer = stock_factor == 0 and fin_factor == 0 and (doc_type_raw in ['TR', 'TRANSFER', 'TRANSFERT'])
        if 'TR' in ref_text.upper() and (not is_transfer):
            is_transfer = True
        items = transaction_data.get('items', [])
        stored_total_amount = to_decimal(transaction_data.get('amount', transaction_data.get('total_amount', 0)))
        is_simple = False
        if not is_transfer:
            if transaction_data.get('is_simple_payment'):
                is_simple = True
            if doc_type_raw in ['CLIENT_PAY', 'SUPPLIER_PAY', 'VERSEMENT', 'REGLEMENT']:
                is_simple = True
            if not items and abs(stored_total_amount) > 0:
                is_simple = True
        is_supplier = (stock_factor == 1 or 'SUPPLIER' in doc_type_raw or 'PURCHASE' in doc_type_raw or ('BA' in doc_type_raw)) and (not is_transfer)
        if 'REGLEMENT' in str(ref_text).upper():
            is_supplier = True
        if is_transfer:
            doc_title = 'BON DE TRANSFERT'
        elif is_simple:
            doc_title = ref_text.upper() or ('REGLEMENT' if is_supplier else 'VERSEMENT')
        else:
            visuals = AppConstants.DOC_VISUALS.get(doc_type_raw, {'name': doc_type_raw})
            doc_title = visuals['name'].upper()
        y = 10
        store_name = 'MagPro Store'
        store_address = ''
        store_phone = ''
        if self.db.setting_exists('store_name'):
            store_name = self.db.get_setting('store_name', store_name)
            store_address = self.db.get_setting('store_address', '')
            store_phone = self.db.get_setting('store_phone', '')
        y = draw_text_line(store_name, y, font_large, 'center')
        if store_address:
            y = draw_text_line(store_address, y, font_reg, 'center')
        if store_phone:
            y = draw_text_line(f'Tel: {store_phone}', y, font_reg, 'center')
        y += 5
        y = draw_separator(y)
        y = draw_text_line(doc_title, y, font_large, 'center')
        y += 5
        y = draw_separator(y)
        if (not is_simple or is_transfer) and ref_text and (ref_text != 'None') and (ref_text != doc_title):
            y = draw_text_line(f'Bon N°: {ref_text}', y, font_med, 'left')
        ts_str = str(transaction_data.get('timestamp', ''))[:16] or datetime.now().strftime('%Y-%m-%d %H:%M')
        y = draw_text_line(f'Date: {ts_str}', y, font_reg, 'left')
        user_str = transaction_data.get('user_name', self.current_user_name)
        if user_str and user_str != 'ADMIN':
            y = draw_text_line(f'User: {user_str}', y, font_reg, 'left')
        if is_transfer:
            label_entity = 'Trajet'
            loc = transaction_data.get('purchase_location') or transaction_data.get('location') or 'store'
            entity_name_raw = 'Magasin -> Dépôt' if loc == 'store' else 'Dépôt -> Magasin'
        else:
            label_entity = 'Fournisseur' if is_supplier else 'Client'
            entity_name_raw = transaction_data.get('entity_name') or transaction_data.get('client_name')
            if not entity_name_raw:
                ent_id = transaction_data.get('entity_id')
                if ent_id:
                    target_type = 'supplier' if is_supplier else 'account'
                    found = self.db.get_entity_by_id(ent_id, target_type)
                    if found:
                        entity_name_raw = found.get('name')
            if not entity_name_raw:
                entity_name_raw = AppConstants.DEFAULT_SUPPLIER_NAME if is_supplier else AppConstants.DEFAULT_CLIENT_NAME
        y = draw_text_line(f'{label_entity}: {entity_name_raw}', y, font_med, 'left')
        y += 5
        y = draw_separator(y)
        if is_simple and (not is_transfer):
            y += 20
            y = draw_text_line('MONTANT:', y, font_large, 'center')
            abs_amount = abs(stored_total_amount)
            y = draw_text_line(f'{abs_amount:,.2f} DA'.replace(',', ' ').replace('.', ','), y, font_large, 'center')
            note = str(transaction_data.get('note', '')).strip()
            if note and note != doc_title and (note != 'None'):
                y += 20
                y = draw_text_line(f'Note: {note}', y, font_reg, 'center')
            y += 20
            draw.line([(margin + 100, y), (PAPER_WIDTH - margin - 100, y)], fill=(0, 0, 0), width=1)
            y += 15
            y = draw_text_line('Merci de votre Fidélité !', y, font_med, 'center')
            y += 120
            return image.crop((0, 0, PAPER_WIDTH, y))
        calc_total_ht = Decimal(0)
        calc_total_tva = Decimal(0)
        for item in items:
            raw_prod = item.get('name') or item.get('product_name') or 'Article'
            qty = to_decimal(item.get('qty', 0))
            prod_lines = self.get_wrapped_text(raw_prod, font_bold, PAPER_WIDTH - 2 * margin)
            for line in prod_lines:
                y = draw_text_line(line, y, font_bold, 'left')
            qty_str = str(int(qty)) if qty == qty.to_integral_value() else str(float(qty))
            if is_transfer:
                qty_display = f'Qté : {qty_str}'
                y = draw_text_line(qty_display, y, font_large, 'center')
                draw.line([(margin + 50, y - 2), (PAPER_WIDTH - margin - 50, y - 2)], fill=(200, 200, 200), width=1)
                y += 10
                continue
            price = to_decimal(item.get('price', 0))
            tva_rate = to_decimal(item.get('tva', 0))
            price_str = f'{price:,.2f}'
            line_ht = qty * price
            line_tva = line_ht * (tva_rate / Decimal(100))
            calc_total_ht += line_ht
            calc_total_tva += line_tva
            line_ttc = line_ht + line_tva
            if tva_rate > 0:
                line_calc = f'{qty_str} x {price_str} (TVA {int(tva_rate)}%)'
            else:
                line_calc = f'{qty_str} x {price_str}'
            line_total_str = f'{line_ttc:,.2f}'
            y += draw_lr(line_calc, line_total_str, font_reg, y)
            draw.line([(margin + 50, y - 2), (PAPER_WIDTH - margin - 50, y - 2)], fill=(220, 220, 220), width=1)
            y += 5
        y += 10
        y = draw_separator(y)
        if not is_transfer:
            payment_info = transaction_data.get('payment_info', {})
            saved_paid = Decimal(0)
            if 'amount' in payment_info:
                saved_paid = to_decimal(payment_info['amount'])
            elif 'paid_amount' in transaction_data:
                saved_paid = to_decimal(transaction_data['paid_amount'])
            elif 'payment_details' in transaction_data:
                try:
                    import json
                    details = json.loads(transaction_data['payment_details'])
                    saved_paid = to_decimal(details.get('amount', 0))
                except:
                    pass
            saved_timbre = to_decimal(payment_info.get('timbre', 0))
            pay_method = transaction_data.get('payment_method', '') or payment_info.get('method', '')
            is_cash = any((k in str(pay_method).lower() for k in ['espèce', 'espece']))
            final_timbre = saved_timbre
            if doc_type_raw == 'FC' and is_cash and (saved_timbre == 0):
                base = calc_total_ht + calc_total_tva
                final_timbre = to_decimal(AppConstants.calculate_stamp_duty(base))
            if stored_total_amount > 0:
                total_net = stored_total_amount
            else:
                total_net = calc_total_ht + calc_total_tva + final_timbre
            if calc_total_tva > 0:
                y += draw_lr('Total HT:', f'{calc_total_ht:,.2f}', font_med, y)
                y += draw_lr('Total TVA:', f'{calc_total_tva:,.2f}', font_med, y)
            if final_timbre > 0:
                y += draw_lr('Droit Timbre:', f'{final_timbre:,.2f}', font_med, y)
            if calc_total_tva > 0 or final_timbre > 0:
                y += 5
                draw.line([(margin, y), (PAPER_WIDTH - margin, y)], fill=(0, 0, 0), width=1)
                y += 5
            y += draw_lr('TOTAL:', f'{total_net:,.2f} DA', font_large, y, True)
            is_comptoir = False
            if entity_name_raw and ('COMPTOIR' in str(entity_name_raw).upper() or str(entity_name_raw).upper() == AppConstants.DEFAULT_CLIENT_NAME.upper()):
                is_comptoir = True
            if not is_comptoir:
                y += 5
                y += draw_lr('VERSEMENT:', f'{saved_paid:,.2f} DA', font_med, y)
                reste = total_net - saved_paid
                if abs(reste) < Decimal('0.01'):
                    reste = Decimal(0)
                label_reste = 'RESTE:' if reste >= 0 else 'RENDU:'
                y += draw_lr(label_reste, f'{abs(reste):,.2f} DA', font_med, y, True)
        y += 25
        draw.line([(margin + 100, y), (PAPER_WIDTH - margin - 100, y)], fill=(0, 0, 0), width=1)
        y += 15
        y = draw_text_line('Merci de votre Fidélité !', y, font_med, 'center')
        y += 120
        final_image = image.crop((0, 0, PAPER_WIDTH, y))
        return final_image

    def print_ticket_bluetooth(self, transaction_data):
        if platform != 'android':
            return
        target_mac = self.db.get_setting('printer_mac', '').strip()
        if not target_mac:
            self.notify('Imprimante non configurée', 'error')
            return
        socket = None
        try:
            adapter = BluetoothAdapter.getDefaultAdapter()
            if not adapter or not adapter.isEnabled():
                self.notify('Bluetooth OFF', 'error')
                return
            device = adapter.getRemoteDevice(target_mac)
            uuid = UUID.fromString('00001101-0000-1000-8000-00805F9B34FB')
            img = self.create_receipt_image(transaction_data)
            raster_data = self.get_image_raster_data(img)
            socket = device.createRfcommSocketToServiceRecord(uuid)
            socket.connect()
            time.sleep(0.2)
            output_stream = socket.getOutputStream()
            ESC = b'\x1b'
            GS = b'\x1d'
            INIT = ESC + b'@'
            CUT = GS + b'V\x00'
            output_stream.write(INIT)
            output_stream.flush()
            time.sleep(0.1)
            chunk_size = 1024
            total_size = len(raster_data)
            for i in range(0, total_size, chunk_size):
                chunk = raster_data[i:i + chunk_size]
                output_stream.write(chunk)
                output_stream.flush()
                time.sleep(0.03)
            output_stream.write(b'\n\n')
            output_stream.write(CUT)
            output_stream.flush()
            time.sleep(0.5)
            socket.close()
        except Exception as e:
            try:
                if socket:
                    socket.close()
            except:
                pass
            print(f'Print Error: {e}')
            self.notify("Erreur d'impression (Connexion)", 'error')

    def _round_num(self, value):
        try:
            return round(float(value), 2)
        except (ValueError, TypeError):
            return 0.0

    def calculate_cart_totals(self, cart_items, is_invoice_mode):
        total_ht = Decimal('0.00')
        total_tva = Decimal('0.00')
        for item in cart_items:
            try:
                p = to_decimal(item.get('price', 0))
                q = to_decimal(item.get('qty', 0))
                t_rate = to_decimal(item.get('tva', 0)) if is_invoice_mode else Decimal('0.00')
                line_ht = quantize_decimal(p * q)
                line_tva = quantize_decimal(line_ht * (t_rate / Decimal('100')))
                total_ht += line_ht
                total_tva += line_tva
            except (ValueError, TypeError):
                continue
        return (total_ht, total_tva)

    def load_local_entities(self, entity_type):
        if hasattr(self, 'mgmt_dialog') and self.mgmt_dialog and self.mgmt_dialog.parent:
            if self.current_entity_type_mgmt == entity_type:
                self.load_more_entities(reset=True)
                return
        if hasattr(self, 'entity_dialog') and self.entity_dialog and self.entity_dialog.parent:
            current_source = getattr(self, 'entities_source_type', None)
            if current_source == entity_type:
                self.load_more_entities(reset=True)
                return

    def open_mode(self, mode, skip_dialog=False):
        self.current_mode = mode
        if not skip_dialog:
            self.cart = []
            self.selected_entity = None
        self.selected_location = 'store'
        self.update_cart_button()
        self.selected_family_filter = 'TOUS'
        if hasattr(self, 'btn_main_family_filter'):
            self.btn_main_family_filter.text = 'TOUS'
        title = AppConstants.MODE_TITLES.get(mode, 'Produits')
        color_name = AppConstants.MODE_COLORS.get(mode, 'Blue')
        if hasattr(self, 'prod_toolbar'):
            self.prod_toolbar.title = title
            self.prod_toolbar.right_action_items = []
        self.theme_cls.primary_palette = color_name
        if mode == 'manage_products':
            self.btn_add_prod.opacity = 1
            self.btn_add_prod.disabled = False
            self.btn_add_prod.size_hint_x = None
            self.btn_add_prod.width = dp(48)
            self.btn_scan_prod.opacity = 0
            self.btn_scan_prod.disabled = True
            self.btn_scan_prod.width = 0
            if self.cart_bar:
                self.cart_bar.height = 0
                self.cart_bar.opacity = 0
                self.cart_bar.disabled = True
        else:
            self.btn_add_prod.opacity = 0
            self.btn_add_prod.disabled = True
            self.btn_add_prod.width = 0
            self.btn_scan_prod.opacity = 1
            self.btn_scan_prod.disabled = False
            self.btn_scan_prod.width = dp(48)
            if self.cart_bar:
                self.cart_bar.height = dp(60)
                self.cart_bar.opacity = 1
                self.cart_bar.disabled = False
        self.load_more_products(reset=True)

        def enter_products_screen():
            self.sm.current = 'products'
            if hasattr(self, 'search_field') and self.search_field:
                self.search_field.text = ''
                self.search_field.focus = False
        modes_requiring_entity = ['sale', 'purchase', 'return_sale', 'return_purchase', 'invoice_sale', 'invoice_purchase', 'proforma', 'order_purchase']
        if mode in modes_requiring_entity and (not skip_dialog):
            self.show_entity_selection_dialog(None, next_action=enter_products_screen)
        else:
            enter_products_screen()

    def go_back(self):
        try:
            Window.release_all_keyboards()
            self.editing_transaction_key = None
            self.editing_payment_amount = None
            if hasattr(self, 'editing_payment_method'):
                del self.editing_payment_method
            if hasattr(self, 'search_field') and self.search_field:
                self.search_field.text = ''
            self.cart = []
            self.update_cart_button()
            self.sm.current = 'dashboard'
            self._reset_notification_state(0)
        except Exception as e:
            print(f'Back Error: {e}')
            self.sm.current = 'dashboard'

    def show_zoomed_image(self, image_path, title='Image'):
        from kivymd.uix.fitimage import FitImage
        if not image_path or not os.path.exists(image_path):
            self.notify('الصورة غير متوفرة', 'error')
            return
        content = MDBoxLayout(orientation='vertical', padding=0, spacing=0, size_hint_y=None, height=dp(400))
        img = FitImage(source=image_path, radius=[10, 10, 0, 0], size_hint=(1, 1), mipmap=True)
        content.add_widget(img)
        btn_close = MDRaisedButton(text='Fermer', md_bg_color=(0.2, 0.2, 0.2, 1), text_color=(1, 1, 1, 1), size_hint_x=1, elevation=0, on_release=lambda x: self.zoom_dialog.dismiss())
        content.add_widget(btn_close)
        self.zoom_dialog = MDDialog(title=title, type='custom', content_cls=content, size_hint=(0.9, None))
        self.zoom_dialog.open()

    def open_image_selector(self, instance):
        if platform == 'android':
            self.open_android_image_picker()
        else:
            self.open_desktop_image_picker()

    def open_desktop_image_picker(self):
        try:
            default_path = os.path.join(os.path.expanduser('~'), 'Pictures')
            if not os.path.exists(default_path):
                default_path = os.path.expanduser('~')
            content = MDBoxLayout(orientation='vertical', spacing=10, size_hint_y=None, height=dp(550))
            content.add_widget(MDLabel(text='Sélectionnez une image (.png, .jpg)', font_style='Caption', theme_text_color='Secondary', size_hint_y=None, height=dp(20)))
            list_bg = MDCard(orientation='vertical', size_hint_y=1, md_bg_color=(0.15, 0.15, 0.15, 1), radius=[10], padding='5dp', elevation=0)
            self.img_file_chooser = FileChooserListView(path=default_path, filters=['*.png', '*.jpg', '*.jpeg', '*.webp'], size_hint_y=1)
            list_bg.add_widget(self.img_file_chooser)
            content.add_widget(list_bg)
            btn_box = MDBoxLayout(spacing=10, size_hint_y=None, height=dp(50))
            btn_cancel = MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=self.theme_cls.primary_color, on_release=lambda x: self.img_picker_dialog.dismiss())
            btn_select = MDRaisedButton(text='SÉLECTIONNER', md_bg_color=(0.1, 0.4, 0.8, 1), text_color=(1, 1, 1, 1), on_release=lambda x: self.confirm_image_selection(self.img_file_chooser.selection))
            btn_box.add_widget(btn_cancel)
            btn_box.add_widget(btn_select)
            content.add_widget(btn_box)
            self.img_picker_dialog = MDDialog(title='Choisir une image', type='custom', content_cls=content, size_hint=(0.95, 0.9))
            self.img_picker_dialog.open()
        except Exception as e:
            self.notify(f'Erreur UI: {e}', 'error')

    def open_android_image_picker(self):
        try:
            from jnius import autoclass, cast
            from android import activity
            Intent = autoclass('android.content.Intent')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            intent = Intent(Intent.ACTION_GET_CONTENT)
            intent.setType('image/*')
            intent.addCategory(Intent.CATEGORY_OPENABLE)
            activity.bind(on_activity_result=self._on_image_file_chosen)
            currentActivity = cast('android.app.Activity', PythonActivity.mActivity)
            self.notify('Sélectionnez une image...', 'info')
            currentActivity.startActivityForResult(intent, 103)
        except Exception as e:
            self.notify(f'Erreur Android Picker: {e}', 'error')

    def _on_image_file_chosen(self, requestCode, resultCode, intent):
        from android import activity
        activity.unbind(on_activity_result=self._on_image_file_chosen)
        if requestCode == 103 and resultCode == -1:
            if intent:
                uri = intent.getData()
                if uri:
                    self._copy_image_uri_to_temp(uri)
                else:
                    self.notify('Aucune image sélectionnée (URI Null)', 'error')
        else:
            pass

    def _copy_image_uri_to_temp(self, uri):
        self.notify("Traitement de l'image...", 'info')
        threading.Thread(target=self._background_image_copy_task, args=(uri,), daemon=True).start()

    def _background_image_copy_task(self, uri):
        if platform != 'android':
            if os.path.exists(str(uri)):
                Clock.schedule_once(lambda dt: self.confirm_image_selection([str(uri)]), 0)
            return
        try:
            from jnius import autoclass, cast
            temp_img_path = os.path.join(self.user_data_dir, 'temp_selected_image.jpg')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            currentActivity = cast('android.app.Activity', PythonActivity.mActivity)
            content_resolver = currentActivity.getContentResolver()
            input_stream = content_resolver.openInputStream(uri)
            with open(temp_img_path, 'wb') as f_out:
                buffer = bytearray(8192)
                while True:
                    bytes_read = input_stream.read(buffer)
                    if bytes_read == -1:
                        break
                    f_out.write(buffer[:bytes_read])
            input_stream.close()
            if os.path.exists(temp_img_path) and os.path.getsize(temp_img_path) > 0:
                Clock.schedule_once(lambda dt: self.confirm_image_selection([temp_img_path]), 0)
            else:
                Clock.schedule_once(lambda dt: self.notify('Erreur: Fichier image vide', 'error'), 0)
        except Exception as e:
            print(f'Android Copy Error: {e}')
            Clock.schedule_once(lambda dt: self.notify(f'Erreur import image: {str(e)}', 'error'), 0)

    def confirm_image_selection(self, selection):
        if hasattr(self, 'img_picker_dialog') and self.img_picker_dialog:
            self.img_picker_dialog.dismiss()
        if selection and len(selection) > 0:
            file_path = selection[0]
            self.temp_selected_image_path = file_path
            self.notify('Image sélectionnée', 'success')
            if hasattr(self, 'lbl_image_status'):
                self.lbl_image_status.text = f'Image: {os.path.basename(file_path)}'
                self.lbl_image_status.theme_text_color = 'Custom'
                self.lbl_image_status.text_color = (0, 0.6, 0, 1)
        else:
            self.notify('Aucune sélection', 'warning')

    def save_product_image_local(self, source_path):
        if not source_path or not os.path.exists(source_path):
            return ''
        try:
            img_dir = os.path.join(self.user_data_dir, 'product_images')
            if not os.path.exists(img_dir):
                os.makedirs(img_dir)
            filename = os.path.basename(source_path)
            name, ext = os.path.splitext(filename)
            if not ext or len(ext) > 5:
                ext = '.jpg'
            new_filename = f'prod_{int(time.time())}_{random.randint(1000, 9999)}{ext}'
            dest_path = os.path.join(img_dir, new_filename)
            saved_successfully = False
            try:
                from PIL import Image
                with Image.open(source_path) as img:
                    if img.mode in ('RGBA', 'P'):
                        img = img.convert('RGB')
                        dest_path = os.path.join(img_dir, f'prod_{int(time.time())}_{random.randint(1000, 9999)}.jpg')
                    img.thumbnail((1024, 1024), Image.Resampling.LANCZOS)
                    img.save(dest_path, quality=85, optimize=True)
                    saved_successfully = True
                    print(f'[INFO] Image compressed and saved at: {dest_path}')
            except ImportError:
                print('[WARNING] PIL not found, falling back to shutil')
            except Exception as e:
                print(f'[WARNING] PIL Error: {e}, falling back to shutil')
            if not saved_successfully:
                import shutil
                shutil.copyfile(source_path, dest_path)
                print(f'[INFO] Image copied raw to: {dest_path}')
            return dest_path
        except Exception as e:
            print(f'[ERROR] Save Image Exception: {e}')
            self.notify(f'Erreur sauvegarde image: {e}', 'error')
            return ''

    def build(self):
        Builder.load_string(KV_BUILDER)
        self.title = 'MagPro Gestion de Stock'
        self._search_event = None
        self._entity_search_event = None
        self._notify_event = None
        self.theme_cls.primary_palette = 'Blue'
        self.theme_cls.accent_palette = 'Amber'
        self.theme_cls.theme_style = 'Light'
        self.theme_cls.font_styles['H4'] = ['ArabicFont', 34, False, 0.25]
        self.theme_cls.font_styles['H5'] = ['ArabicFont', 24, False, 0]
        self.theme_cls.font_styles['H6'] = ['ArabicFont', 20, False, 0.15]
        self.theme_cls.font_styles['Subtitle1'] = ['ArabicFont', 16, False, 0.15]
        self.theme_cls.font_styles['Subtitle2'] = ['ArabicFont', 14, False, 0.1]
        self.theme_cls.font_styles['Body1'] = ['ArabicFont', 16, False, 0.5]
        self.theme_cls.font_styles['Body2'] = ['ArabicFont', 14, False, 0.25]
        self.theme_cls.font_styles['Button'] = ['ArabicFont', 14, True, 1.25]
        self.theme_cls.font_styles['Caption'] = ['ArabicFont', 12, False, 0.4]
        self.data_dir = self.user_data_dir
        self.db = DatabaseManager()
        self.is_seller_mode = self.db.get_setting('config_seller_mode', 'False') == 'True'
        self.root_box = MDBoxLayout(orientation='vertical')
        self.sm = MDScreenManager()
        self.sm.add_widget(self._build_login_screen())
        self.sm.add_widget(self._build_dashboard_screen())
        self.sm.add_widget(self._build_products_screen())
        self.sm.add_widget(self._build_cart_screen())
        self.root_box.add_widget(self.sm)
        self.status_bar_bg = MDCard(size_hint_y=None, height=dp(40), radius=[0], md_bg_color=(0.2, 0.2, 0.2, 1), elevation=0)
        self.status_bar_label = MDLabel(text='Initialisation...', halign='center', theme_text_color='Custom', text_color=(1, 1, 1, 1), font_style='Caption', bold=True)
        self.status_bar_bg.add_widget(self.status_bar_label)
        self.root_box.add_widget(self.status_bar_bg)
        Window.bind(on_keyboard=self.on_keyboard)
        return self.root_box

    def get_device_id(self):
        from kivy.utils import platform
        if platform == 'android':
            try:
                from jnius import autoclass
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                content_resolver = PythonActivity.mActivity.getContentResolver()
                Secure = autoclass('android.provider.Settings$Secure')
                android_id = Secure.getString(content_resolver, Secure.ANDROID_ID)
                return str(android_id) if android_id else 'ANDROID_UNKNOWN'
            except Exception:
                return 'ANDROID_ERR_ID'
        elif platform == 'win':
            return 'PC_DEBUG_ID_12345'
        return 'UNKNOWN_DEVICE_ID'

    def check_license_validity(self):
        try:
            stored_key = self.db.get_setting('license_key')
            if not stored_key:
                return False
            device_id = self.get_device_id()
            salt = f'magpro_mobile_v6_{device_id}_secure_key'
            expected_key = hashlib.sha256(salt.encode()).hexdigest()
            return stored_key == expected_key
        except Exception as e:
            print(f'License check error: {e}')
            return False

    def copy_to_clipboard(self, text):
        from kivy.core.clipboard import Clipboard
        Clipboard.copy(text)
        self.notify('ID copié dans le presse-papiers', 'success')

    def validate_activation(self, key_input, dialog_ref):
        try:
            device_id = self.get_device_id()
            salt = f'magpro_mobile_v6_{device_id}_secure_key'
            expected_key = hashlib.sha256(salt.encode()).hexdigest()
            if key_input.strip() == expected_key:
                self.db.set_setting('license_key', expected_key)
                self.notify('Activation réussie ! Bienvenue.', 'success')
                if dialog_ref:
                    dialog_ref.dismiss()
                Clock.schedule_once(self._deferred_start, 0.5)
            else:
                self.notify('Clé invalide. Veuillez vérifier.', 'error')
        except Exception as e:
            self.notify(f'Erreur: {e}', 'error')

    def show_activation_dialog(self):
        from kivy.core.clipboard import Clipboard
        device_id = self.get_device_id()
        content = MDBoxLayout(orientation='vertical', spacing='12dp', size_hint_y=None, adaptive_height=True, padding=['20dp', '20dp', '20dp', '10dp'])
        content.add_widget(MDIcon(icon='shield-check', halign='center', font_size='64sp', theme_text_color='Custom', text_color=self.theme_cls.primary_color, pos_hint={'center_x': 0.5}))
        content.add_widget(MDLabel(text='Activation Requise', halign='center', font_style='H5', bold=True, theme_text_color='Primary', adaptive_height=True))
        id_card = MDCard(orientation='vertical', radius=[10], padding=['15dp', '12dp', '15dp', '12dp'], md_bg_color=(0.96, 0.96, 0.96, 1), elevation=0, size_hint_y=None, adaptive_height=True, spacing='5dp')
        id_card.add_widget(MDLabel(text="ID d'appareil :", halign='left', font_style='Caption', theme_text_color='Secondary', adaptive_height=True))
        id_row = MDBoxLayout(orientation='horizontal', spacing='10dp', adaptive_height=True)
        field_id = MDTextField(text=device_id, readonly=True, font_size='16sp', mode='line', active_line=False, size_hint_x=0.85, pos_hint={'center_y': 0.5})
        btn_copy = MDIconButton(icon='content-copy', theme_text_color='Custom', text_color=self.theme_cls.primary_color, on_release=lambda x: Clipboard.copy(device_id), pos_hint={'center_y': 0.5}, icon_size='22sp')
        id_row.add_widget(field_id)
        id_row.add_widget(btn_copy)
        id_card.add_widget(id_row)
        content.add_widget(id_card)
        key_row = MDBoxLayout(orientation='horizontal', spacing='10dp', adaptive_height=True)
        self.field_key = NoMenuTextField(hint_text='Saisir la clé de licence', mode='rectangle', size_hint_x=0.85, pos_hint={'center_y': 0.5})
        btn_paste = MDIconButton(icon='content-paste', theme_text_color='Custom', text_color=self.theme_cls.primary_color, on_release=lambda x: setattr(self.field_key, 'text', Clipboard.paste()), pos_hint={'center_y': 0.5}, icon_size='22sp')
        key_row.add_widget(self.field_key)
        key_row.add_widget(btn_paste)
        content.add_widget(key_row)
        btn_activate = MDRaisedButton(text="ACTIVER L'APPLICATION", md_bg_color=(0, 0.7, 0, 1), font_size='16sp', elevation=1, size_hint_x=1, size_hint_y=None, height='52dp', on_release=lambda x: self.validate_activation(self.field_key.text, self.activation_dialog_ref))
        content.add_widget(btn_activate)
        self.activation_dialog_ref = MDDialog(title='', type='custom', content_cls=content, size_hint=(0.9, None), auto_dismiss=False, radius=[16, 16, 16, 16])
        self.activation_dialog_ref.open()

    def on_start(self):
        from kivy.clock import Clock
        self.request_android_permissions()
        try:
            from jnius import autoclass
            ToneGenerator = autoclass('android.media.ToneGenerator')
            AudioManager = autoclass('android.media.AudioManager')
            self.tone_gen = ToneGenerator(3, 100)
        except Exception:
            self.tone_gen = None
        self.db = DatabaseManager()
        Clock.schedule_once(self._deferred_start, 0.5)

    def request_android_permissions(self):
        if platform != 'android':
            return
        try:
            from android.permissions import request_permissions, Permission
            from jnius import autoclass

            def callback(permissions, results):
                pass
            Build = autoclass('android.os.Build')
            VERSION = autoclass('android.os.Build$VERSION')
            permissions_list = [Permission.BLUETOOTH, Permission.BLUETOOTH_ADMIN, Permission.ACCESS_COARSE_LOCATION, Permission.ACCESS_FINE_LOCATION]
            if VERSION.SDK_INT >= 31:
                permissions_list.extend(['android.permission.BLUETOOTH_CONNECT', 'android.permission.BLUETOOTH_SCAN'])
            request_permissions(permissions_list, callback)
        except Exception:
            pass

    def _deferred_start(self, dt):
        if not self.check_license_validity():
            self.show_activation_dialog()
            return
        self._auto_login_check(0)
        self.check_and_load_stats()
        self.update_dashboard_layout()

    def select_entity_from_rv(self, entity_data):
        final_name = entity_data.get('name', '')
        category = entity_data.get('price_category', 'Détail')
        self.selected_entity = {'id': entity_data['id'], 'name': final_name, 'category': category}
        if hasattr(self, 'btn_ent_screen'):
            self.btn_ent_screen.text = self.fix_text(final_name)[:15]
            if self.current_mode in ['sale', 'return_sale', 'client_payment', 'invoice_sale', 'proforma']:
                self.btn_ent_screen.md_bg_color = (0, 0.6, 0.6, 1)
            else:
                self.btn_ent_screen.md_bg_color = (0.8, 0.4, 0, 1)
        self.recalculate_cart_prices()
        if hasattr(self, 'entity_dialog') and self.entity_dialog:
            self.entity_dialog.dismiss()
        if hasattr(self, 'pending_entity_next_action') and self.pending_entity_next_action:
            self.pending_entity_next_action()
            self.pending_entity_next_action = None

    def check_and_load_stats(self, target_date=None):
        if target_date:
            date_display = str(target_date)
            if hasattr(self, 'lbl_dashboard_date'):
                self.lbl_dashboard_date.text = f'Tableau de Bord ({date_display})'
        elif hasattr(self, 'lbl_dashboard_date'):
            self.lbl_dashboard_date.text = "Tableau de Bord (Aujourd'hui)"
        stats = self.get_comprehensive_stats(target_date)

        def fmt(val):
            return '{:,.2f}'.format(val).replace(',', ' ').replace('.', ',')
        self.stat_sales_today = stats['sales']
        self.stat_purchases_today = stats['purchases']
        self.stat_client_payments = stats['cash_in']
        self.stat_supplier_payments = stats['cash_out']
        self.stat_net_total = stats['sales'] - stats['purchases']
        try:
            if hasattr(self, 'lbl_new_sales') and self.lbl_new_sales:
                self.lbl_new_sales.text = fmt(stats['sales']) + ' DA'
            if hasattr(self, 'lbl_new_purchases') and self.lbl_new_purchases:
                self.lbl_new_purchases.text = fmt(stats['purchases']) + ' DA'
            if hasattr(self, 'lbl_new_in') and self.lbl_new_in:
                self.lbl_new_in.text = fmt(stats['cash_in']) + ' DA'
            if hasattr(self, 'lbl_new_out') and self.lbl_new_out:
                self.lbl_new_out.text = fmt(stats['cash_out']) + ' DA'
            if hasattr(self, 'lbl_new_stock_val') and self.lbl_new_stock_val:
                self.lbl_new_stock_val.text = fmt(stats['stock_value']) + ' DA'
            if hasattr(self, 'lbl_new_profit') and self.lbl_new_profit:
                profit = stats['profit']
                self.lbl_new_profit.text = fmt(profit) + ' DA'
                if profit < 0:
                    self.lbl_new_profit.text_color = (0.8, 0, 0, 1)
                else:
                    self.lbl_new_profit.text_color = (0, 0.5, 0.5, 1)
        except Exception as e:
            print(f'Error UI stats: {e}')

    def open_stats_date_picker(self, instance):
        date_dialog = MDDatePicker()
        date_dialog.bind(on_save=self.on_stats_date_save)
        date_dialog.open()

    def on_stats_date_save(self, instance, value, date_range):
        self.check_and_load_stats(value)

    def open_history_date_picker(self, instance):
        date_dialog = MDDatePicker()
        date_dialog.bind(on_save=self.on_history_date_save)
        date_dialog.open()

    def on_history_date_save(self, instance, value, date_range):
        self.btn_hist_date.text = str(value)
        self.filter_history_list(specific_date=value)

    def reset_local_stats(self):
        self.stat_sales_today = 0
        self.stat_purchases_today = 0
        self.stat_client_payments = 0
        self.stat_supplier_payments = 0
        self.save_local_stats()

    def save_local_stats(self):
        today_str = str(datetime.now().date())
        self.db.save_stats_data(today_str, self.stat_sales_today, self.stat_purchases_today, self.stat_client_payments, self.stat_supplier_payments)

    def calculate_net_total(self):
        self.stat_net_total = self.stat_sales_today + self.stat_client_payments - (self.stat_purchases_today + self.stat_supplier_payments)
        self.update_dashboard_labels()

    def update_local_entity_balance(self, entity_id, change_amount):
        if not entity_id:
            return
        try:
            target_entity = self.db.get_entity_by_id(entity_id, 'account')
            is_client = True
            if not target_entity:
                target_entity = self.db.get_entity_by_id(entity_id, 'supplier')
                is_client = False
            if target_entity:
                self.db.update_entity_balance(entity_id, change_amount, 'account' if is_client else 'supplier')
                if hasattr(self, 'mgmt_dialog') and self.mgmt_dialog:
                    current_type = getattr(self, 'current_entity_type_mgmt', '')
                    target_type = 'account' if is_client else 'supplier'
                    if current_type == target_type:
                        self.load_local_entities(current_type)
        except Exception as e:
            print(f'Error updating balance: {e}')

    def filter_entity_history_list(self, day_offset=None, specific_date=None):
        if not hasattr(self, 'rv_entity_history') or not self.rv_entity_history:
            return
        if not hasattr(self, 'history_target_entity') or not self.history_target_entity:
            self.rv_entity_history.data = []
            return
        raw_id = self.history_target_entity.get('id')
        if raw_id is None:
            self.rv_entity_history.data = []
            return
        current_cat = 'client'
        if hasattr(self, 'current_entity_type_mgmt') and self.current_entity_type_mgmt == 'supplier':
            current_cat = 'supplier'
        elif self.history_target_entity.get('price_category') == 'Gros':
            current_cat = 'supplier'
        inactive_color = (0.5, 0.5, 0.5, 1)
        active_color = self.theme_cls.primary_color
        target_date = None
        has_ui = hasattr(self, 'btn_ent_hist_today') and self.btn_ent_hist_today
        if specific_date:
            target_date = specific_date
            if has_ui:
                self.btn_ent_hist_today.md_bg_color = inactive_color
                self.btn_ent_hist_yesterday.md_bg_color = inactive_color
                self.btn_ent_hist_date.md_bg_color = active_color
                self.btn_ent_hist_date.text = str(specific_date)
        else:
            if day_offset is None:
                day_offset = 0
            target_date = datetime.now().date() - timedelta(days=day_offset)
            if has_ui:
                self.btn_ent_hist_today.md_bg_color = active_color if day_offset == 0 else inactive_color
                self.btn_ent_hist_yesterday.md_bg_color = active_color if day_offset == 1 else inactive_color
                self.btn_ent_hist_date.md_bg_color = inactive_color
                self.btn_ent_hist_date.text = 'CALENDRIER'
        self.rv_entity_history.data = [{'raw_text': 'Chargement...', 'raw_sec': '', 'amount_text': '', 'icon': 'timer-sand', 'icon_color': [0.5, 0.5, 0.5, 1], 'bg_color': [1, 1, 1, 1], 'is_local': True, 'raw_data': None}]

        def job():
            transactions = self.db.get_transactions(target_date=target_date, entity_id=raw_id, entity_category=current_cat)
            Clock.schedule_once(lambda dt: self.render_transactions_list(transactions, self.rv_entity_history, is_global_mode=False, reset=True))
        threading.Thread(target=job).start()

    def open_entity_history_dialog(self, entity):
        self.history_target_entity = entity
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, height=dp(550))
        tabs_box = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(50), spacing=5)
        self.btn_ent_hist_today = MDRaisedButton(text='AUJ.', size_hint_x=0.33, elevation=0)
        self.btn_ent_hist_today.bind(on_release=lambda x: self.filter_entity_history_list(day_offset=0))
        self.btn_ent_hist_yesterday = MDRaisedButton(text='HIER', size_hint_x=0.33, elevation=0, md_bg_color=(0.5, 0.5, 0.5, 1))
        self.btn_ent_hist_yesterday.bind(on_release=lambda x: self.filter_entity_history_list(day_offset=1))
        self.btn_ent_hist_date = MDRaisedButton(text='CALENDRIER', size_hint_x=0.33, elevation=0, md_bg_color=(0.5, 0.5, 0.5, 1))
        self.btn_ent_hist_date.bind(on_release=self.open_entity_history_date_picker)
        tabs_box.add_widget(self.btn_ent_hist_today)
        tabs_box.add_widget(self.btn_ent_hist_yesterday)
        tabs_box.add_widget(self.btn_ent_hist_date)
        content.add_widget(tabs_box)
        self.rv_entity_history = HistoryRecycleView()
        content.add_widget(self.rv_entity_history)
        c_name = self.fix_text(entity.get('name', 'Client'))
        self.entity_hist_dialog = MDDialog(title=f'Historique: {c_name}', type='custom', content_cls=content, size_hint=(0.95, 0.9))
        self.entity_hist_dialog.open()
        self.filter_entity_history_list(day_offset=0)

    def submit_simple_payment(self, x):
        current_time = time.time()
        if current_time - getattr(self, '_last_click_time', 0) < 1.0:
            return
        self._last_click_time = current_time
        if getattr(self, 'is_transaction_in_progress', False):
            return
        self.is_transaction_in_progress = True
        try:
            val_str = self.txt_simple_amount.get_value()
            if not val_str:
                raise ValueError
            amount = float(val_str)
        except:
            self.notify('Montant invalide', 'error')
            self.is_transaction_in_progress = False
            return
        if amount == 0:
            self.notify('Le montant ne peut pas être 0', 'error')
            self.is_transaction_in_progress = False
            return
        if self.simple_pay_dialog:
            self.simple_pay_dialog.dismiss()
        is_client_mode = self.current_mode in ['client_payment', 'client_pay']
        base_type = self.editing_doc_type if self.editing_transaction_key and self.editing_doc_type else 'CLIENT_PAY' if is_client_mode else 'SUPPLIER_PAY'
        visuals = AppConstants.DOC_VISUALS.get(base_type, {'name': 'Opération'})
        custom_label = visuals['name']
        if amount < 0:
            custom_label = 'Dette Initiale' if not is_client_mode else 'Crédit Initial'
        final_note = self.temp_note.strip() if hasattr(self, 'temp_note') and self.temp_note.strip() else custom_label
        timestamp = str(datetime.now()).split('.')[0]
        data = {'entity_id': self.selected_entity['id'], 'amount': amount, 'doc_type': base_type, 'custom_label': custom_label, 'user_name': self.current_user_name, 'note': final_note, 'is_simple_payment': True, 'timestamp': timestamp, 'items': [], 'payment_info': {'amount': amount, 'method': 'Espèce', 'total': amount}}
        if self.editing_transaction_key:
            data['id'] = self.editing_transaction_key
        try:
            self.db.save_transaction(data)
            self.notify(f'{custom_label} Enregistré', 'success')
            try:
                if self.db.get_setting('printer_auto', 'False') == 'True' and self.db.get_setting('printer_mac', ''):
                    threading.Thread(target=self.print_ticket_bluetooth, args=(data,), daemon=True).start()
            except:
                pass
            self.is_transaction_in_progress = False
            self.editing_transaction_key = None
            self.editing_doc_type = None
            self.temp_note = ''
            self.go_back()
            entity_type = 'account' if is_client_mode else 'supplier'
            self.load_local_entities(entity_type)
            if hasattr(self, 'pending_dialog') and self.pending_dialog:
                try:
                    self.filter_history_list(day_offset=0)
                except:
                    pass
            if hasattr(self, 'entity_hist_dialog') and self.entity_hist_dialog:
                try:
                    self.filter_entity_history_list(day_offset=0)
                except:
                    pass
            self.check_and_load_stats()
        except Exception as e:
            print(f'Payment Error: {e}')
            self.notify('Erreur Sauvegarde', 'error')
            self.is_transaction_in_progress = False

    def update_dashboard_labels(self):
        try:
            if hasattr(self, 'lbl_stat_sales') and self.lbl_stat_sales:
                self.lbl_stat_sales.text = f'{self.stat_sales_today:.2f} DA'
            if hasattr(self, 'lbl_stat_purchases') and self.lbl_stat_purchases:
                self.lbl_stat_purchases.text = f'{self.stat_purchases_today:.2f} DA'
            if hasattr(self, 'lbl_stat_client_pay') and self.lbl_stat_client_pay:
                self.lbl_stat_client_pay.text = f'{self.stat_client_payments:.2f} DA'
            if hasattr(self, 'lbl_stat_supp_pay') and self.lbl_stat_supp_pay:
                self.lbl_stat_supp_pay.text = f'{self.stat_supplier_payments:.2f} DA'
            if hasattr(self, 'lbl_stat_net') and self.lbl_stat_net:
                self.lbl_stat_net.text = f'{self.stat_net_total:.2f} DA'
        except:
            pass

    def update_dashboard_layout(self):
        if not self.buttons_container or not self.stats_card_container:
            return
        self.buttons_container.clear_widgets()
        col_green = (0, 0.7, 0, 1)
        col_blue = (0, 0, 0.8, 1)
        col_purple = (0.5, 0, 0.5, 1)
        col_red = (0.8, 0, 0, 1)
        col_teal = (0, 0.5, 0.5, 1)
        col_orange = (1, 0.6, 0, 1)
        col_deep_orange = (1, 0.3, 0, 1)
        col_brown = (0.4, 0.2, 0.1, 1)
        col_cyan = (0, 0.6, 0.6, 1)
        bg_green = (0.9, 1, 0.9, 1)
        bg_blue = (0.9, 0.95, 1, 1)
        bg_purple = (0.95, 0.9, 1, 1)
        bg_red = (1, 0.9, 0.9, 1)
        bg_teal = (0.8, 1, 1, 1)
        bg_orange = (1, 0.95, 0.8, 1)
        bg_deep_orange = (1, 0.9, 0.8, 1)
        bg_brown = (1, 0.85, 0.85, 1)
        if self.is_seller_mode:
            self.buttons_container.add_widget(self._create_dash_btn('cart', 'VENTE (BV)', bg_green, col_green, lambda x: self.open_mode('sale')))
            grid = MDGridLayout(cols=2, spacing=dp(10), adaptive_height=True)
            grid.add_widget(self._create_dash_btn('keyboard-return', 'RETOUR CL.', bg_red, col_red, lambda x: self.open_mode('return_sale')))
            grid.add_widget(self._create_dash_btn('account-group', 'CLIENTS', bg_teal, col_teal, lambda x: self.open_entity_manager('account')))
            self.buttons_container.add_widget(grid)
        else:
            grid = MDGridLayout(cols=2, spacing=dp(10), adaptive_height=True)
            grid.add_widget(self._create_dash_btn('cart', 'VENTE (BV)', bg_green, col_green, lambda x: self.open_mode('sale')))
            grid.add_widget(self._create_dash_btn('truck', 'ACHAT (BA)', bg_orange, col_orange, lambda x: self.open_mode('purchase')))
            grid.add_widget(self._create_dash_btn('file-document', 'FACTURE (FC)', bg_blue, col_blue, lambda x: self.open_mode('invoice_sale')))
            grid.add_widget(self._create_dash_btn('file-document-edit', 'FACT. ACHAT (FF)', bg_deep_orange, col_deep_orange, lambda x: self.open_mode('invoice_purchase')))
            grid.add_widget(self._create_dash_btn('file-document-outline', 'PROFORMA (FP)', bg_purple, col_purple, lambda x: self.open_mode('proforma')))
            grid.add_widget(self._create_dash_btn('clipboard-list', 'COMMANDE (DP)', bg_teal, col_cyan, lambda x: self.open_mode('order_purchase')))
            grid.add_widget(self._create_dash_btn('keyboard-return', 'RETOUR CL.', bg_red, col_red, lambda x: self.open_mode('return_sale')))
            grid.add_widget(self._create_dash_btn('undo', 'RETOUR FR.', bg_blue, col_blue, lambda x: self.open_mode('return_purchase')))
            grid.add_widget(self._create_dash_btn('account-group', 'CLIENTS', bg_teal, col_teal, lambda x: self.open_entity_manager('account')))
            grid.add_widget(self._create_dash_btn('truck-delivery', 'FOURNISSEURS', bg_brown, col_brown, lambda x: self.open_entity_manager('supplier')))
            grid.add_widget(self._create_dash_btn('database-edit', 'PRODUITS', bg_blue, col_blue, lambda x: self.open_mode('manage_products')))
            grid.add_widget(self._create_dash_btn('transfer', 'TRANSFERT (TR)', bg_purple, col_purple, lambda x: self.open_mode('transfer')))
            self.buttons_container.add_widget(grid)
        self.stats_card_container.clear_widgets()
        self.stats_card_container.md_bg_color = (0, 0, 0, 0)
        self.stats_card_container.elevation = 0
        self.stats_card_container.size_hint_y = None
        self.stats_card_container.adaptive_height = True
        self.stats_card_container.padding = [0, dp(5), 0, dp(20)]
        header_box = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(40))
        self.lbl_dashboard_date = MDLabel(text="Tableau de Bord (Aujourd'hui)", font_style='H6', bold=True, theme_text_color='Primary', valign='center')
        btn_date = MDIconButton(icon='calendar-month', theme_text_color='Custom', text_color=col_blue, on_release=self.open_stats_date_picker)
        header_box.add_widget(self.lbl_dashboard_date)
        header_box.add_widget(btn_date)
        self.stats_card_container.add_widget(header_box)
        stats_grid = MDGridLayout(cols=2, spacing=dp(10), adaptive_height=True)

        def create_modern_stat_card(icon, title, value_id, color_bg, color_icon):
            card = MDCard(orientation='vertical', padding=dp(12), radius=[15], md_bg_color=color_bg, elevation=1, size_hint_y=None, height=dp(100))
            top_box = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30))
            top_box.add_widget(MDIcon(icon=icon, theme_text_color='Custom', text_color=color_icon, font_size='26sp'))
            top_box.add_widget(MDLabel(text=title, halign='right', font_style='Caption', bold=True, theme_text_color='Secondary'))
            val_label = MDLabel(text='...', halign='left', font_style='H5', bold=True, theme_text_color='Custom', text_color=(0.2, 0.2, 0.2, 1))
            setattr(self, value_id, val_label)
            card.add_widget(top_box)
            card.add_widget(MDBoxLayout(size_hint_y=None, height=dp(10)))
            card.add_widget(val_label)
            return card
        stats_grid.add_widget(create_modern_stat_card('cart-outline', 'Ventes', 'lbl_new_sales', (0.9, 0.95, 1, 1), (0, 0, 0.8, 1)))
        if not self.is_seller_mode:
            stats_grid.add_widget(create_modern_stat_card('truck-outline', 'Total Achats', 'lbl_new_purchases', (1, 0.95, 0.8, 1), (0.9, 0.5, 0, 1)))
        stats_grid.add_widget(create_modern_stat_card('cash-plus', 'Encaissements', 'lbl_new_in', (0.95, 1, 0.95, 1), (0, 0.5, 0.5, 1)))
        if not self.is_seller_mode:
            stats_grid.add_widget(create_modern_stat_card('cash-minus', 'Décaissements', 'lbl_new_out', (1, 0.9, 0.9, 1), (0.8, 0, 0, 1)))
        if not self.is_seller_mode:
            stats_grid.add_widget(create_modern_stat_card('package-variant', 'Valeur Stock', 'lbl_new_stock_val', (1, 0.95, 0.8, 1), (0.8, 0.4, 0, 1)))
            stats_grid.add_widget(create_modern_stat_card('chart-line', 'Bénéfice', 'lbl_new_profit', (0.95, 0.9, 1, 1), (0.5, 0, 0.5, 1)))
        self.stats_card_container.add_widget(stats_grid)
        self.check_and_load_stats()

    def open_entity_history_date_picker(self, instance):
        date_dialog = MDDatePicker()
        date_dialog.bind(on_save=self.on_entity_history_date_save)
        date_dialog.open()

    def on_entity_history_date_save(self, instance, value, date_range):
        self.btn_ent_hist_date.text = str(value)
        self.filter_entity_history_list(specific_date=value)

    def open_entity_manager(self, entity_type):
        self.current_entity_type_mgmt = entity_type
        self.current_entity_sort = 'name'
        title_text = 'Gestion Clients' if entity_type == 'account' else 'Gestion Fournisseurs'
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, height=dp(600), spacing=dp(5))
        search_layout = MDBoxLayout(orientation='horizontal', spacing=dp(5), size_hint_y=None, height=dp(55))
        self.entity_search = SmartTextField(hint_text='Rechercher...', icon_right='magnify', size_hint_x=0.85)
        self.entity_search.bind(text=self.filter_entities_paginated)
        self.btn_sort_entity = MDIconButton(icon='sort-alphabetical-variant', theme_text_color='Custom', text_color=(0.2, 0.2, 0.2, 1), md_bg_color=(0.95, 0.95, 0.95, 1), size_hint=(None, None), size=(dp(48), dp(48)), pos_hint={'center_y': 0.5}, on_release=self.toggle_entity_sort)
        search_layout.add_widget(self.entity_search)
        search_layout.add_widget(self.btn_sort_entity)
        content.add_widget(search_layout)
        self.rv_mgmt_entity = MgmtEntityRecycleView()
        content.add_widget(self.rv_mgmt_entity)
        btn_add = MDFillRoundFlatButton(text='AJOUTER NOUVEAU', size_hint_x=1, md_bg_color=(0, 0.7, 0, 1), on_release=lambda x: self.show_add_edit_entity_dialog(None))
        content.add_widget(btn_add)
        self.mgmt_dialog = MDDialog(title=title_text, type='custom', content_cls=content, size_hint=(0.95, 0.9))
        self.mgmt_dialog.open()
        self.active_entity_rv = self.rv_mgmt_entity
        self.load_more_entities(reset=True)

    def toggle_entity_sort(self, instance):
        if self.current_entity_sort == 'name':
            self.current_entity_sort = 'balance'
            instance.icon = 'sort-numeric-descending'
            instance.md_bg_color = (0.8, 0.9, 1, 1)
            self.notify('Tri: Par Solde (Décroissant)', 'info')
        elif self.current_entity_sort == 'balance':
            self.current_entity_sort = 'active'
            instance.icon = 'check-bold'
            instance.md_bg_color = (0.8, 1, 0.8, 1)
            self.notify("Tri: Opérations d'aujourd'hui d'abord", 'info')
        else:
            self.current_entity_sort = 'name'
            instance.icon = 'sort-alphabetical-variant'
            instance.md_bg_color = (0.95, 0.95, 0.95, 1)
            self.notify('Tri: Alphabétique (A-Z)', 'info')
        self.load_more_entities(reset=True)

    def start_direct_payment_from_manager(self, entity):
        self.selected_entity = entity
        if self.current_entity_type_mgmt == 'account':
            self.current_mode = 'client_payment'
        else:
            self.current_mode = 'supplier_payment'
        self.show_simple_payment_dialog()

    def open_entity_edit_menu(self, entity):
        self.mgmt_selected_entity = entity
        title_text = self.fix_text(entity['name'])
        content = MDBoxLayout(orientation='vertical', spacing=dp(10), size_hint_y=None, height=dp(160), padding=[dp(15), 0, dp(15), dp(10)])
        gps_data = entity.get('gps_location', '')
        from kivymd.uix.button import MDFillRoundFlatIconButton
        btn_map = MDFillRoundFlatIconButton(text='Localiser (Maps)', icon='google-maps', size_hint_x=1, height=dp(50), md_bg_color=(0.1, 0.7, 0.3, 1) if gps_data else (0.8, 0.8, 0.8, 1), theme_text_color='Custom', text_color=(1, 1, 1, 1), icon_color=(1, 1, 1, 1), font_size='17sp', on_release=lambda x: [self.options_dialog.dismiss(), self.open_client_location(gps_data)])
        content.add_widget(btn_map)
        from kivymd.uix.card import MDSeparator
        content.add_widget(MDSeparator(height=dp(1), color=(0.9, 0.9, 0.9, 1)))
        actions_row = MDBoxLayout(orientation='horizontal', spacing=dp(10), size_hint_y=None, height=dp(50))
        btn_edit = MDFillRoundFlatIconButton(text='Modifier', icon='pencil', size_hint_x=0.5, height=dp(48), md_bg_color=(0.15, 0.45, 0.8, 1), theme_text_color='Custom', text_color=(1, 1, 1, 1), icon_color=(1, 1, 1, 1), font_size='16sp', on_release=lambda x: [self.options_dialog.dismiss(), self.show_add_edit_entity_dialog(entity)])
        from kivymd.uix.button import MDRoundFlatIconButton
        btn_del = MDRoundFlatIconButton(text='Supprimer', icon='delete', size_hint_x=0.5, height=dp(48), theme_text_color='Custom', text_color=(0.85, 0.1, 0.1, 1), icon_color=(0.85, 0.1, 0.1, 1), line_color=(0.85, 0.1, 0.1, 1), font_size='16sp', on_release=lambda x: self.confirm_delete_entity(entity))
        actions_row.add_widget(btn_edit)
        actions_row.add_widget(btn_del)
        content.add_widget(actions_row)
        self.options_dialog = MDDialog(title=title_text, type='custom', content_cls=content, size_hint=(0.95, None), radius=[16, 16, 16, 16])
        self.options_dialog.open()

    def show_add_edit_entity_dialog(self, entity=None):
        from kivy.core.clipboard import Clipboard
        is_edit = entity is not None
        title = 'Modifier la Fiche' if is_edit else 'Ajouter un Nouveau'
        val_name = entity.get('name', '') if is_edit else ''
        val_phone = entity.get('phone', '') if is_edit else ''
        val_address = entity.get('address', '') if is_edit else ''
        val_activity = entity.get('activity', '') if is_edit else ''
        val_email = entity.get('email', '') if is_edit else ''
        val_rc = entity.get('rc', '') if is_edit else ''
        val_nif = entity.get('nif', '') if is_edit else ''
        val_nis = entity.get('nis', '') if is_edit else ''
        val_nai = entity.get('nai', '') if is_edit else ''
        val_gps = entity.get('gps_location', '') if is_edit else ''
        raw_cat = str(entity.get('price_category', '')).strip() if is_edit else ''
        if raw_cat == 'Gros':
            display_cat = 'Gros'
        elif raw_cat == 'Demi-Gros':
            display_cat = 'Demi-Gros'
        else:
            display_cat = 'Détail'
        scroll = MDScrollView(size_hint_y=None, height=dp(600))
        main_box = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(15), padding=[dp(10), dp(10), dp(10), dp(20)])
        card_info = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.99, 0.99, 0.99, 1))
        header_info = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
        header_info.add_widget(MDIcon(icon='account-box-outline', theme_text_color='Primary', font_size='22sp'))
        header_info.add_widget(MDLabel(text='Identité', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
        card_info.add_widget(header_info)
        card_info.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
        f_name = SmartTextField(text=val_name, hint_text='Nom Complet / Raison Sociale *', required=True, icon_right='account')
        f_activity = SmartTextField(text=val_activity, hint_text='Activité', icon_right='briefcase')
        card_info.add_widget(f_name)
        card_info.add_widget(f_activity)
        main_box.add_widget(card_info)
        card_contact = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.96, 0.98, 1, 1))
        header_contact = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
        header_contact.add_widget(MDIcon(icon='card-account-phone-outline', theme_text_color='Primary', font_size='22sp'))
        header_contact.add_widget(MDLabel(text='Coordonnées', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
        card_contact.add_widget(header_contact)
        card_contact.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
        f_phone = SmartTextField(text=val_phone, hint_text='Téléphone', input_filter='int', icon_right='phone')
        f_address = SmartTextField(text=val_address, hint_text='Adresse', icon_right='map-marker')
        box_gps = MDBoxLayout(orientation='horizontal', spacing=dp(5), adaptive_height=True)
        f_gps = SmartTextField(text=val_gps, hint_text='Lien Maps / Position GPS', icon_right='google-maps', size_hint_x=0.85)
        btn_paste_gps = MDIconButton(icon='content-paste', theme_text_color='Custom', text_color=self.theme_cls.primary_color, pos_hint={'center_y': 0.5}, on_release=lambda x: setattr(f_gps, 'text', Clipboard.paste()))
        box_gps.add_widget(f_gps)
        box_gps.add_widget(btn_paste_gps)
        f_email = SmartTextField(text=val_email, hint_text='Email', icon_right='email')
        card_contact.add_widget(f_phone)
        card_contact.add_widget(f_address)
        card_contact.add_widget(box_gps)
        card_contact.add_widget(f_email)
        main_box.add_widget(card_contact)
        f_price_cat = MDTextField(text=display_cat, hint_text='Catégorie de Prix', readonly=True, icon_right='tag')

        def on_cat_touch(instance, touch):
            if instance.collide_point(*touch.pos):
                self.show_price_cat_selector(instance)
                return True
            return False
        f_price_cat.bind(on_touch_down=on_cat_touch)
        if self.current_entity_type_mgmt == 'account':
            card_comm = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(1, 0.98, 0.96, 1))
            header_comm = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
            header_comm.add_widget(MDIcon(icon='store-cog-outline', theme_text_color='Primary', font_size='22sp'))
            header_comm.add_widget(MDLabel(text='Commercial', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
            card_comm.add_widget(header_comm)
            card_comm.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
            card_comm.add_widget(f_price_cat)
            main_box.add_widget(card_comm)
        card_fisc = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.95, 0.95, 0.95, 1))
        header_fisc = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
        header_fisc.add_widget(MDIcon(icon='file-document-multiple-outline', theme_text_color='Primary', font_size='22sp'))
        header_fisc.add_widget(MDLabel(text='Information Fiscale', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
        card_fisc.add_widget(header_fisc)
        card_fisc.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
        f_rc = SmartTextField(text=val_rc, hint_text='N° Registre Commerce (RC)')
        f_nif = SmartTextField(text=val_nif, hint_text='N.I.F (Identifiant Fiscal)')
        f_nis = SmartTextField(text=val_nis, hint_text='N.I.S (Statistique)')
        f_nai = SmartTextField(text=val_nai, hint_text='N.A.I (Article)')
        card_fisc.add_widget(f_rc)
        card_fisc.add_widget(f_nif)
        card_fisc.add_widget(f_nis)
        card_fisc.add_widget(f_nai)
        main_box.add_widget(card_fisc)
        footer_box = MDBoxLayout(orientation='vertical', spacing=dp(10), adaptive_height=True, padding=[0, dp(10), 0, 0])

        def save(x):
            name_val = f_name.get_value().strip()
            if not name_val:
                f_name.error = True
                self.notify('Le nom est obligatoire', 'error')
                return
            cat_val = 'Détail'
            if self.current_entity_type_mgmt == 'account':
                cat_val = f_price_cat.text
            payload = {'action': 'update' if is_edit else 'add', 'type': self.current_entity_type_mgmt, 'name': name_val, 'phone': f_phone.get_value().strip(), 'address': f_address.get_value().strip(), 'activity': f_activity.get_value().strip(), 'email': f_email.get_value().strip(), 'price_category': cat_val, 'rc': f_rc.get_value().strip(), 'nif': f_nif.get_value().strip(), 'nis': f_nis.get_value().strip(), 'nai': f_nai.get_value().strip(), 'gps_location': f_gps.get_value().strip(), 'id': entity.get('id') if is_edit else None}
            try:
                self.db.save_entity(payload)
                self.ae_dialog.dismiss()
                self.notify('Enregistré avec succès', 'success')
                self.load_local_entities(self.current_entity_type_mgmt)
            except Exception as e:
                self.notify(f'Erreur: {e}', 'error')
        btn_save = MDRaisedButton(text='ENREGISTRER', md_bg_color=(0, 0.7, 0, 1), text_color=(1, 1, 1, 1), size_hint_x=1, height=dp(50), elevation=2, on_release=save)
        btn_cancel = MDRaisedButton(text='ANNULER', md_bg_color=(0.9, 0.9, 0.9, 1), text_color=(0.3, 0.3, 0.3, 1), size_hint_x=1, height=dp(50), elevation=0, on_release=lambda x: self.ae_dialog.dismiss())
        footer_box.add_widget(btn_save)
        footer_box.add_widget(btn_cancel)
        main_box.add_widget(footer_box)
        scroll.add_widget(main_box)
        self.ae_dialog = MDDialog(title=title, type='custom', content_cls=scroll, buttons=[], size_hint=(0.98, 0.96))
        self.ae_dialog.open()

    def show_price_cat_selector(self, text_field_instance):
        content = MDBoxLayout(orientation='vertical', spacing=10, size_hint_y=None, height=dp(160), padding=dp(10))

        def select(value):
            text_field_instance.text = value
            self.cat_dialog.dismiss()
        content.add_widget(MDRaisedButton(text='Détail', size_hint_x=1, md_bg_color=(0, 0.6, 0.6, 1), on_release=lambda x: select('Détail')))
        content.add_widget(MDRaisedButton(text='Demi-Gros', size_hint_x=1, md_bg_color=(0.9, 0.6, 0, 1), on_release=lambda x: select('Demi-Gros')))
        content.add_widget(MDRaisedButton(text='Gros', size_hint_x=1, md_bg_color=(0.5, 0, 0.5, 1), on_release=lambda x: select('Gros')))
        self.cat_dialog = MDDialog(title='Choisir Catégorie', type='custom', content_cls=content, size_hint=(0.8, None))
        self.cat_dialog.open()

    def notify(self, text, type='info'):
        if not self.status_bar_label:
            return
        color_map = {'success': (0, 0.6, 0, 1), 'error': (0.8, 0.1, 0.1, 1), 'warning': (0.9, 0.5, 0, 1), 'info': (0.2, 0.2, 0.2, 1)}
        self.status_bar_label.text = text
        self.status_bar_bg.md_bg_color = color_map.get(type, (0.2, 0.2, 0.2, 1))
        if self._notify_event:
            self._notify_event.cancel()
        self._notify_event = Clock.schedule_once(self._reset_notification_state, 3)

    def _reset_notification_state(self, dt):
        if not self.status_bar_label:
            return
        self.status_bar_label.markup = True
        self._notify_event = None
        self.status_bar_label.text = 'Prêt'
        self.status_bar_bg.md_bg_color = (0.2, 0.2, 0.2, 1)

    def confirm_delete_entity(self, entity):
        if hasattr(self, 'options_dialog') and self.options_dialog:
            self.options_dialog.dismiss()
        entity_id = entity['id']
        has_transactions = False
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM transactions WHERE entity_id = ?', (entity_id,))
            count = cursor.fetchone()[0]
            if count > 0:
                has_transactions = True
            conn.close()
        except Exception as e:
            print(f'Error checking transactions: {e}')
        if has_transactions:
            self.play_sound('error')
            self.notify('Impossible: Ce client/fournisseur a des opérations !', 'error')
            self.dialog = MDDialog(title='Suppression Impossible', text=f"Le client '{self.fix_text(entity['name'])}' ne peut pas être supprimé car il possède un historique d'opérations.\n\nVeuillez supprimer ses opérations d'abord.", buttons=[MDFlatButton(text='OK', on_release=lambda x: self.dialog.dismiss())])
            self.dialog.open()
            return

        def do_delete(x):
            if hasattr(self, 'del_conf_dialog') and self.del_conf_dialog:
                self.del_conf_dialog.dismiss()
            try:
                self.db.delete_entity(entity['id'], self.current_entity_type_mgmt)
                self.notify('Supprimé avec succès', 'success')
                self.load_local_entities(self.current_entity_type_mgmt)
            except Exception as e:
                self.notify(f'Erreur suppresssion: {e}', 'error')
        name_display = self.fix_text(entity['name'])
        self.del_conf_dialog = MDDialog(title='Confirmation', text=f'Voulez-vous vraiment supprimer {name_display} ?\nCette action est irréversible.', buttons=[MDFlatButton(text='NON', on_release=lambda x: self.del_conf_dialog.dismiss()), MDRaisedButton(text='OUI, SUPPRIMER', md_bg_color=(1, 0, 0, 1), on_release=do_delete)])
        self.del_conf_dialog.open()

    def _auto_login_check(self, dt):
        self.check_and_load_stats()
        if hasattr(self, 'password_field'):
            self.password_field.text = ''

    def do_login(self, x):
        self.notify('Connexion en cours...', 'info')
        self.current_user_name = 'ADMIN'
        p = self.password_field.get_value()
        if p is None:
            p = ''
        user = self.db.login(p)
        if user:
            self.db.set_setting('cred_password', p)
            self.sm.current = 'dashboard'
            self.load_more_products(reset=True)
            self.check_and_load_stats()
            self.notify(f'Bienvenue', 'success')
        else:
            self.notify('Mot de passe incorrect', 'error')
            self.play_sound('error')

    def show_change_password_dialog_login(self, instance):
        content = MDBoxLayout(orientation='vertical', spacing='15dp', size_hint_y=None, height=dp(280), padding=[0, 10, 0, 0])
        old_pass = MDTextField(hint_text='Mot de passe actuel', password=True)
        new_pass = MDTextField(hint_text='Nouveau mot de passe', password=True)
        conf_pass = MDTextField(hint_text='Confirmer le mot de passe', password=True)
        content.add_widget(old_pass)
        content.add_widget(new_pass)
        content.add_widget(conf_pass)

        def confirm_change(x):
            current_p = old_pass.text if old_pass.text else ''
            new_p = new_pass.text if new_pass.text else ''
            conf_p = conf_pass.text if conf_pass.text else ''
            if not new_p:
                self.notify('Le nouveau mot de passe est vide', 'error')
                new_pass.error = True
                return
            if new_p != conf_p:
                self.notify('Les mots de passe ne correspondent pas', 'error')
                new_pass.error = True
                conf_pass.error = True
                return
            user = self.db.login(current_p)
            if user:
                self.db.update_admin_password(new_p)
                self.db.set_setting('cred_password', '')
                self.password_field.text = ''
                self.notify('Mot de passe modifié avec succès', 'success')
                self.pass_dialog.dismiss()
            else:
                self.notify('Mot de passe actuel incorrect', 'error')
                old_pass.error = True
        self.pass_dialog = MDDialog(title='Changer le mot de passe', type='custom', content_cls=content, buttons=[MDFlatButton(text='ANNULER', on_release=lambda x: self.pass_dialog.dismiss()), MDRaisedButton(text='CONFIRMER', md_bg_color=(0, 0.6, 0, 1), text_color=(1, 1, 1, 1), on_release=confirm_change)])
        self.pass_dialog.open()

    def logout(self):

        def perform_exit(x):
            if hasattr(self, 'logout_diag') and self.logout_diag:
                self.logout_diag.dismiss()
            self.stop()
        self.logout_diag = MDDialog(title='Fermeture', text="Voulez-vous vraiment fermer l'application ?", buttons=[MDFlatButton(text='ANNULER', on_release=lambda x: self.logout_diag.dismiss()), MDRaisedButton(text='FERMER', md_bg_color=(0.8, 0, 0, 1), text_color=(1, 1, 1, 1), on_release=perform_exit)])
        self.logout_diag.open()

    def _create_stat_item(self, title, ref_name, color):
        box = MDBoxLayout(orientation='vertical', padding=dp(5), md_bg_color=(1, 1, 1, 1), radius=[5])
        box.add_widget(MDLabel(text=title, font_style='Caption', halign='center'))
        val_lbl = MDLabel(text='0.00 DA', font_style='Subtitle2', bold=True, halign='center', theme_text_color='Custom', text_color=color)
        setattr(self, ref_name, val_lbl)
        box.add_widget(val_lbl)
        return box

    def _create_dash_btn(self, icon, text, bg_color, icon_color, action):
        card = MDCard(orientation='vertical', padding=dp(15), radius=[12], ripple_behavior=True, on_release=action, md_bg_color=bg_color, elevation=2, size_hint_y=None, height=dp(110))
        card.add_widget(MDIcon(icon=icon, font_size='38sp', pos_hint={'center_x': 0.5}, theme_text_color='Custom', text_color=icon_color))
        card.add_widget(MDLabel(text=text, halign='center', bold=True, font_style='Caption'))
        return card

    def _build_login_screen(self):
        screen = MDScreen(name='login')
        layout = MDFloatLayout()
        card = MDCard(orientation='vertical', size_hint=(0.85, None), height=dp(320), pos_hint={'center_x': 0.5, 'center_y': 0.5}, padding=dp(20), spacing=dp(15), radius=[20], elevation=4)
        icon_box = MDFloatLayout(size_hint_y=None, height=dp(70))
        icon_box.add_widget(MDIcon(icon='lock-open-variant', font_size='60sp', pos_hint={'center_x': 0.5, 'center_y': 0.5}, theme_text_color='Primary'))
        card.add_widget(icon_box)
        card.add_widget(MDLabel(text='Authentification', halign='center', font_style='H5', bold=True))
        self.password_field = SmartTextField(hint_text='Mot de passe', password=True, icon_right='key', halign='center')
        card.add_widget(self.password_field)
        card.add_widget(MDFillRoundFlatButton(text='SE CONNECTER', font_size='18sp', size_hint_x=1, on_release=self.do_login))
        layout.add_widget(card)
        btn_settings = MDFlatButton(text='Changer / Définir Mot de passe', pos_hint={'center_x': 0.5, 'y': 0.1}, theme_text_color='Custom', text_color=(0.4, 0.4, 0.4, 1), on_release=self.show_change_password_dialog_login)
        layout.add_widget(btn_settings)
        footer_label = MDLabel(text='MagPro v7.1.0 © 2026', halign='center', pos_hint={'center_x': 0.5, 'y': 0.02}, size_hint_y=None, height=dp(20), font_style='Caption', theme_text_color='Hint')
        layout.add_widget(footer_label)
        screen.add_widget(layout)
        return screen

    def _build_dashboard_screen(self):
        screen = MDScreen(name='dashboard')
        layout = MDBoxLayout(orientation='vertical')
        self.dash_toolbar = MDTopAppBar(title='Accueil', left_action_items=[['clipboard-text-clock', lambda x: self.show_pending_dialog()]], right_action_items=[['cog', lambda x: self.open_settings_menu()], ['logout', lambda x: self.logout()]])
        layout.add_widget(self.dash_toolbar)
        scroll = MDScrollView()
        self.main_dash_content = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(20), padding=dp(15))
        self.buttons_container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(15))
        self.main_dash_content.add_widget(self.buttons_container)
        self.stats_card_container = MDCard(orientation='vertical', size_hint_y=None, height=dp(280), padding=dp(10), radius=[10], elevation=2, md_bg_color=(0.97, 0.97, 0.97, 1))
        self.main_dash_content.add_widget(self.stats_card_container)
        scroll.add_widget(self.main_dash_content)
        layout.add_widget(scroll)
        screen.add_widget(layout)
        return screen

    def show_store_settings_dialog(self, *args):
        if hasattr(self, 'settings_menu_dialog') and self.settings_menu_dialog:
            self.settings_menu_dialog.dismiss()
        s_name = self.db.get_setting('store_name', '')
        s_phone = self.db.get_setting('store_phone', '')
        s_address = self.db.get_setting('store_address', '')
        s_activity = self.db.get_setting('store_activity', '')
        s_email = self.db.get_setting('store_email', '')
        s_rc = self.db.get_setting('store_rc', '')
        s_nif = self.db.get_setting('store_nif', '')
        s_nis = self.db.get_setting('store_nis', '')
        s_nai = self.db.get_setting('store_nai', '')
        scroll = MDScrollView(size_hint_y=None, height=dp(600))
        main_box = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(15), padding=[dp(10), dp(10), dp(10), dp(20)])
        card_info = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.99, 0.99, 0.99, 1))
        header_info = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
        header_info.add_widget(MDIcon(icon='store', theme_text_color='Primary', font_size='22sp'))
        header_info.add_widget(MDLabel(text='Identité du Magasin', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
        card_info.add_widget(header_info)
        card_info.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
        f_name = SmartTextField(text=s_name, hint_text='Nom du Magasin', icon_right='domain')
        f_activity = SmartTextField(text=s_activity, hint_text='Activité', icon_right='briefcase')
        card_info.add_widget(f_name)
        card_info.add_widget(f_activity)
        main_box.add_widget(card_info)
        card_contact = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.96, 0.98, 1, 1))
        header_contact = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
        header_contact.add_widget(MDIcon(icon='card-account-phone', theme_text_color='Primary', font_size='22sp'))
        header_contact.add_widget(MDLabel(text='Coordonnées', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
        card_contact.add_widget(header_contact)
        card_contact.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
        f_phone = SmartTextField(text=s_phone, hint_text='Téléphone', input_filter='int', icon_right='phone')
        f_address = SmartTextField(text=s_address, hint_text='Adresse', icon_right='map-marker')
        f_email = SmartTextField(text=s_email, hint_text='Email', icon_right='email')
        card_contact.add_widget(f_phone)
        card_contact.add_widget(f_address)
        card_contact.add_widget(f_email)
        main_box.add_widget(card_contact)
        card_fisc = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.95, 0.95, 0.95, 1))
        header_fisc = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
        header_fisc.add_widget(MDIcon(icon='file-document-multiple', theme_text_color='Primary', font_size='22sp'))
        header_fisc.add_widget(MDLabel(text='Information Fiscale', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
        card_fisc.add_widget(header_fisc)
        card_fisc.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
        f_rc = SmartTextField(text=s_rc, hint_text='N° Registre Commerce (RC)')
        f_nif = SmartTextField(text=s_nif, hint_text='N.I.F')
        f_nis = SmartTextField(text=s_nis, hint_text='N.I.S')
        f_nai = SmartTextField(text=s_nai, hint_text='N.A.I')
        card_fisc.add_widget(f_rc)
        card_fisc.add_widget(f_nif)
        card_fisc.add_widget(f_nis)
        card_fisc.add_widget(f_nai)
        main_box.add_widget(card_fisc)
        footer_box = MDBoxLayout(orientation='vertical', spacing=dp(10), adaptive_height=True, padding=[0, dp(10), 0, 0])

        def save_info(x):
            self.db.set_setting('store_name', f_name.get_value().strip())
            self.db.set_setting('store_phone', f_phone.get_value().strip())
            self.db.set_setting('store_address', f_address.get_value().strip())
            self.db.set_setting('store_activity', f_activity.get_value().strip())
            self.db.set_setting('store_email', f_email.get_value().strip())
            self.db.set_setting('store_rc', f_rc.get_value().strip())
            self.db.set_setting('store_nif', f_nif.get_value().strip())
            self.db.set_setting('store_nis', f_nis.get_value().strip())
            self.db.set_setting('store_nai', f_nai.get_value().strip())
            self.notify('Enregistré avec succès', 'success')
            self.store_settings_dialog.dismiss()
            self.open_settings_menu()
        btn_save = MDRaisedButton(text='ENREGISTRER', md_bg_color=(0, 0.7, 0, 1), text_color=(1, 1, 1, 1), size_hint_x=1, height=dp(50), elevation=2, on_release=save_info)
        btn_cancel = MDRaisedButton(text='ANNULER', md_bg_color=(0.9, 0.9, 0.9, 1), text_color=(0.3, 0.3, 0.3, 1), size_hint_x=1, height=dp(50), elevation=0, on_release=lambda x: [self.store_settings_dialog.dismiss(), self.open_settings_menu()])
        footer_box.add_widget(btn_save)
        footer_box.add_widget(btn_cancel)
        main_box.add_widget(footer_box)
        scroll.add_widget(main_box)
        self.store_settings_dialog = MDDialog(title='Informations du Magasin', type='custom', content_cls=scroll, buttons=[], size_hint=(0.98, 0.96))
        self.store_settings_dialog.open()

    def open_settings_menu(self, *args):
        try:
            if hasattr(self, 'settings_menu_dialog') and self.settings_menu_dialog:
                self.settings_menu_dialog.dismiss()
            scroll_view = MDScrollView(size_hint_y=None, height=dp(600))
            content_list = MDList()

            def add_section(text):
                lbl = MDLabel(text=text, theme_text_color='Custom', text_color=self.theme_cls.primary_color, font_style='Subtitle2', bold=True, size_hint_y=None, height=dp(40), padding=(dp(20), dp(10)))
                content_list.add_widget(lbl)

            def add_option(title, details, icon_name, action_callback, icon_color=None):
                item = TwoLineAvatarIconListItem(text=title, secondary_text=details, on_release=action_callback)
                icon = IconLeftWidget(icon=icon_name)
                if icon_color:
                    icon.text_color = icon_color
                item.add_widget(icon)
                content_list.add_widget(item)
            add_section('DONNÉES & SAUVEGARDE')
            add_option('Exporter Produits', 'Direct Excel (.xlsx)', 'file-export', lambda x: self.perform_export())
            add_option('Importer Produits', 'Depuis Excel (.xlsx)', 'file-import', lambda x: [self.settings_menu_dialog.dismiss(), self.import_data_dialog()])
            add_option('Sauvegarde Locale', 'Backup complet (.db)', 'database-export', lambda x: self.perform_local_backup())
            add_option('Sauvegarde Cloud', 'Partager (Drive, Email...)', 'cloud-upload', lambda x: self.share_database_file())
            add_option('Restaurer', 'Depuis une sauvegarde', 'backup-restore', lambda x: [self.settings_menu_dialog.dismiss(), self.show_restore_dialog()])
            add_section('MAGASIN')
            add_option('Info Magasin', 'Nom, Adresse, Entête...', 'store', lambda x: self.show_store_settings_dialog(x))
            add_section('IMPRIMANTE & PDF')
            printer_name = self.db.get_setting('printer_name', 'Non configurée')
            add_option('Configurer', f'Actuelle: {printer_name}', 'printer-wireless', lambda x: [self.settings_menu_dialog.dismiss(), self.open_bluetooth_selector(x)])
            add_option("Oublier l'imprimante", 'Déconnecter', 'printer-off', lambda x: self.clear_printer_selection(x), icon_color=(0.8, 0, 0, 1))
            auto_layout = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(50), padding=(dp(20), 0))
            lbl_auto = MDLabel(text='Impression Auto (Bluetooth)', theme_text_color='Primary', size_hint_x=0.8)
            is_auto = self.db.get_setting('printer_auto', 'False') == 'True'
            chk_auto = MDCheckbox(active=is_auto, size_hint=(None, None), size=(dp(40), dp(40)), pos_hint={'center_y': 0.5})
            chk_auto.bind(active=self.toggle_auto_print)
            auto_layout.add_widget(lbl_auto)
            auto_layout.add_widget(chk_auto)
            content_list.add_widget(auto_layout)
            bal_layout = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(50), padding=(dp(20), 0))
            lbl_bal = MDLabel(text='Afficher Solde PDF (BV/BA)', theme_text_color='Primary', size_hint_x=0.8)
            is_bal = self.db.get_setting('show_balance_in_pdf', 'False') == 'True'
            chk_bal = MDCheckbox(active=is_bal, size_hint=(None, None), size=(dp(40), dp(40)), pos_hint={'center_y': 0.5})
            chk_bal.bind(active=self.toggle_pdf_balance)
            bal_layout.add_widget(lbl_bal)
            bal_layout.add_widget(chk_bal)
            content_list.add_widget(bal_layout)
            add_section('SÉCURITÉ')
            add_option('Admin / Vendeur', 'Gérer les accès', 'shield-account', lambda x: [self.settings_menu_dialog.dismiss(), self.open_seller_auth_dialog(x)])
            scroll_view.add_widget(content_list)
            self.settings_menu_dialog = MDDialog(title='Paramètres', type='custom', content_cls=scroll_view, buttons=[MDFlatButton(text='FERMER', theme_text_color='Custom', text_color=self.theme_cls.primary_color, on_release=lambda x: self.settings_menu_dialog.dismiss())], size_hint=(0.96, None))
            self.settings_menu_dialog.open()
        except Exception as e:
            self.notify(f'Erreur Menu: {e}', 'error')

    def toggle_auto_print(self, instance, value):
        self.db.set_setting('printer_auto', str(value))
        msg = 'Activée' if value else 'Désactivée'

    def toggle_pdf_balance(self, instance, value):
        self.db.set_setting('show_balance_in_pdf', str(value))

    def _build_products_screen(self):
        screen = MDScreen(name='products')
        root_layout = MDFloatLayout()
        self.rv_products = ProductRecycleView()
        self.rv_products.size_hint = (1, 1)
        self.rv_products.pos_hint = {'center_x': 0.5, 'center_y': 0.5}

        def set_rv_padding(dt):
            if hasattr(self.rv_products, 'layout_manager') and self.rv_products.layout_manager:
                self.rv_products.layout_manager.padding = [dp(5), dp(205), dp(5), dp(70)]
        Clock.schedule_once(set_rv_padding, 0.1)
        root_layout.add_widget(self.rv_products)
        header_container = MDBoxLayout(orientation='vertical', adaptive_height=True, md_bg_color=(1, 1, 1, 1))
        header_container.pos_hint = {'top': 1}
        self.prod_toolbar = MDTopAppBar(title='Produits', left_action_items=[['arrow-left', lambda x: self.go_back()]], elevation=2)
        header_container.add_widget(self.prod_toolbar)
        self.prod_search_layout = MDBoxLayout(orientation='horizontal', padding=[dp(10), dp(5), dp(10), dp(0)], spacing=dp(5), size_hint_y=None, height=dp(55), md_bg_color=(1, 1, 1, 1))
        self.search_field = SmartTextField(hint_text='Rechercher...', mode='rectangle', icon_right='magnify', size_hint_y=None, height=dp(40), pos_hint={'center_y': 0.5}, size_hint_x=1)
        self.search_field.bind(text=self.filter_products)
        self.prod_search_layout.add_widget(self.search_field)
        self.btn_scan_prod = MDIconButton(icon='barcode-scan', theme_text_color='Custom', text_color=(0.2, 0.2, 0.2, 1), pos_hint={'center_y': 0.5}, icon_size='28sp', on_release=self.open_barcode_scanner)
        self.prod_search_layout.add_widget(self.btn_scan_prod)
        self.btn_add_prod = MDIconButton(icon='plus-circle', theme_text_color='Custom', text_color=(0, 0.7, 0, 1), pos_hint={'center_y': 0.5}, icon_size='36sp', on_release=lambda x: self.show_manage_product_dialog(None))
        self.prod_search_layout.add_widget(self.btn_add_prod)
        header_container.add_widget(self.prod_search_layout)
        header_container.add_widget(MDBoxLayout(size_hint_y=None, height=dp(8)))
        self.selected_family_filter = 'TOUS'
        filter_box = MDBoxLayout(orientation='vertical', size_hint_y=None, height=dp(55), padding=[dp(10), dp(0), dp(10), dp(5)])
        family_filter_card = MDCard(size_hint=(1, 1), radius=[4], md_bg_color=(0.96, 0.96, 0.96, 1), line_color=(0.5, 0.5, 0.5, 1), line_width=1.1, elevation=1, ripple_behavior=False)
        self.btn_main_family_filter = MDFlatButton(text='TOUS', theme_text_color='Custom', text_color=(0.15, 0.15, 0.15, 1), font_style='Subtitle1', size_hint=(1, 1), pos_hint={'center_x': 0.5, 'center_y': 0.5}, on_release=self.open_filter_menu)
        family_filter_card.add_widget(self.btn_main_family_filter)
        filter_box.add_widget(family_filter_card)
        header_container.add_widget(filter_box)
        from kivymd.uix.card import MDSeparator
        header_container.add_widget(MDSeparator(height=dp(1), color=(0.9, 0.9, 0.9, 1)))
        root_layout.add_widget(header_container)
        self.cart_bar = MDCard(size_hint_y=None, height=dp(60), padding=[dp(15), dp(5)], md_bg_color=self.theme_cls.primary_color, radius=[15, 15, 0, 0], ripple_behavior=True, on_release=self.open_cart_screen, elevation=4, pos_hint={'bottom': 1})
        cart_box = MDBoxLayout(orientation='horizontal')
        self.lbl_cart_count = MDLabel(text='PANIER (0)', theme_text_color='Custom', text_color=(1, 1, 1, 1), bold=True, halign='left', size_hint_x=0.5, font_style='Subtitle1')
        self.lbl_cart_total = MDLabel(text='0.00 DA', theme_text_color='Custom', text_color=(1, 1, 1, 1), bold=True, halign='right', font_style='H6', size_hint_x=0.5)
        cart_box.add_widget(self.lbl_cart_count)
        cart_box.add_widget(self.lbl_cart_total)
        self.cart_bar.add_widget(cart_box)
        root_layout.add_widget(self.cart_bar)
        screen.add_widget(root_layout)
        return screen

    def open_filter_menu(self, instance):
        families = self.db.get_families()
        if 'TOUS' in families:
            families.remove('TOUS')
        full_list = ['TOUS'] + families
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, adaptive_height=True, padding=dp(0))
        scroll = MDScrollView(size_hint_y=None, height=dp(300))
        list_layout = MDBoxLayout(orientation='vertical', adaptive_height=True)
        for fam in full_list:
            btn = MDFlatButton(text=fam, theme_text_color='Custom', text_color=(0.2, 0.2, 0.2, 1), font_style='Subtitle1', size_hint_x=1, height=dp(50))
            btn.bind(on_release=lambda x, f=fam: self.apply_filter(f))
            list_layout.add_widget(btn)
            from kivymd.uix.card import MDSeparator
            list_layout.add_widget(MDSeparator(height=dp(1), color=(0.9, 0.9, 0.9, 1)))
        scroll.add_widget(list_layout)
        content.add_widget(scroll)
        self.filter_dialog = MDDialog(title='Filtrer par famille', type='custom', content_cls=content, size_hint=(0.85, None))
        self.filter_dialog.open()

    def apply_filter(self, family_name):
        self.selected_family_filter = family_name
        self.btn_main_family_filter.text = family_name
        self.filter_dialog.dismiss()
        self.load_more_products(reset=True)

    def load_more_products(self, reset=False):
        if self.is_loading_more and (not reset):
            return
        if reset:
            self.current_page_offset = 0
            self.is_loading_more = False
            if self.rv_products:
                self.rv_products.scroll_y = 1.0
                self.rv_products.data = []
                self.rv_products.refresh_from_data()
        self.is_loading_more = True
        threading.Thread(target=self._load_products_worker, args=(reset,), daemon=True).start()

    def _load_products_worker(self, reset):
        try:
            from datetime import datetime
            limit = 50
            offset = self.current_page_offset
            search_text = ''
            if hasattr(self, 'search_field') and self.search_field.text:
                search_text = self.search_field.text
            family = getattr(self, 'selected_family_filter', 'TOUS')
            products = self.db.get_products(limit=limit, offset=offset, search_query=search_text, family_filter=family)
            allowed_autre_modes = ['sale', 'invoice_sale', 'proforma', 'order_purchase']
            if offset == 0 and self.current_mode in allowed_autre_modes and (not search_text) and (family == 'TOUS'):
                virtual_item = {'id': -999, 'name': 'Autre Article', 'price': 0, 'purchase_price': 0, 'stock': -1000000, 'stock_warehouse': 0, 'barcode': '', 'reference': '', 'is_virtual': True}
                if products is None:
                    products = []
                products.insert(0, virtual_item)
            if not products:
                Clock.schedule_once(lambda dt: self._finish_loading_empty(), 0)
                return
            rv_data = []
            purchase_modes = ['purchase', 'invoice_purchase', 'return_purchase', 'order_purchase', 'bi']
            is_purchase_view = self.current_mode in purchase_modes
            is_transfer = self.current_mode == 'transfer'
            modes_showing_promo = ['sale', 'invoice_sale', 'proforma', 'client_payment', 'manage_products']
            should_show_promo = self.current_mode in modes_showing_promo
            customer_category = 'Détail'
            if self.selected_entity:
                customer_category = str(self.selected_entity.get('category', 'Détail')).strip()

            def fmt_qty(val):
                try:
                    val = float(val)
                    return str(int(val)) if val.is_integer() else str(val)
                except:
                    return '0'
            img_dir = os.path.join(self.user_data_dir, 'product_images')
            for p in products:
                try:
                    s_store = float(p.get('stock', 0) or 0)
                    is_unlimited = s_store <= -900000 or p.get('id') == -999
                    if is_transfer and is_unlimited:
                        continue
                    raw_name = self.fix_text(str(p.get('name', '')))
                    name_display = raw_name
                    s_wh = float(p.get('stock_warehouse', 0) or 0)
                    total_stock = s_store + s_wh
                    price_fmt = ''
                    price_color = [0, 0, 0, 1]
                    stock_text = ''
                    icon = 'package-variant'
                    icon_col = [0, 0.6, 0, 1]
                    raw_img_path = p.get('image_path', '')
                    final_img_path = ''
                    if raw_img_path:
                        if os.path.exists(raw_img_path):
                            final_img_path = raw_img_path
                        else:
                            filename = os.path.basename(raw_img_path)
                            potential_path = os.path.join(img_dir, filename)
                            if os.path.exists(potential_path):
                                final_img_path = potential_path
                    if is_transfer:
                        price_fmt = f'Tot: {fmt_qty(total_stock)}'
                        price_color = [0.2, 0.2, 0.8, 1]
                        stock_text = f'Mag: {fmt_qty(s_store)} | Dép: {fmt_qty(s_wh)}'
                    elif is_purchase_view:
                        price = float(p.get('purchase_price', p.get('price', 0)) or 0)
                        price_fmt = f'{price:.2f} DA'
                        price_color = [0.9, 0.5, 0, 1]
                        if is_unlimited:
                            stock_text = 'Illimité'
                        else:
                            stock_text = f'Qté: {fmt_qty(s_store)}'
                    else:
                        base_price = float(p.get('price', 0) or 0)
                        final_display_price = base_price
                        if customer_category == 'Gros':
                            wh_price = float(p.get('price_wholesale', 0) or 0)
                            if wh_price > 0:
                                final_display_price = wh_price
                        elif customer_category == 'Demi-Gros':
                            semi_price = float(p.get('price_semi', 0) or 0)
                            if semi_price > 0:
                                final_display_price = semi_price
                        has_promo = False
                        if should_show_promo:
                            raw_active = p.get('is_promo_active', 0)
                            is_active = str(raw_active) == '1' or raw_active == 1
                            if is_active:
                                promo_exp = str(p.get('promo_expiry', '')).strip()
                                date_valid = True
                                if promo_exp and len(promo_exp) > 5:
                                    try:
                                        exp_date = datetime.strptime(promo_exp, '%Y-%m-%d').date()
                                        if datetime.now().date() > exp_date:
                                            date_valid = False
                                    except:
                                        pass
                                if date_valid:
                                    has_promo = True
                                    p_type = str(p.get('promo_type', 'fixed'))
                                    try:
                                        p_val = float(p.get('promo_value', 0))
                                    except:
                                        p_val = 0.0
                                    if p_type == 'fixed':
                                        if p_val > 0:
                                            final_display_price = p_val
                                    else:
                                        final_display_price = base_price * (1 - p_val / 100)
                        price_fmt = f'{final_display_price:.2f} DA'
                        if has_promo:
                            price_color = [0, 0.2, 0.8, 1]
                            icon = 'sale'
                            icon_col = [0.9, 0.1, 0.1, 1]
                        else:
                            price_color = [0, 0.6, 0, 1]
                        if is_unlimited:
                            stock_text = 'Illimité'
                            if p.get('id') == -999:
                                price_fmt = 'Prix Libre'
                                price_color = [0, 0.4, 0.8, 1]
                                icon = 'pencil-plus'
                                icon_col = [0, 0.4, 0.8, 1]
                            elif not has_promo:
                                icon = 'package-variant'
                        elif s_wh == 0:
                            stock_text = f'Qté: {fmt_qty(s_store)}'
                            if not has_promo:
                                icon = 'package-variant' if s_store > 0 else 'package-variant-closed'
                                icon_col = [0, 0.6, 0, 1] if s_store > 0 else [0.8, 0, 0, 1]
                        else:
                            stock_text = f'Qté: {fmt_qty(s_store)} | Dép: {fmt_qty(s_wh)}'
                            if not has_promo:
                                icon = 'package-variant' if total_stock > 0 else 'package-variant-closed'
                                icon_col = [0, 0.6, 0, 1] if total_stock > 0 else [0.8, 0, 0, 1]
                    rv_data.append({'text_name': name_display, 'text_price': price_fmt, 'text_stock': stock_text, 'icon_name': icon, 'icon_color': icon_col, 'price_color': price_color, 'image_path': final_img_path, 'raw_data': p})
                except Exception as ex:
                    print(f'Error processing item: {ex}')
                    continue
            Clock.schedule_once(lambda dt: self._append_to_rv(rv_data, reset), 0)
        except Exception as e:
            print(f'Worker Error: {e}')
            import traceback
            traceback.print_exc()
            Clock.schedule_once(lambda dt: self._finish_loading_empty(), 0)

    def _finish_loading_empty(self):
        self.is_loading_more = False
        if self.rv_products:
            self.rv_products.loading_lock = False

    def _build_cart_screen(self):
        screen = MDScreen(name='cart')
        layout = MDBoxLayout(orientation='vertical')
        self.cart_toolbar = MDTopAppBar(title='Panier', left_action_items=[['arrow-left', lambda x: self.back_to_products()]])
        layout.add_widget(self.cart_toolbar)
        selectors = MDCard(orientation='horizontal', size_hint_y=None, height=dp(70), padding=dp(10), radius=0, md_bg_color=(0.95, 0.95, 0.95, 1))
        self.btn_ent_screen = MDFillRoundFlatButton(text='Client', size_hint_x=0.45, on_release=self.handle_entity_button_click)
        self.btn_loc_screen = MDFillRoundFlatButton(text='Magasin', size_hint_x=0.45, on_release=self.toggle_location)
        selectors.add_widget(self.btn_ent_screen)
        selectors.add_widget(MDBoxLayout(size_hint_x=0.1))
        selectors.add_widget(self.btn_loc_screen)
        layout.add_widget(selectors)
        self.rv_cart = CartRecycleView()
        layout.add_widget(self.rv_cart)
        self.footer_card = MDCard(orientation='vertical', size_hint_y=None, height=dp(150), padding=dp(15), spacing=dp(10), radius=[20, 20, 0, 0], elevation=4)
        total_row = MDBoxLayout(orientation='horizontal')
        self.lbl_total_title = MDLabel(text='TOTAL:', bold=True, font_style='Subtitle1')
        self.lbl_cart_screen_total = MDLabel(text='0.00 DA', halign='right', font_style='H5', bold=True, theme_text_color='Primary')
        total_row.add_widget(self.lbl_total_title)
        total_row.add_widget(self.lbl_cart_screen_total)
        self.btn_validate_cart = MDFillRoundFlatButton(text='VALIDER LA COMMANDE', size_hint_x=1, height=dp(55), md_bg_color=(0, 0.7, 0, 1), on_release=self.open_payment_dialog)
        self.footer_card.add_widget(total_row)
        self.footer_card.add_widget(self.btn_validate_cart)
        layout.add_widget(self.footer_card)
        screen.add_widget(layout)
        return screen

    def open_cart_screen(self, x=None):
        if not self.cart:
            self.dialog = MDDialog(title='Panier vide', text='Veuillez ajouter au moins un produit pour continuer.', buttons=[MDFlatButton(text='OK', on_release=lambda x: self.dialog.dismiss())])
            self.dialog.open()
            return
        if self.current_mode != 'transfer' and self.selected_entity is None:
            self.show_entity_selection_dialog(None, next_action=lambda: self.open_cart_screen(None))
            return
        self.refresh_cart_screen_items()
        self.sm.transition.direction = 'left'
        self.sm.current = 'cart'

    def back_to_products(self):
        self.sm.transition.direction = 'right'
        self.sm.current = 'products'

    def handle_entity_button_click(self, instance):
        if self.current_mode == 'transfer':
            self.toggle_location(instance)
        else:
            self.show_entity_selection_dialog(instance)

    def refresh_cart_screen_items(self):
        doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
        doc_type = doc_type_map.get(self.current_mode, 'BV')
        stock_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
        fin_f = AppConstants.FINANCIAL_FACTORS.get(doc_type, 0)
        is_invoice_mode = doc_type in ['FC', 'FF', 'FP']
        is_transfer = stock_f == 0 and fin_f == 0 and (doc_type in ['TR', 'TRANSFER'])
        should_hide_location = is_transfer or doc_type in ['FP', 'DP', 'PROFORMA', 'ORDER']
        total_ht, total_tva = self.calculate_cart_totals(self.cart, is_invoice_mode)
        timbre = Decimal('0.00')
        total_ttc = total_ht + total_tva + timbre
        total_ttc = quantize_decimal(total_ttc)
        items_count = len(self.cart)
        if hasattr(self, 'cart_toolbar'):
            self.cart_toolbar.title = f'Panier ({items_count})'
        if hasattr(self, 'lbl_cart_screen_total'):
            if is_transfer:
                self.lbl_cart_screen_total.text = ''
                self.lbl_cart_screen_total.opacity = 0
            else:
                self.lbl_cart_screen_total.text = f'{total_ttc:,.2f} DA'.replace(',', ' ').replace('.', ',')
                self.lbl_cart_screen_total.opacity = 1
        if hasattr(self, 'lbl_total_title'):
            if is_transfer:
                self.lbl_total_title.text = 'Résumé:'
            else:
                self.lbl_total_title.text = 'TOTAL:'
        self.update_location_display()
        if hasattr(self, 'btn_ent_screen'):
            if is_transfer:
                src = 'Magasin' if self.selected_location == 'store' else 'Dépôt'
                dst = 'Dépôt' if self.selected_location == 'store' else 'Magasin'
                self.btn_ent_screen.text = f'{src}  >>>  {dst}'
                self.btn_ent_screen.md_bg_color = (0.5, 0, 0.5, 1)
            else:
                is_supp_mode = stock_f == 1 or 'SUPPLIER' in doc_type
                if self.selected_entity:
                    self.btn_ent_screen.text = self.fix_text(self.selected_entity.get('name', 'Tiers'))[:15]
                vis = AppConstants.DOC_VISUALS.get(doc_type, {'color': (0, 0.6, 0.6, 1)})
                self.btn_ent_screen.md_bg_color = vis['color']
            if should_hide_location:
                self.btn_ent_screen.size_hint_x = 0.95
                if hasattr(self, 'btn_loc_screen'):
                    self.btn_loc_screen.opacity = 0
                    self.btn_loc_screen.disabled = True
                    self.btn_loc_screen.size_hint_x = 0
                    self.btn_loc_screen.width = 0
            else:
                self.btn_ent_screen.size_hint_x = 0.45
                if hasattr(self, 'btn_loc_screen'):
                    self.btn_loc_screen.opacity = 1
                    self.btn_loc_screen.disabled = False
                    self.btn_loc_screen.size_hint_x = 0.45
        rv_data = []
        for item in self.cart:
            try:
                p = to_decimal(item.get('price', 0))
                q = to_decimal(item.get('qty', 0))
                t_rate = to_decimal(item.get('tva', 0)) if is_invoice_mode else Decimal('0.00')
                line_ht = quantize_decimal(p * q)
                line_ttc = quantize_decimal(line_ht * (Decimal('1') + t_rate / Decimal('100')))
                q_disp = str(int(q)) if q % 1 == 0 else str(float(q))
                if is_transfer:
                    details_text = f'Qté: {q_disp}'
                    d_color = [0.1, 0.4, 0.8, 1]
                else:
                    details_text = f'{p:,.2f} DA x {q_disp}'
                    if t_rate > 0 and is_invoice_mode:
                        details_text += f' (+{int(t_rate)}% TVA)'
                    details_text += f' = {line_ttc:,.2f} DA'
                    d_color = [0.4, 0.4, 0.4, 1]
                rv_data.append({'name': item.get('name', 'Produit'), 'details': details_text, 'd_color': d_color, 'raw_item': item})
            except Exception as e:
                print(f'Error processing cart item: {e}')
        if hasattr(self, 'rv_cart'):
            self.rv_cart.data = rv_data
            self.rv_cart.refresh_from_data()
        if hasattr(self, 'total_bg_card'):
            self.total_bg_card.opacity = 0 if is_transfer else 1
        if hasattr(self, 'btn_validate_cart'):
            if is_transfer:
                self.btn_validate_cart.text = 'VALIDER LE TRANSFERT'
            else:
                self.btn_validate_cart.text = 'VALIDER LA COMMANDE'

    def edit_cart_item(self, item):

        def fmt_num(value):
            try:
                val_float = float(value)
                return str(int(val_float)) if val_float.is_integer() else str(val_float)
            except:
                return '0'
        doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
        doc_type = doc_type_map.get(self.current_mode, 'BV')
        fin_f = AppConstants.FINANCIAL_FACTORS.get(doc_type, 0)
        is_invoice = doc_type in ['FC', 'FF', 'FP']
        hide_price = fin_f == 0 and doc_type != 'FP'
        is_virtual = item.get('id') == -999 or str(item.get('name')).startswith('Autre Article')
        base_height = 600 if is_invoice else 520
        dialog_height = dp(base_height + 20) if is_virtual else dp(base_height)
        content = MDBoxLayout(orientation='vertical', spacing='10dp', size_hint_y=None, height=dialog_height, padding=[0, '5dp', 0, 0])
        self.active_edit_target = 'qty'
        self.input_reset_mode = True
        self.name_cleared_once = False

        def update_edit_colors():
            ACTIVE_BG = (0.9, 1, 0.9, 1)
            INACTIVE_BG = (0.95, 0.95, 0.95, 1)
            if hasattr(self, 'edit_qty_card'):
                self.edit_qty_card.md_bg_color = ACTIVE_BG if self.active_edit_target == 'qty' else INACTIVE_BG
                self.edit_qty_card.elevation = 3 if self.active_edit_target == 'qty' else 0
            if hasattr(self, 'edit_price_card'):
                self.edit_price_card.md_bg_color = ACTIVE_BG if self.active_edit_target == 'price' else INACTIVE_BG
                self.edit_price_card.elevation = 3 if self.active_edit_target == 'price' else 0
            if hasattr(self, 'edit_tva_card'):
                self.edit_tva_card.md_bg_color = ACTIVE_BG if self.active_edit_target == 'tva' else INACTIVE_BG
                self.edit_tva_card.elevation = 3 if self.active_edit_target == 'tva' else 0
            if hasattr(self, 'edit_name_card'):
                self.edit_name_card.md_bg_color = ACTIVE_BG if self.active_edit_target == 'name' else INACTIVE_BG
                self.edit_name_card.elevation = 3 if self.active_edit_target == 'name' else 0
        raw_name = item.get('name', 'Produit')
        if is_virtual:
            self.edit_name_card = MDCard(size_hint_y=None, height='70dp', radius=[10], padding=[10, 0, 10, 0], elevation=0)
            self.edit_name_field = SmartTextField(text=self.fix_text(raw_name), hint_text='Nom', font_size='22sp', halign='center', mode='line', line_color_normal=(0, 0, 0, 0), line_color_focus=(0, 0, 0, 0), pos_hint={'center_y': 0.5})

            def on_name_touch(instance, touch):
                if instance.collide_point(*touch.pos):
                    self.active_edit_target = 'name'
                    update_edit_colors()
                    if not self.name_cleared_once:
                        self.old_name_temp = instance.text
                    return False
                return False
            self.edit_name_field.bind(on_touch_down=on_name_touch)
            self.edit_name_card.add_widget(self.edit_name_field)
            name_row = MDBoxLayout(size_hint_y=None, height='75dp', padding=[20, 0, 20, 0])
            name_row.add_widget(self.edit_name_card)
            content.add_widget(name_row)
        else:
            lbl_prod = MDLabel(text=self.fix_text(raw_name), halign='center', bold=True, font_style='Subtitle1', theme_text_color='Primary', adaptive_height=True)
            content.add_widget(lbl_prod)
        if not hide_price:
            price_val = item.get('price', 0)
            self.edit_price_card = MDCard(size_hint_y=None, height='70dp', radius=[10], padding=[10, 0, 10, 0], elevation=0)
            self.edit_price_field = NoMenuTextField(text=fmt_num(price_val), hint_text='Prix (DA)', font_size='26sp', halign='center', mode='line', readonly=True, line_color_normal=(0, 0, 0, 0), line_color_focus=(0, 0, 0, 0), pos_hint={'center_y': 0.5})
            self.edit_price_field.theme_text_color = 'Custom'
            self.edit_price_field.text_color_normal = (0, 0, 0, 1)
            self.edit_price_field.text_color_focus = (0, 0, 0, 1)

            def on_price_touch(instance, touch):
                if instance.collide_point(*touch.pos):
                    if self.active_edit_target != 'price':
                        self.input_reset_mode = True
                    self.active_edit_target = 'price'
                    update_edit_colors()
                    return True
                return False
            self.edit_price_field.bind(on_touch_down=on_price_touch)
            self.edit_price_card.add_widget(self.edit_price_field)
            price_row = MDBoxLayout(size_hint_y=None, height='75dp', padding=[60, 0, 60, 0])
            price_row.add_widget(self.edit_price_card)
            content.add_widget(price_row)
        if is_invoice:
            tva_val = item.get('tva', 0)
            self.edit_tva_card = MDCard(size_hint_y=None, height='60dp', radius=[10], padding=[10, 0, 10, 0], elevation=0)
            self.edit_tva_field = NoMenuTextField(text=fmt_num(tva_val), hint_text='TVA %', font_size='24sp', halign='center', mode='line', readonly=True, line_color_normal=(0, 0, 0, 0), line_color_focus=(0, 0, 0, 0), pos_hint={'center_y': 0.5})
            self.edit_tva_field.theme_text_color = 'Custom'
            self.edit_tva_field.text_color_normal = (0, 0, 0, 1)
            self.edit_tva_field.text_color_focus = (0, 0, 0, 1)

            def on_tva_touch(instance, touch):
                if instance.collide_point(*touch.pos):
                    if self.active_edit_target != 'tva':
                        self.input_reset_mode = True
                    self.active_edit_target = 'tva'
                    update_edit_colors()
                    return True
                return False
            self.edit_tva_field.bind(on_touch_down=on_tva_touch)
            self.edit_tva_card.add_widget(self.edit_tva_field)
            tva_row = MDBoxLayout(size_hint_y=None, height='65dp', padding=[100, 0, 100, 0])
            tva_row.add_widget(self.edit_tva_card)
            content.add_widget(tva_row)
        qty_row = MDBoxLayout(orientation='horizontal', spacing='10dp', size_hint_y=None, height='65dp', padding=[40, 0])
        btn_minus = MDIconButton(icon='minus', theme_text_color='Custom', text_color=(1, 1, 1, 1), md_bg_color=(0.9, 0.3, 0.3, 1), pos_hint={'center_y': 0.5}, icon_size='20sp')
        qty_val = item.get('qty', 1)
        self.edit_qty_card = MDCard(size_hint_x=1, size_hint_y=None, height='60dp', radius=[10], padding=[10, 0, 10, 0], elevation=0, pos_hint={'center_y': 0.5})
        self.edit_qty_field = NoMenuTextField(text=fmt_num(qty_val), hint_text='Qté', font_size='28sp', halign='center', readonly=True, mode='line', line_color_normal=(0, 0, 0, 0), line_color_focus=(0, 0, 0, 0), pos_hint={'center_y': 0.5})
        self.edit_qty_field.theme_text_color = 'Custom'
        self.edit_qty_field.text_color_normal = (0, 0, 0, 1)
        self.edit_qty_field.text_color_focus = (0, 0, 0, 1)

        def on_qty_touch(instance, touch):
            if instance.collide_point(*touch.pos):
                if self.active_edit_target != 'qty':
                    self.input_reset_mode = True
                self.active_edit_target = 'qty'
                update_edit_colors()
                return True
            return False
        self.edit_qty_field.bind(on_touch_down=on_qty_touch)
        self.edit_qty_card.add_widget(self.edit_qty_field)
        btn_plus = MDIconButton(icon='plus', theme_text_color='Custom', text_color=(1, 1, 1, 1), md_bg_color=(0.2, 0.7, 0.2, 1), pos_hint={'center_y': 0.5}, icon_size='20sp')
        qty_row.add_widget(btn_minus)
        qty_row.add_widget(self.edit_qty_card)
        qty_row.add_widget(btn_plus)
        content.add_widget(qty_row)
        self.btn_save_edit = MDRaisedButton(text='MODIFIER', md_bg_color=(0, 0.6, 0, 1), text_color=(1, 1, 1, 1), size_hint_x=0.7, size_hint_y=1, font_size='18sp', elevation=3)

        def update_calculations():
            try:
                q = float(self.edit_qty_field.text or 0)
            except:
                q = 0.0
            p = 0.0
            if hasattr(self, 'edit_price_field'):
                try:
                    p = float(self.edit_price_field.text or 0)
                except:
                    p = 0.0
            else:
                p = float(item.get('price', 0))
            tva = 0.0
            if hasattr(self, 'edit_tva_field'):
                try:
                    tva = float(self.edit_tva_field.text or 0)
                except:
                    tva = 0.0
            line_ht = self._round_num(q * p)
            total = self._round_num(line_ht * (1 + tva / 100.0))
            if not hide_price:
                self.btn_save_edit.text = f'MODIFIER\n{total:.2f} DA'
            else:
                self.btn_save_edit.text = 'MODIFIER'

        def change_qty(amount):
            try:
                current = float(self.edit_qty_field.text or 0)
                new_val = current + amount
                if new_val < 1:
                    new_val = 1
                self.edit_qty_field.text = fmt_num(new_val)
                update_calculations()
            except:
                self.edit_qty_field.text = '1'
        btn_plus.bind(on_release=lambda x: change_qty(1))
        btn_minus.bind(on_release=lambda x: change_qty(-1))

        def get_active_field():
            if self.active_edit_target == 'name':
                return None
            if self.active_edit_target == 'price' and hasattr(self, 'edit_price_field'):
                return self.edit_price_field
            if self.active_edit_target == 'tva' and hasattr(self, 'edit_tva_field'):
                return self.edit_tva_field
            return self.edit_qty_field

        def add_digit(digit):
            field = get_active_field()
            if not field:
                return
            current = field.text
            if self.input_reset_mode:
                if digit == '.':
                    field.text = '0.'
                else:
                    field.text = str(digit)
                self.input_reset_mode = False
            elif digit == '.':
                if '.' in current:
                    return
                if not current:
                    field.text = '0.'
                else:
                    field.text = current + '.'
            elif current == '0':
                field.text = str(digit)
            else:
                field.text = current + str(digit)
            update_calculations()

        def backspace(x=None):
            field = get_active_field()
            if not field:
                return
            current = field.text
            self.input_reset_mode = False
            if len(current) > 0:
                field.text = current[:-1]
            update_calculations()
        grid = MDGridLayout(cols=3, spacing='8dp', size_hint_y=1, padding=[20, 0])
        keys = ['7', '8', '9', '4', '5', '6', '1', '2', '3', '.', '0', 'DEL']
        for key in keys:
            if key == 'DEL':
                btn = MDIconButton(icon='backspace-outline', theme_text_color='Custom', text_color=(0, 0, 0, 1), md_bg_color=(0.8, 0.8, 0.8, 1), size_hint=(1, 1), icon_size='22sp', on_release=backspace)
            else:
                btn = MDRaisedButton(text=key, md_bg_color=(0.95, 0.95, 0.95, 1), theme_text_color='Custom', text_color=(0, 0, 0, 1), font_size='22sp', size_hint=(1, 1), elevation=1, on_release=lambda x, k=key: add_digit(k))
            grid.add_widget(btn)
        content.add_widget(grid)
        buttons_box = MDBoxLayout(orientation='horizontal', spacing='10dp', size_hint_y=None, height='60dp')
        btn_cancel = MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=(0.5, 0.5, 0.5, 1), size_hint_x=0.3, on_release=lambda x: self.edit_dialog.dismiss())

        def save_changes(x):
            try:
                q_text = self.edit_qty_field.text or '0'
                new_q = float(q_text)
                if new_q <= 0:
                    raise ValueError
                item['qty'] = new_q
                if is_virtual:
                    if hasattr(self, 'edit_name_field'):
                        new_name = self.edit_name_field.text.strip()
                        if new_name:
                            item['name'] = new_name
                    if hasattr(self, 'edit_price_field'):
                        item['price'] = float(self.edit_price_field.text or 0)
                elif hasattr(self, 'edit_price_field') and self.active_edit_target == 'price':
                    item['price'] = float(self.edit_price_field.text or 0)
                if hasattr(self, 'edit_tva_field'):
                    item['tva'] = float(self.edit_tva_field.text or 0)
                self.refresh_cart_screen_items()
                self.update_cart_button()
                self.edit_dialog.dismiss()
                self.notify('Modifié avec succès', 'success')
            except:
                self.notify('Valeurs invalides', 'error')
        self.btn_save_edit.bind(on_release=save_changes)
        buttons_box.add_widget(btn_cancel)
        buttons_box.add_widget(self.btn_save_edit)
        content.add_widget(buttons_box)
        update_edit_colors()
        update_calculations()
        self.edit_dialog = MDDialog(title='', type='custom', content_cls=content, buttons=[], size_hint=(0.85, None))
        self.edit_dialog.open()

    def open_seller_auth_dialog(self, x):
        if hasattr(self, 'settings_menu_dialog') and self.settings_menu_dialog:
            self.settings_menu_dialog.dismiss()
        stored_pass = self.db.get_setting('seller_password')
        has_pass = stored_pass is not None and stored_pass != ''
        self.temp_new_seller_pass = None
        if has_pass:
            title = 'Accès Admin'
            hint = 'Code PIN / Mot de passe'
            btn_text = 'SE CONNECTER'
        else:
            title = 'Nouveau Mot de Passe'
            hint = 'Créez un mot de passe'
            btn_text = 'SUIVANT'
        content = MDBoxLayout(orientation='vertical', spacing=10, size_hint_y=None, height=dp(80))
        self.seller_pass_field = MDTextField(hint_text=hint, password=True, halign='center', input_type='number')
        content.add_widget(self.seller_pass_field)
        self.btn_auth_action = MDRaisedButton(text=btn_text, md_bg_color=(0, 0.6, 0.8, 1), on_release=lambda x: self.check_create_seller_pass(has_pass, stored_pass))
        self.auth_dialog = MDDialog(title=title, type='custom', content_cls=content, buttons=[MDFlatButton(text='ANNULER', on_release=lambda x: self.auth_dialog.dismiss()), self.btn_auth_action])
        self.auth_dialog.open()

    def check_create_seller_pass(self, exists, stored_pass):
        pwd = self.seller_pass_field.text.strip()
        if not pwd:
            self.notify('Le mot de passe est vide !', 'error')
            return
        if exists:
            if pwd == stored_pass:
                self.auth_dialog.dismiss()
                self.open_seller_toggle_dialog()
            else:
                self.notify('Mot de passe incorrect', 'error')
                self.seller_pass_field.text = ''
        elif self.temp_new_seller_pass is None:
            self.temp_new_seller_pass = pwd
            self.auth_dialog.title = 'Confirmer le mot de passe'
            self.seller_pass_field.text = ''
            self.seller_pass_field.hint_text = 'Répétez le mot de passe'
            self.btn_auth_action.text = 'CONFIRMER'
            Clock.schedule_once(lambda dt: setattr(self.seller_pass_field, 'focus', True), 0.2)
        elif pwd == self.temp_new_seller_pass:
            self.db.set_setting('seller_password', pwd)
            self.notify('Mot de passe créé avec succès', 'success')
            self.auth_dialog.dismiss()
            self.open_seller_toggle_dialog()
        else:
            self.notify('Les mots de passe ne correspondent pas', 'error')
            self.temp_new_seller_pass = None
            self.auth_dialog.title = 'Nouveau Mot de Passe'
            self.seller_pass_field.text = ''
            self.seller_pass_field.hint_text = 'Créez un mot de passe'
            self.btn_auth_action.text = 'SUIVANT'

    def open_seller_toggle_dialog(self):
        content = MDBoxLayout(orientation='horizontal', spacing=20, size_hint_y=None, height=dp(50), padding=[20, 0])
        content.add_widget(MDLabel(text='Mode Vendeur (Restreint)'))
        current_state = self.db.get_setting('config_seller_mode', 'False') == 'True'
        chk = MDCheckbox(active=current_state, size_hint=(None, None), size=(dp(48), dp(48)))
        chk.bind(active=self.on_seller_mode_switch)
        content.add_widget(chk)
        self.toggle_dialog = MDDialog(title='Configuration Mode', type='custom', content_cls=content, buttons=[MDFlatButton(text='FERMER', on_release=lambda x: self.toggle_dialog.dismiss())])
        self.toggle_dialog.open()

    def on_seller_mode_switch(self, instance, value):
        self.is_seller_mode = value
        self.db.set_setting('config_seller_mode', str(value))
        self.update_dashboard_layout()
        status = 'Activé' if value else 'Désactivé'
        self.notify(f'Mode Vendeur: {status}', 'info')

    def show_manage_product_dialog(self, product, prefilled_barcode=None):
        try:
            self.temp_selected_image_path = None
            self.remove_image_order = False
            if product and product.get('name') == 'Autre Article':
                self.notify('Modification interdite (Système)', 'error')
                return
            is_edit = product is not None
            has_movements = False
            current_image = ''
            val_name = ''
            val_barcode = prefilled_barcode if prefilled_barcode else ''
            val_reference = ''
            val_num_prod = ''
            val_family = 'TOUS'
            val_stock = ''
            val_cost = ''
            val_p1 = ''
            val_p2 = ''
            val_p3 = ''
            is_used = False
            is_unlimited = False
            val_is_promo = False
            val_promo_type = 'fixed'
            val_promo_value = ''
            val_promo_expiry = ''

            def fmt(v):
                try:
                    val = float(v)
                    return f'{val:.2f}'.replace('.', ',') if val > 0 else ''
                except:
                    return ''

            def fmt_int(v):
                try:
                    f = float(v)
                    return str(int(f)) if f.is_integer() else str(f)
                except:
                    return ''
            if is_edit:
                has_movements = self.db.check_product_has_movements(product['id'])
                val_name = product.get('name', '')
                val_barcode = str(product.get('barcode') or '')
                val_reference = product.get('reference', '') or ''
                is_used = product.get('is_used', 0) == 1
                current_image = product.get('image_path', '')
                val_num_prod = str(product.get('product_ref') or product.get('ref') or '')
                raw_fam = product.get('family', '')
                val_family = raw_fam if raw_fam else 'TOUS'
                if val_num_prod == 'None':
                    val_num_prod = ''
                raw_cost = product.get('purchase_price')
                if raw_cost is None:
                    raw_cost = product.get('cost', 0)
                val_cost = fmt(raw_cost)
                val_p1 = fmt(product.get('price', 0))
                val_p2 = fmt(product.get('price_semi', 0))
                val_p3 = fmt(product.get('price_wholesale', 0))
                raw_stock = float(product.get('stock', 0) or 0.0)
                is_unlimited = raw_stock <= -900000
                val_stock = '' if is_unlimited else fmt_int(raw_stock)
                val_is_promo = product.get('is_promo_active', 0) == 1
                val_promo_type = product.get('promo_type', 'fixed')
                val_promo_value = fmt(product.get('promo_value', 0))
                val_promo_expiry = str(product.get('promo_expiry', ''))
            title = 'Fiche Produit' if is_edit else 'Nouveau Produit'
            scroll = MDScrollView(size_hint_y=None, height=dp(600))
            main_box = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(15), padding=[dp(10), dp(10), dp(10), dp(20)])
            card_info = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.98, 0.98, 0.98, 1))
            header_info = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
            header_info.add_widget(MDIcon(icon='information-outline', theme_text_color='Primary', font_size='20sp'))
            header_info.add_widget(MDLabel(text='Informations', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
            card_info.add_widget(header_info)
            card_info.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
            self.field_num = MDTextField(text=val_num_prod, hint_text='N° Produit', size_hint_x=1, icon_right='pound')
            card_info.add_widget(self.field_num)
            box_bar = MDBoxLayout(orientation='horizontal', spacing=dp(5), size_hint_x=1, adaptive_height=True)
            self.field_bar = MDTextField(text=val_barcode, hint_text='Code-Barres', size_hint_x=1, icon_right='barcode')

            def scan_into_field(x):
                self.target_scan_field = self.field_bar
                self.open_barcode_scanner(None)
            btn_scan_field = MDIconButton(icon='barcode-scan', theme_text_color='Custom', text_color=(0, 0, 0, 1), md_bg_color=(0.9, 0.9, 0.9, 1), on_release=scan_into_field)
            btn_gen = MDIconButton(icon='refresh', theme_text_color='Custom', text_color=self.theme_cls.primary_color, on_release=lambda x: setattr(self.field_bar, 'text', '7' + ''.join([str(random.randint(0, 9)) for _ in range(12)])))
            box_bar.add_widget(self.field_bar)
            box_bar.add_widget(btn_scan_field)
            box_bar.add_widget(btn_gen)
            card_info.add_widget(box_bar)
            # ================================================
            self.field_name = SmartTextField(text=val_name, hint_text='Désignation*', required=True, icon_right='tag-text-outline')
            card_info.add_widget(self.field_name)
            self.field_reference = SmartTextField(text=val_reference, hint_text='Référence', icon_right='text')
            card_info.add_widget(self.field_reference)
            card_info.add_widget(MDLabel(text='La famille:', font_style='Caption', theme_text_color='Secondary'))
            family_border_card = MDCard(size_hint_y=None, height=dp(50), radius=[4], md_bg_color=(0, 0, 0, 0), line_color=(0, 0, 0, 1), line_width=1, elevation=0)
            family_container = MDFloatLayout()
            self.btn_select_family = MDFlatButton(text=val_family, theme_text_color='Custom', text_color=(0.1, 0.1, 0.1, 1), size_hint=(1, 1), pos_hint={'center_x': 0.5, 'center_y': 0.5}, on_release=lambda x: self.open_family_selector_dialog())
            btn_add_family = MDIconButton(icon='plus-circle', theme_text_color='Custom', text_color=(0, 0.7, 0, 1), size_hint=(None, None), size=(dp(40), dp(40)), icon_size='28sp', pos_hint={'right': 1, 'center_y': 0.5}, on_release=lambda x: self.show_add_family_dialog())
            family_container.add_widget(self.btn_select_family)
            family_container.add_widget(btn_add_family)
            family_border_card.add_widget(family_container)
            card_info.add_widget(family_border_card)
            box_img = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True, padding=[0, dp(5), 0, 0])
            img_status_text = 'Aucune'
            if current_image:
                if os.path.exists(current_image):
                    img_status_text = os.path.basename(current_image)
                else:
                    potential = os.path.join(self.user_data_dir, 'product_images', os.path.basename(current_image))
                    if os.path.exists(potential):
                        img_status_text = os.path.basename(current_image)
            self.lbl_image_status = MDLabel(text=f'Image: {img_status_text}', font_style='Caption', size_hint_x=0.5, theme_text_color='Secondary', valign='center', shorten=True)

            def clear_image_selection(x):
                self.temp_selected_image_path = None
                self.remove_image_order = True
                self.lbl_image_status.text = 'Image: Aucune (Sera supprimée)'
                self.lbl_image_status.text_color = (0.8, 0, 0, 1)
            btn_del_img = MDIconButton(icon='trash-can-outline', theme_text_color='Custom', text_color=(0.8, 0, 0, 1), on_release=clear_image_selection)
            btn_img = MDRaisedButton(text='Choisir', size_hint_x=0.3, md_bg_color=(0.2, 0.6, 0.8, 1), on_release=self.open_image_selector, elevation=0)
            box_img.add_widget(self.lbl_image_status)
            box_img.add_widget(btn_del_img)
            box_img.add_widget(btn_img)
            card_info.add_widget(box_img)
            main_box.add_widget(card_info)
            card_stock = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.96, 0.99, 0.96, 1))
            header_stock = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
            header_stock.add_widget(MDIcon(icon='package-variant-closed', theme_text_color='Primary', font_size='20sp'))
            header_stock.add_widget(MDLabel(text='Stock', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
            card_stock.add_widget(header_stock)
            card_stock.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
            row_stock = MDBoxLayout(orientation='horizontal', spacing=dp(15), adaptive_height=True, padding=[0, dp(10), 0, 0])
            box_chk = MDBoxLayout(orientation='horizontal', spacing=dp(5), size_hint_x=0.5, adaptive_height=True)
            self.chk_unlimited = MDCheckbox(active=is_unlimited, size_hint=(None, None), size=(dp(40), dp(40)), pos_hint={'center_y': 0.5})
            lbl_unlimited = MDLabel(text='Illimité', font_style='Body2', theme_text_color='Secondary', pos_hint={'center_y': 0.5})
            if has_movements:
                self.chk_unlimited.disabled = True
                lbl_unlimited.text = 'Illimité\n(Verrouillé)'
            box_chk.add_widget(self.chk_unlimited)
            box_chk.add_widget(lbl_unlimited)
            row_stock.add_widget(box_chk)
            self.field_stock = MDTextField(text=val_stock, hint_text='Quantité', input_filter='float', size_hint_x=0.5)

            def on_checkbox_active(checkbox, value):
                try:
                    if value:
                        self.field_stock.disabled = True
                        self.field_stock.text = ''
                        self.field_stock.hint_text = '---'
                        self.field_stock.line_color_normal = (0, 0, 0, 0)
                    elif has_movements:
                        self.field_stock.disabled = True
                        self.field_stock.helper_text = 'Verrouillé (Mouvements)'
                        self.field_stock.helper_text_mode = 'persistent'
                    else:
                        self.field_stock.disabled = False
                        self.field_stock.hint_text = 'Stock Initial'
                        self.field_stock.helper_text = ''
                except:
                    pass
            self.chk_unlimited.bind(active=on_checkbox_active)
            on_checkbox_active(self.chk_unlimited, is_unlimited)
            row_stock.add_widget(self.field_stock)
            card_stock.add_widget(row_stock)
            main_box.add_widget(card_stock)
            card_price = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(0.96, 0.96, 0.99, 1))
            header_price = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
            header_price.add_widget(MDIcon(icon='cash-multiple', theme_text_color='Primary', font_size='20sp'))
            header_price.add_widget(MDLabel(text='Tarification', bold=True, theme_text_color='Primary', font_style='Subtitle1'))
            card_price.add_widget(header_price)
            card_price.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
            self.field_cost = MDTextField(text=val_cost, hint_text="Prix d'Achat (Coût)", input_filter='float', icon_right='truck-delivery')
            if is_used:
                self.field_cost.helper_text = 'Produit utilisé'
            card_price.add_widget(self.field_cost)
            self.field_p1 = MDTextField(text=val_p1, hint_text='Prix Détail', input_filter='float', icon_right='tag')
            card_price.add_widget(self.field_p1)
            self.field_p2 = MDTextField(text=val_p2, hint_text='Prix Demi-Gros', input_filter='float', icon_right='tag-multiple')
            card_price.add_widget(self.field_p2)
            self.field_p3 = MDTextField(text=val_p3, hint_text='Prix Gros', input_filter='float', icon_right='currency-usd')
            card_price.add_widget(self.field_p3)
            main_box.add_widget(card_price)
            card_promo = MDCard(orientation='vertical', radius=[12], padding=dp(15), spacing=dp(10), elevation=1, adaptive_height=True, md_bg_color=(1, 0.95, 0.95, 1))
            header_promo = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
            header_promo.add_widget(MDIcon(icon='sale', theme_text_color='Custom', text_color=(0.8, 0, 0, 1), font_size='20sp'))
            header_promo.add_widget(MDLabel(text='Promotion (Solde)', bold=True, theme_text_color='Custom', text_color=(0.8, 0, 0, 1), font_style='Subtitle1'))
            self.chk_promo_active = MDCheckbox(active=val_is_promo, size_hint=(None, None), size=(dp(40), dp(40)))
            header_promo.add_widget(MDLabel(text='Activer', halign='right', font_style='Caption'))
            header_promo.add_widget(self.chk_promo_active)
            card_promo.add_widget(header_promo)
            card_promo.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
            type_box = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
            self.btn_promo_fixed = MDRaisedButton(text='Prix Fixe', size_hint_x=0.5, elevation=0)
            self.btn_promo_percent = MDRaisedButton(text='Pourcentage %', size_hint_x=0.5, elevation=0)
            self.promo_type_selected = val_promo_type

            def update_promo_buttons():
                if self.promo_type_selected == 'fixed':
                    self.btn_promo_fixed.md_bg_color = (0.8, 0, 0, 1)
                    self.btn_promo_fixed.text_color = (1, 1, 1, 1)
                    self.btn_promo_percent.md_bg_color = (0.9, 0.9, 0.9, 1)
                    self.btn_promo_percent.text_color = (0, 0, 0, 1)
                    self.field_promo_val.hint_text = 'Nouveau Prix (DA)'
                else:
                    self.btn_promo_fixed.md_bg_color = (0.9, 0.9, 0.9, 1)
                    self.btn_promo_fixed.text_color = (0, 0, 0, 1)
                    self.btn_promo_percent.md_bg_color = (0.8, 0, 0, 1)
                    self.btn_promo_percent.text_color = (1, 1, 1, 1)
                    self.field_promo_val.hint_text = 'Remise (%)'
            self.btn_promo_fixed.on_release = lambda: [setattr(self, 'promo_type_selected', 'fixed'), update_promo_buttons()]
            self.btn_promo_percent.on_release = lambda: [setattr(self, 'promo_type_selected', 'percent'), update_promo_buttons()]
            type_box.add_widget(self.btn_promo_fixed)
            type_box.add_widget(self.btn_promo_percent)
            card_promo.add_widget(type_box)
            row_promo_val = MDBoxLayout(orientation='horizontal', spacing=dp(10), adaptive_height=True)
            self.field_promo_val = MDTextField(text=val_promo_value, hint_text='Valeur', input_filter='float', size_hint_x=1)
            row_promo_val.add_widget(self.field_promo_val)
            card_promo.add_widget(row_promo_val)
            box_date = MDBoxLayout(orientation='horizontal', spacing=dp(5), adaptive_height=True)
            self.btn_promo_date = MDRaisedButton(text=val_promo_expiry if val_promo_expiry else 'Date Fin (Optionnel)', md_bg_color=(0.3, 0.3, 0.3, 1), size_hint_x=0.85)

            def on_promo_date_save(instance, value, date_range):
                self.btn_promo_date.text = str(value)

            def open_promo_picker(x):
                date_dialog = MDDatePicker()
                date_dialog.bind(on_save=on_promo_date_save)
                date_dialog.open()
            self.btn_promo_date.on_release = lambda: open_promo_picker(None)
            btn_reset_date = MDIconButton(icon='calendar-remove', theme_text_color='Custom', text_color=(0.8, 0.2, 0.2, 1), size_hint_x=0.15, pos_hint={'center_y': 0.5}, on_release=lambda x: setattr(self.btn_promo_date, 'text', 'Date Fin (Optionnel)'))
            box_date.add_widget(self.btn_promo_date)
            box_date.add_widget(btn_reset_date)
            card_promo.add_widget(box_date)
            update_promo_buttons()
            main_box.add_widget(card_promo)
            if not is_edit:
                try:
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    cursor.execute('SELECT product_ref FROM products')
                    rows = cursor.fetchall()
                    conn.close()
                    max_num = 0
                    for row in rows:
                        ref_str = str(row[0])
                        if ref_str.isdigit():
                            val_int = int(ref_str)
                            if val_int > max_num:
                                max_num = val_int
                    next_ref = f'{max_num + 1:05d}'
                    self.field_num.text = next_ref
                except:
                    self.field_num.text = '00001'

            def save_product_action(x):
                try:
                    name_val = self.field_name.get_value().strip()
                    if not name_val:
                        self.field_name.error = True
                        self.notify('Erreur: Le nom est obligatoire', 'error')
                        return

                    def safe_float_standard(text_widget):
                        if not text_widget or not text_widget.text:
                            return 0.0
                        clean_str = str(text_widget.text).replace(',', '.').replace(' ', '').strip()
                        try:
                            return float(clean_str)
                        except:
                            return 0.0
                    cost_val = safe_float_standard(self.field_cost)
                    p1_val = safe_float_standard(self.field_p1)
                    p2_val = safe_float_standard(self.field_p2)
                    p3_val = safe_float_standard(self.field_p3)
                    is_p_active = self.chk_promo_active.active
                    p_type = self.promo_type_selected
                    p_val = safe_float_standard(self.field_promo_val)
                    p_lim = 0.0
                    p_exp = self.btn_promo_date.text if self.btn_promo_date.text != 'Date Fin (Optionnel)' else ''
                    initial_stock_val = 0.0
                    stock_to_save = 0.0
                    if self.chk_unlimited.active:
                        stock_to_save = -1000000.0
                    elif has_movements and is_edit:
                        stock_to_save = float(product.get('stock', 0))
                        initial_stock_val = stock_to_save
                    else:
                        stock_to_save = safe_float_standard(self.field_stock)
                        initial_stock_val = stock_to_save
                    final_img_path = current_image
                    if self.temp_selected_image_path:
                        if is_edit and current_image and os.path.exists(current_image):
                            try:
                                os.remove(current_image)
                            except:
                                pass
                        saved_path = self.save_product_image_local(self.temp_selected_image_path)
                        if saved_path:
                            final_img_path = saved_path
                    elif self.remove_image_order:
                        if current_image and os.path.exists(current_image):
                            try:
                                os.remove(current_image)
                            except Exception as e:
                                pass
                        final_img_path = ''
                    family_to_save = self.btn_select_family.text
                    if family_to_save == 'TOUS':
                        family_to_save = ''
                    payload = {'name': name_val, 'barcode': self.field_bar.text.strip(), 'reference': self.field_reference.get_value().strip(), 'stock': stock_to_save, 'product_ref': self.field_num.text.strip(), 'purchase_price': cost_val, 'price': p1_val, 'price_semi': p2_val, 'price_wholesale': p3_val, 'image_path': final_img_path, 'category': '', 'family': family_to_save, 'unit': '', 'user_name': self.current_user_name if hasattr(self, 'current_user_name') else 'ADMIN', 'action': 'update' if is_edit else 'add', 'is_promo_active': is_p_active, 'promo_type': p_type, 'promo_value': p_val, 'promo_qty_limit': p_lim, 'promo_expiry': p_exp}
                    if is_edit:
                        payload['id'] = product['id']
                    saved_p_id = self.db.save_product(payload)
                    if not saved_p_id and (not is_edit):
                        raise Exception('DB Error: ID not returned')
                    target_pid = int(product['id']) if is_edit else int(saved_p_id)
                    if not self.chk_unlimited.active and (not has_movements):
                        existing_bi_id = self.db.get_product_bi_transaction(target_pid)
                        if existing_bi_id:
                            self.db.update_bi_transaction_qty(existing_bi_id, target_pid, initial_stock_val, cost_val)
                        elif initial_stock_val > 0:
                            supplier_id = None
                            try:
                                conn = self.db.get_connection()
                                cursor = conn.cursor()
                                cursor.execute('SELECT id FROM suppliers WHERE name = ?', (AppConstants.DEFAULT_SUPPLIER_NAME,))
                                row = cursor.fetchone()
                                if row:
                                    supplier_id = row[0]
                                conn.close()
                            except:
                                pass
                            if supplier_id:
                                timestamp = str(datetime.now()).split('.')[0]
                                total_amount = initial_stock_val * cost_val
                                trans_data = {'doc_type': 'BI', 'items': [], 'user_name': payload['user_name'], 'timestamp': timestamp, 'purchase_location': 'store', 'entity_id': supplier_id, 'payment_info': {'amount': total_amount, 'method': 'Initial', 'total': total_amount}, 'is_simple_payment': False, 'amount': total_amount, 'note': 'Stock Initial'}
                                t_id = self.db.save_transaction(trans_data)
                                conn = self.db.get_connection()
                                cursor = conn.cursor()
                                cursor.execute('\n                                    INSERT INTO transaction_items \n                                    (transaction_id, product_id, product_name, qty, price, tva, is_return, cost_price) \n                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)\n                                ', (t_id, target_pid, payload['name'], initial_stock_val, cost_val, 0, 0, cost_val))
                                conn.commit()
                                conn.close()
                    if self.dialog:
                        self.dialog.dismiss()
                    if hasattr(self, 'search_field') and self.search_field:
                        self.search_field.text = ''
                    self.load_more_products(reset=True)
                    self.update_family_filter_ui()
                    self.notify('Produit enregistré avec succès', 'success')
                except ValueError as ve:
                    self.notify('Erreur de format numérique (Vérifiez les prix/stock)', 'error')
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    self.notify(f'Erreur Sauvegarde: {e}', 'error')

            def delete_product_flow(x):
                if has_movements:
                    self.notify('Impossible: Mouvements détectés', 'error')
                    return

                def confirm(y):
                    if self.conf_diag:
                        self.conf_diag.dismiss()
                    try:
                        img_to_del = product.get('image_path', '')
                        if img_to_del and os.path.exists(img_to_del):
                            try:
                                os.remove(img_to_del)
                            except:
                                pass
                        self.db.delete_product(product['id'])
                        bi_id = self.db.get_product_bi_transaction(product['id'])
                        if bi_id:
                            self.db.delete_transaction(bi_id)
                        if self.dialog:
                            self.dialog.dismiss()
                        if hasattr(self, 'search_field') and self.search_field:
                            self.search_field.text = ''
                        self.load_more_products(reset=True)
                        self.notify('Produit supprimé', 'success')
                    except Exception as e:
                        self.notify(f'Erreur: {e}', 'error')
                self.conf_diag = MDDialog(title='Confirmation', text='Voulez-vous vraiment supprimer ce produit ?', buttons=[MDFlatButton(text='NON', on_release=lambda z: self.conf_diag.dismiss()), MDRaisedButton(text='OUI', md_bg_color=(1, 0, 0, 1), on_release=confirm)])
                self.conf_diag.open()
            footer_box = MDBoxLayout(orientation='vertical', spacing=dp(10), adaptive_height=True, padding=[0, dp(20), 0, 0])
            btn_save = MDRaisedButton(text='ENREGISTRER', md_bg_color=(0, 0.7, 0, 1), text_color=(1, 1, 1, 1), elevation=2, size_hint_x=1, height=dp(50), on_release=save_product_action)
            footer_box.add_widget(btn_save)
            if is_edit:
                btn_del = MDRaisedButton(text='SUPPRIMER', md_bg_color=(0.8, 0, 0, 1), text_color=(1, 1, 1, 1), elevation=2, size_hint_x=1, height=dp(50), on_release=delete_product_flow)
                footer_box.add_widget(btn_del)
            btn_close = MDRaisedButton(text='FERMER', md_bg_color=(0.9, 0.9, 0.9, 1), text_color=(0.2, 0.2, 0.2, 1), elevation=0, size_hint_x=1, height=dp(50), on_release=lambda x: self.dialog.dismiss())
            footer_box.add_widget(btn_close)
            main_box.add_widget(footer_box)
            scroll.add_widget(main_box)
            self.dialog = MDDialog(title=title, type='custom', content_cls=scroll, buttons=[], size_hint=(0.95, 0.85))
            self.dialog.open()
        except Exception as e:
            self.notify(f'Erreur UI: {e}', 'error')
            print(f'Error in show_manage_product_dialog: {e}')

    def add_to_cart(self, product):
        try:
            if hasattr(self.qty_field, 'get_value'):
                qty = float(self.qty_field.get_value())
            else:
                qty = float(self.qty_field.text)
            if qty <= 0:
                raise ValueError
        except:
            self.notify('Quantité invalide', 'error')
            return
        if product.get('id') == -999 or str(product.get('name')).startswith('Autre Article'):
            count = sum((1 for item in self.cart if str(item.get('name')).startswith('Autre Article')))
            new_name = f'Autre Article {count + 1}'
            final_price = float(product.get('price', 0))
            self.cart.append({'id': -999, 'name': new_name, 'price': final_price, 'qty': qty, 'original_unit_price': final_price, 'tva': 0, 'is_virtual': True})
            self._finish_add_to_cart(new_name)
            return
        final_price = float(product.get('price', 0))
        found = False
        for item in self.cart:
            if item['id'] == product['id']:
                item['qty'] += qty
                item['price'] = final_price
                found = True
                break
        if not found:
            self.cart.append({'id': product['id'], 'name': product['name'], 'price': final_price, 'qty': qty, 'original_unit_price': final_price, 'tva': 0, 'has_promo': False})
        self._finish_add_to_cart('Article ajouté')

    def _finish_add_to_cart(self, msg):
        if hasattr(self, 'dialog') and self.dialog:
            self.dialog.dismiss()
        self.update_cart_button()
        self.notify(msg, 'success')
        if hasattr(self, 'search_field') and self.search_field:
            self.search_field.text = ''
            self.filter_products(None, '')
            Clock.schedule_once(lambda x: setattr(self.search_field, 'focus', True), 0.2)

    def update_cart_button(self):
        try:
            count = len(self.cart)
            doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
            doc_type = doc_type_map.get(self.current_mode, 'BV')
            fin_f = AppConstants.FINANCIAL_FACTORS.get(doc_type, 0)
            is_invoice_mode = doc_type in ['FC', 'FF', 'FP']
            total_decimal = Decimal('0.00')
            for item in self.cart:
                p = to_decimal(item.get('price', 0))
                q = to_decimal(item.get('qty', 0))
                tva = to_decimal(item.get('tva', 0)) if is_invoice_mode else Decimal('0.00')
                line_ht = p * q
                line_ttc = line_ht * (Decimal('1') + tva / Decimal('100'))
                total_decimal += line_ttc
            total_val = quantize_decimal(total_decimal)
            if self.lbl_cart_count:
                self.lbl_cart_count.text = f'PANIER ({count})'
            if self.lbl_cart_total:
                if fin_f == 0 and doc_type != 'FP':
                    self.lbl_cart_total.text = ''
                else:
                    self.lbl_cart_total.text = f'{total_val:.2f} DA'
        except Exception as e:
            print(f'Cart Update Error: {e}')

    def remove_from_cart(self, item):
        if item in self.cart:
            self.cart.remove(item)
        self.refresh_cart_screen_items()
        self.update_cart_button()

    def start_payment_flow(self, mode):
        self.current_mode = mode
        entity_type = 'account' if mode == 'client_payment' else 'supplier'
        self.theme_cls.primary_palette = 'Teal' if mode == 'client_payment' else 'Brown'
        self.load_local_entities(entity_type)
        self.show_entity_selection_dialog(None, next_action=self.show_simple_payment_dialog)

    def show_simple_payment_dialog(self, amount=None):
        if not self.selected_entity:
            return
        if not hasattr(self, 'temp_note'):
            self.temp_note = ''
        if self.current_mode == 'client_payment':
            title = 'Versement Client'
            theme_col = (0, 0.6, 0, 1)
        else:
            title = 'Règlement Fournisseur'
            theme_col = (0.8, 0.4, 0, 1)
        content = MDBoxLayout(orientation='vertical', spacing='12dp', size_hint_y=None, height=dp(560), padding=[0, '10dp', 0, 0])
        header_box = MDBoxLayout(orientation='vertical', adaptive_height=True, padding=[0, 0, 0, '5dp'])
        ent_name = self.fix_text(self.selected_entity['name'])
        header_box.add_widget(MDLabel(text=ent_name, halign='center', font_style='H5', bold=True, theme_text_color='Primary', shorten=True, shorten_from='right'))
        content.add_widget(header_box)

        def backspace(x=None):
            current = self.txt_simple_amount.text
            if current:
                self.txt_simple_amount.text = current[:-1]
            if not self.txt_simple_amount.text:
                self.txt_simple_amount.text = ''

        def add_digit(digit):
            current = self.txt_simple_amount.text
            if digit == '.':
                if '.' in current:
                    return
                if not current:
                    self.txt_simple_amount.text = '0.'
                else:
                    self.txt_simple_amount.text = current + '.'
            elif current == '0':
                self.txt_simple_amount.text = str(digit)
            elif current == '-0':
                self.txt_simple_amount.text = '-' + str(digit)
            else:
                self.txt_simple_amount.text = current + str(digit)
        input_row = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(110), spacing='10dp', padding=[dp(20), 0])
        val = ''
        if amount:
            try:
                f_val = float(amount)
                if f_val.is_integer():
                    val = str(int(f_val))
                else:
                    val = str(f_val)
            except:
                val = str(amount)
        self.txt_simple_amount = MDTextField(text=val, hint_text='Montant (DA)', font_size='45sp', halign='center', readonly=True, mode='fill', line_color_focus=theme_col, size_hint_x=0.7)
        self.txt_simple_amount.get_value = lambda: self.txt_simple_amount.text
        note_bg_color = (0.2, 0.2, 0.2, 1)
        if self.temp_note and self.temp_note.strip():
            note_bg_color = (0, 0.6, 0, 1)
        self.btn_note_icon = MDIconButton(icon='note-edit-outline', theme_text_color='Custom', text_color=(1, 1, 1, 1), md_bg_color=note_bg_color, size_hint=(None, None), size=(dp(55), dp(55)), pos_hint={'center_y': 0.5}, on_release=self.open_note_input)
        btn_del = MDIconButton(icon='backspace-outline', theme_text_color='Custom', text_color=(1, 1, 1, 1), md_bg_color=(0.9, 0.1, 0.1, 1), size_hint=(None, None), size=(dp(55), dp(55)), pos_hint={'center_y': 0.5}, on_release=backspace)
        input_row.add_widget(self.txt_simple_amount)
        input_row.add_widget(self.btn_note_icon)
        input_row.add_widget(btn_del)
        content.add_widget(input_row)
        grid = MDGridLayout(cols=3, spacing='10dp', size_hint_y=1, padding=[dp(20), dp(10)])
        keys = ['7', '8', '9', '4', '5', '6', '1', '2', '3', '-', '0', '.']
        for key in keys:
            btn = MDRaisedButton(text=key, md_bg_color=(0.96, 0.96, 0.96, 1), theme_text_color='Custom', text_color=(0.1, 0.1, 0.1, 1), font_size='28sp', elevation=1, size_hint=(1, 1), on_release=lambda x, k=key: add_digit(k))
            if key == '-':
                btn.font_size = '38sp'
                btn.text_color = (0, 0, 0, 1)
            grid.add_widget(btn)
        content.add_widget(grid)
        buttons_box = MDBoxLayout(orientation='horizontal', spacing='10dp', size_hint_y=None, height='70dp', padding=[0, '10dp', 0, 0])
        btn_cancel = MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=(0.5, 0.5, 0.5, 1), size_hint_x=0.25, on_release=lambda x: self.simple_pay_dialog.dismiss())
        btn_valid = MDRaisedButton(text='VALIDER', md_bg_color=theme_col, text_color=(1, 1, 1, 1), size_hint_x=0.75, size_hint_y=1, font_size='22sp', elevation=3, on_release=self.submit_simple_payment)
        buttons_box.add_widget(btn_cancel)
        buttons_box.add_widget(btn_valid)
        content.add_widget(buttons_box)
        self.simple_pay_dialog = MDDialog(title=title, type='custom', content_cls=content, size_hint=(0.92, None), buttons=[])
        self.simple_pay_dialog.open()

    def open_note_input(self, instance):
        content = MDBoxLayout(orientation='vertical', spacing='10dp', size_hint_y=None, height=dp(100), padding=dp(10))
        note_field = SmartTextField(text=self.temp_note, hint_text='Entrez une note (Optionnel)')
        content.add_widget(note_field)

        def save_note(x):
            self.temp_note = note_field.get_value().strip()
            if hasattr(self, 'btn_note_icon'):
                if self.temp_note:
                    self.btn_note_icon.md_bg_color = (0, 0.6, 0, 1)
                else:
                    self.btn_note_icon.md_bg_color = (0.2, 0.2, 0.2, 1)
            note_dialog.dismiss()
        note_dialog = MDDialog(title='Ajouter une note', type='custom', content_cls=content, buttons=[MDFlatButton(text='ANNULER', on_release=lambda x: note_dialog.dismiss()), MDRaisedButton(text='OK', on_release=save_note)])
        note_dialog.open()

    def toggle_location(self, x=None):
        self.selected_location = 'warehouse' if self.selected_location == 'store' else 'store'
        self.update_location_display()
        if self.current_mode == 'transfer' and hasattr(self, 'btn_ent_screen'):
            src = 'Magasin' if self.selected_location == 'store' else 'Dépôt'
            dst = 'Dépôt' if self.selected_location == 'store' else 'Magasin'
            self.btn_ent_screen.text = f'{src}  >>>  {dst}'

    def update_location_display(self):
        if hasattr(self, 'btn_loc_screen'):
            if self.selected_location == 'store':
                self.btn_loc_screen.text = 'Magasin'
                self.btn_loc_screen.md_bg_color = self.theme_cls.primary_color
            else:
                self.btn_loc_screen.text = 'Dépôt'
                self.btn_loc_screen.md_bg_color = (0.8, 0.4, 0, 1)

    def show_entity_selection_dialog(self, x, next_action=None):
        self.pending_entity_next_action = next_action
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, height=dp(600))
        self.entity_search = SmartTextField(hint_text='Rechercher...', icon_right='magnify')
        self.entity_search.bind(text=self.filter_entities_paginated)
        content.add_widget(self.entity_search)
        self.rv_entity = EntityRecycleView()
        content.add_widget(self.rv_entity)
        sales_modes = ['sale', 'return_sale', 'client_payment', 'invoice_sale', 'proforma']
        e_type = 'account' if self.current_mode in sales_modes else 'supplier'
        self.entities_source_type = e_type
        title_text = 'Choisir un Client' if e_type == 'account' else 'Choisir un Fournisseur'
        self.entity_dialog = MDDialog(title=title_text, type='custom', content_cls=content, size_hint=(0.9, 0.8))
        self.entity_dialog.open()
        self.active_entity_rv = self.rv_entity
        self.load_more_entities(reset=True)

    def recalculate_cart_prices(self):
        if not self.cart or not self.selected_entity:
            return
        doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
        doc_type = doc_type_map.get(self.current_mode, 'BV')
        stock_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
        if stock_f != -1 and doc_type not in ['FP', 'FC']:
            return
        cat = str(self.selected_entity.get('category', 'Détail')).strip()
        for item in self.cart:
            if item.get('is_virtual', False) or item.get('id') == -999:
                continue
            original_product = self.db.get_product_by_id(item['id'])
            if not original_product:
                continue
            base_price = float(original_product.get('price', 0) or 0)
            new_price = base_price
            is_promo_valid = False
            is_active_val = original_product.get('is_promo_active', 0)
            is_active_bool = str(is_active_val) == '1' or is_active_val == 1
            if is_active_bool:
                promo_exp = str(original_product.get('promo_expiry', '') or '')
                date_valid = True
                if promo_exp and len(promo_exp) > 5:
                    try:
                        from datetime import datetime
                        exp_date = datetime.strptime(promo_exp, '%Y-%m-%d').date()
                        if datetime.now().date() > exp_date:
                            date_valid = False
                    except:
                        pass
                if date_valid:
                    is_promo_valid = True
                    p_type = str(original_product.get('promo_type', 'fixed'))
                    try:
                        p_val = float(original_product.get('promo_value', 0))
                    except:
                        p_val = 0.0
                    if p_type == 'fixed':
                        if p_val > 0:
                            new_price = p_val
                    else:
                        new_price = base_price * (1 - p_val / 100)
            if not is_promo_valid:
                target_price_column = 'price'
                if cat == 'Gros':
                    target_price_column = 'price_wholesale'
                elif cat == 'Demi-Gros':
                    target_price_column = 'price_semi'
                try:
                    cat_price = float(original_product.get(target_price_column, 0) or 0)
                    if cat_price > 0:
                        new_price = cat_price
                    else:
                        new_price = base_price
                except:
                    new_price = base_price
            item['price'] = new_price
            item['original_unit_price'] = new_price
        self.update_cart_button()

    def open_payment_dialog(self, x):
        current_time = time.time()
        if current_time - getattr(self, '_last_click_time', 0) < 1.0:
            return
        self._last_click_time = current_time
        if getattr(self, 'is_transaction_in_progress', False):
            return
        if not self.cart:
            self.dialog = MDDialog(title='Attention', text='Le panier est vide !', buttons=[MDFlatButton(text='OK', on_release=lambda x: self.dialog.dismiss())])
            self.dialog.open()
            return
        doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
        doc_type = doc_type_map.get(self.current_mode, 'BV')
        fin_f = AppConstants.FINANCIAL_FACTORS.get(doc_type, 0)
        is_invoice_mode = doc_type in ['FC', 'FF', 'FP']
        total_ht, total_tva = self.calculate_cart_totals(self.cart, is_invoice_mode)
        self.temp_total_ht = total_ht
        self.temp_total_tva = total_tva
        base_ttc = self._round_num(total_ht + total_tva)
        is_zero_pay_mode = fin_f == 0
        is_COMPTOIR_entity = False
        if self.selected_entity:
            if self.selected_entity.get('name') == AppConstants.DEFAULT_CLIENT_NAME:
                is_COMPTOIR_entity = True
        else:
            is_COMPTOIR_entity = True
        should_skip_dialog = False
        if is_zero_pay_mode:
            should_skip_dialog = True
        elif is_COMPTOIR_entity:
            should_skip_dialog = True
        if should_skip_dialog:
            final_paid_amount = 0.0
            method_val = ''
            if is_zero_pay_mode:
                final_paid_amount = 0.0
            else:
                method_val = 'Espèce'
                timbre = 0.0
                if doc_type == 'FC' and is_COMPTOIR_entity:
                    timbre = AppConstants.calculate_stamp_duty(base_ttc)
                final_paid_amount = to_decimal(base_ttc) + to_decimal(timbre)
                if fin_f == -1:
                    final_paid_amount = to_decimal(base_ttc)
            self.process_transaction(final_paid_amount, to_decimal(base_ttc), method=method_val)
            return
        self.is_invoice_sale = doc_type == 'FC'
        self.show_details = is_invoice_mode
        if self.is_invoice_sale:
            self.payment_methods = [{'label': 'Par défaut', 'value': ''}, {'label': 'Espèce', 'value': 'Espèce'}, {'label': 'Chèque', 'value': 'Chèque'}, {'label': 'Virement', 'value': 'Virement'}, {'label': 'Versement', 'value': 'Versement'}]
            self.current_method_index = 0
            if hasattr(self, 'editing_payment_method') and self.editing_payment_method:
                for idx, m in enumerate(self.payment_methods):
                    if m['value'] == self.editing_payment_method:
                        self.current_method_index = idx
                        break
        dialog_height = dp(640) if self.show_details else dp(580)
        content = MDBoxLayout(orientation='vertical', spacing='10dp', size_hint_y=None, height=dialog_height, padding=['10dp', '0dp', '10dp', '20dp'])
        header_box = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(50), spacing=dp(10))
        lbl_title = MDLabel(text='Paiement', font_style='H5', bold=True, theme_text_color='Primary', valign='center')
        header_box.add_widget(lbl_title)
        if self.is_invoice_sale:
            self.btn_payment_method = MDRaisedButton(text=self.payment_methods[self.current_method_index]['label'], md_bg_color=(0.2, 0.2, 0.2, 1), elevation=2, pos_hint={'center_y': 0.5}, on_release=self._cycle_payment_method)
            header_box.add_widget(self.btn_payment_method)
        content.add_widget(header_box)
        card_height = dp(125) if self.show_details else dp(90)
        total_card = MDCard(orientation='vertical', size_hint_y=None, height=card_height, radius=[10], md_bg_color=(0.95, 0.95, 0.95, 1), elevation=1, padding='5dp')
        if self.show_details:
            details_box = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(25), padding=[dp(5), 0], spacing=dp(5))
            lbl_ht = MDLabel(text=f'HT: {float(self.temp_total_ht):.2f} DA', theme_text_color='Secondary', font_style='Caption', halign='left', bold=True, size_hint_x=0.33)
            details_box.add_widget(lbl_ht)
            self.lbl_timbre = MDLabel(text='', theme_text_color='Custom', text_color=(0.5, 0, 0.5, 1), font_style='Caption', halign='center', bold=True, size_hint_x=0.33)
            details_box.add_widget(self.lbl_timbre)
            lbl_tva = MDLabel(text=f'TVA: {float(self.temp_total_tva):.2f} DA', theme_text_color='Custom', text_color=(0.8, 0, 0, 1), font_style='Caption', halign='right', bold=True, size_hint_x=0.33)
            details_box.add_widget(lbl_tva)
            total_card.add_widget(details_box)
        else:
            self.lbl_timbre = MDLabel()
        total_box = MDBoxLayout(orientation='vertical', spacing=0)
        total_lbl_title = MDLabel(text='NET À PAYER', halign='center', font_style='Caption', theme_text_color='Secondary', size_hint_y=None, height=dp(20))
        self.lbl_final_total = MDLabel(text='', halign='center', font_style='H4', bold=True, theme_text_color='Primary')
        total_box.add_widget(total_lbl_title)
        total_box.add_widget(self.lbl_final_total)
        total_card.add_widget(total_box)
        content.add_widget(total_card)
        try:
            default_val = f'{float(self.editing_payment_amount or 0):.2f}'
        except:
            default_val = '0.00'
        self.txt_paid = MDTextField(text=default_val, hint_text='Montant Versé (DA)', font_size='40sp', halign='center', readonly=True, size_hint_y=None, height=dp(80), mode='fill', line_color_focus=(0, 0, 0, 0))
        self.txt_paid.get_value = lambda: self.txt_paid.text
        content.add_widget(self.txt_paid)
        self.lbl_rest = MDLabel(text='', halign='center', theme_text_color='Custom', font_style='H6', bold=True, size_hint_y=None, height=dp(30))
        content.add_widget(self.lbl_rest)
        grid = MDGridLayout(cols=3, spacing='10dp', size_hint_y=1)
        keys = ['7', '8', '9', '4', '5', '6', '1', '2', '3', '.', '0', 'DEL']

        def add_digit(digit):
            current = self.txt_paid.text
            if digit == '.':
                if '.' in current:
                    return
                if not current:
                    self.txt_paid.text = '0.'
                else:
                    self.txt_paid.text = current + '.'
            elif current == '0' or current == '0.00':
                self.txt_paid.text = str(digit)
            else:
                self.txt_paid.text = current + str(digit)
            self._recalc_ui_totals()

        def backspace(instance=None):
            current = self.txt_paid.text
            if len(current) > 0:
                self.txt_paid.text = current[:-1]
            if not self.txt_paid.text:
                self.txt_paid.text = '0'
            self._recalc_ui_totals()
        for key in keys:
            if key == 'DEL':
                btn = MDIconButton(icon='backspace', theme_text_color='Custom', text_color=(1, 1, 1, 1), md_bg_color=(0.4, 0.4, 0.4, 1), size_hint=(1, 1), icon_size='24sp', on_release=backspace)
            else:
                btn = MDRaisedButton(text=key, md_bg_color=(1, 1, 1, 1), theme_text_color='Custom', text_color=(0, 0, 0, 1), font_size='24sp', size_hint=(1, 1), elevation=1, on_release=lambda x, k=key: add_digit(k))
            grid.add_widget(btn)
        content.add_widget(grid)
        content.add_widget(MDBoxLayout(size_hint_y=None, height='15dp'))
        buttons_box = MDBoxLayout(orientation='horizontal', spacing='10dp', size_hint_y=None, height='55dp')
        btn_cancel = MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=(0.5, 0.5, 0.5, 1), size_hint_x=0.3, on_release=lambda x: self.pay_dialog.dismiss())
        btn_valid = MDRaisedButton(text='VALIDER', md_bg_color=(0, 0.7, 0, 1), text_color=(1, 1, 1, 1), size_hint_x=0.7, size_hint_y=1, font_size='20sp', elevation=2, on_release=lambda x: self.finalize_submission(self.current_final_total))
        buttons_box.add_widget(btn_cancel)
        buttons_box.add_widget(btn_valid)
        content.add_widget(buttons_box)
        self.pay_dialog = MDDialog(title='', type='custom', content_cls=content, buttons=[], size_hint=(0.94, 0.98))
        self._recalc_ui_totals()
        self.pay_dialog.open()

    def _recalc_ui_totals(self):
        base_ttc = to_decimal(self.temp_total_ht) + to_decimal(self.temp_total_tva)
        base_ttc = quantize_decimal(base_ttc)
        timbre = Decimal('0.00')
        has_timbre_label = hasattr(self, 'lbl_timbre') and self.lbl_timbre is not None
        if hasattr(self, 'is_invoice_sale') and self.is_invoice_sale:
            if hasattr(self, 'payment_methods') and self.payment_methods:
                idx = getattr(self, 'current_method_index', 0)
                if idx < len(self.payment_methods):
                    selected_val = self.payment_methods[idx]['value']
                    if selected_val == 'Espèce':
                        t_val = AppConstants.calculate_stamp_duty(float(base_ttc))
                        timbre = to_decimal(t_val)
                        if has_timbre_label:
                            self.lbl_timbre.text = f'Timbre: {timbre:.2f} DA'
                            self.lbl_timbre.opacity = 1
                    elif has_timbre_label:
                        self.lbl_timbre.text = ''
                        self.lbl_timbre.opacity = 0
                elif has_timbre_label:
                    self.lbl_timbre.opacity = 0
        elif has_timbre_label:
            self.lbl_timbre.text = ''
            self.lbl_timbre.opacity = 0
        self.current_final_total = quantize_decimal(base_ttc + timbre)
        if hasattr(self, 'lbl_final_total') and self.lbl_final_total:
            self.lbl_final_total.text = f'{self.current_final_total:.2f} DA'
        try:
            val_text = self.txt_paid.text
            if hasattr(self.txt_paid, 'get_value'):
                val_text = self.txt_paid.get_value()
            paid = to_decimal(val_text) if val_text else Decimal('0.00')
        except:
            paid = Decimal('0.00')
        diff = self.current_final_total - paid
        diff = quantize_decimal(diff)
        if hasattr(self, 'lbl_rest') and self.lbl_rest:
            if diff >= 0:
                self.lbl_rest.text = f'RESTE: {diff:.2f} DA'
                self.lbl_rest.text_color = (0.8, 0, 0, 1)
            else:
                self.lbl_rest.text = f'RENDU: {abs(diff):.2f} DA'
                self.lbl_rest.text_color = (0, 0.6, 0, 1)

    def finalize_submission(self, total_amount):
        current_time = time.time()
        if current_time - getattr(self, '_last_click_time', 0) < 1.0:
            return
        self._last_click_time = current_time
        if getattr(self, 'is_transaction_in_progress', False):
            return
        if self.pay_dialog:
            self.pay_dialog.dismiss()
        payment_method = ''
        if self.current_mode == 'invoice_sale':
            if hasattr(self, 'payment_methods') and hasattr(self, 'current_method_index'):
                try:
                    payment_method = self.payment_methods[self.current_method_index]['value']
                except:
                    payment_method = ''
        total_dec = to_decimal(total_amount)
        if self.current_mode == 'transfer':
            paid_dec = Decimal('0.00')
        else:
            try:
                val_text = self.txt_paid.text
                if hasattr(self.txt_paid, 'get_value'):
                    val_text = self.txt_paid.get_value()
                paid_dec = to_decimal(val_text) if val_text else Decimal('0.00')
            except:
                paid_dec = Decimal('0.00')
            if paid_dec < total_dec:
                remaining = total_dec - paid_dec
                self.show_credit_warning(paid_dec, total_dec, remaining, payment_method)
                return
            if paid_dec > total_dec and self.current_mode not in ['return_sale', 'return_purchase']:
                excess = paid_dec - total_dec
                self.show_overpayment_dialog(paid_dec, total_dec, excess, payment_method)
                return
        Clock.schedule_once(lambda dt: self.process_transaction(paid_dec, total_dec, method=payment_method), 0.1)

    def process_transaction(self, paid_amount, total_amount, method=None):
        if getattr(self, 'is_transaction_in_progress', False):
            return
        self.is_transaction_in_progress = True
        timestamp = str(datetime.now()).split('.')[0]
        try:
            if self.editing_transaction_key and getattr(self, 'editing_doc_type', None):
                doc_type = self.editing_doc_type
            else:
                doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP', 'client_payment': 'CLIENT_PAY', 'supplier_payment': 'SUPPLIER_PAY'}
                doc_type = doc_type_map.get(self.current_mode, 'BV')
            is_invoice_mode = doc_type in ['FC', 'FF', 'FP']
            ht_val, tva_val = self.calculate_cart_totals(self.cart, is_invoice_mode)
            calc_ht = to_decimal(ht_val)
            calc_tva = to_decimal(tva_val)
            base_ttc = calc_ht + calc_tva
            timbre_amount = Decimal('0.00')
            should_apply_stamp = doc_type in AppConstants.APPLY_STAMP_DUTY
            is_cash = 'Espèce' in (method or '')
            if should_apply_stamp and is_cash:
                t_val = AppConstants.calculate_stamp_duty(float(base_ttc))
                timbre_amount = to_decimal(t_val)
            real_total = quantize_decimal(base_ttc + timbre_amount)
            input_paid = to_decimal(paid_amount)
            inv_payment = input_paid
            excess = Decimal('0.00')
            fin_factor = AppConstants.FINANCIAL_FACTORS.get(doc_type, 0)
            if fin_factor == 1:
                if input_paid > real_total:
                    inv_payment = real_total
                    excess = input_paid - real_total
            ent_id = self.selected_entity['id'] if self.selected_entity else None
            payment_info = {'amount': float(inv_payment), 'total': float(real_total), 'method': method, 'timbre': float(timbre_amount)}
            data = {'doc_type': doc_type, 'items': self.cart, 'user_name': self.current_user_name, 'timestamp': timestamp, 'purchase_location': self.selected_location, 'entity_id': ent_id, 'payment_info': payment_info, 'is_simple_payment': False, 'amount': float(real_total)}
            if self.editing_transaction_key:
                data['id'] = self.editing_transaction_key
            t_id = self.db.save_transaction(data)
            try:
                conn = self.db.get_connection()
                cur = conn.cursor()
                cur.execute('SELECT custom_label FROM transactions WHERE id=?', (t_id,))
                row = cur.fetchone()
                if row:
                    data['custom_label'] = row[0]
                conn.close()
            except:
                pass
            if excess > 0 and ent_id:
                s_factor = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
                pay_type = 'SUPPLIER_PAY' if s_factor == 1 else 'CLIENT_PAY'
                visuals = AppConstants.DOC_VISUALS.get(pay_type, {'name': 'Paiement'})
                clean_label = visuals['name'].upper()
                pay_data = {'doc_type': pay_type, 'amount': float(excess), 'entity_id': ent_id, 'custom_label': clean_label, 'user_name': self.current_user_name, 'note': clean_label, 'is_simple_payment': True, 'payment_info': {'amount': float(excess)}, 'timestamp': timestamp, 'items': []}
                self.db.save_transaction(pay_data)
            if not self.editing_transaction_key:
                s_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
                f_f = AppConstants.FINANCIAL_FACTORS.get(doc_type, 0)
                current_sales = to_decimal(self.stat_sales_today)
                current_purchases = to_decimal(self.stat_purchases_today)
                current_c_pay = to_decimal(self.stat_client_payments)
                current_s_pay = to_decimal(self.stat_supplier_payments)
                if s_f == -1 and f_f == 1:
                    current_sales += inv_payment
                elif s_f == 1 and f_f == 1:
                    current_purchases += inv_payment
                if excess > 0:
                    if s_f == -1:
                        current_c_pay += excess
                    elif s_f == 1:
                        current_s_pay += excess
                self.stat_sales_today = float(current_sales)
                self.stat_purchases_today = float(current_purchases)
                self.stat_client_payments = float(current_c_pay)
                self.stat_supplier_payments = float(current_s_pay)
                self.calculate_net_total()
                self.save_local_stats()
            self.notify('Enregistré avec succès', 'success')
            try:
                if self.db.get_setting('printer_auto', 'False') == 'True' and self.db.get_setting('printer_mac', ''):
                    threading.Thread(target=self.print_ticket_bluetooth, args=(data,), daemon=True).start()
            except:
                pass
            self.is_transaction_in_progress = False
            self.cart = []
            self.selected_entity = None
            self.editing_transaction_key = None
            self.go_back()
            target = 'supplier' if AppConstants.STOCK_MOVEMENTS.get(doc_type) == 1 or 'SUPPLIER' in doc_type else 'account'
            self.load_local_entities(target)
            self.check_and_load_stats()
        except Exception as e:
            self.is_transaction_in_progress = False
            import traceback
            traceback.print_exc()
            self.notify(f'Erreur: {e}', 'error')

    def _cycle_payment_method(self, instance):
        self.current_method_index = (self.current_method_index + 1) % len(self.payment_methods)
        new_label = self.payment_methods[self.current_method_index]['label']
        self.btn_payment_method.text = new_label
        self._recalc_ui_totals()

    def show_overpayment_dialog(self, paid, total, excess, method=None):
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, adaptive_height=True, spacing='15dp', padding=[0, '10dp', 0, 0])
        lbl_info = MDLabel(text=f'[b]Montant saisi:[/b] {paid:.2f} DA\n[b]Total:[/b] {total:.2f} DA', markup=True, halign='center', theme_text_color='Primary', font_style='Body1', size_hint_y=None, adaptive_height=True)
        content.add_widget(lbl_info)
        msg_text = ''
        if self.current_mode in ['return_sale', 'return_purchase']:
            msg_text = f"L'excédent [color=#00C853][b]({excess:.2f} DA)[/b][/color] sera déduit du solde."
        else:
            msg_text = f"L'excédent [color=#00C853][b]({excess:.2f} DA)[/b][/color] sera enregistré comme une opération séparée [b](VERSEMENT/RÈGLEMENT)[/b]."
        lbl_msg = MDLabel(text=msg_text, markup=True, halign='center', theme_text_color='Primary', font_style='Subtitle1', size_hint_y=None, adaptive_height=True)
        content.add_widget(lbl_msg)
        buttons = [MDFlatButton(text='CORRIGER', theme_text_color='Custom', text_color=(0.5, 0.5, 0.5, 1), on_release=lambda x: [self.overpay_dialog.dismiss(), self.open_payment_dialog(None)]), MDRaisedButton(text='CONFIRMER', md_bg_color=(0, 0.7, 0, 1), text_color=(1, 1, 1, 1), elevation=2, on_release=lambda x: [self.overpay_dialog.dismiss(), self.process_transaction(paid, total, method=method)])]
        self.overpay_dialog = MDDialog(title="Création d'un Versement", type='custom', content_cls=content, buttons=buttons)
        self.overpay_dialog.open()

    def show_credit_warning(self, paid, total, remaining, method=None):
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, adaptive_height=True, spacing='15dp', padding=[0, '10dp', 0, 0])
        lbl_info = MDLabel(text=f'[b]Montant saisi:[/b] {paid:.2f} DA\n[b]Total:[/b] {total:.2f} DA', markup=True, halign='center', theme_text_color='Primary', font_style='Body1', size_hint_y=None, adaptive_height=True)
        content.add_widget(lbl_info)
        msg_text = ''
        if self.current_mode in ['return_sale', 'return_purchase']:
            msg_text = f'Vous rendez [b]{paid:.2f} DA[/b].\nLe reste [color=#D32F2F][b]({remaining:.2f} DA)[/b][/color] sera déduit de la dette du tiers.'
        else:
            msg_text = f'Le montant versé est insuffisant.\nLe reste [color=#D32F2F][b]({remaining:.2f} DA)[/b][/color] sera enregistré comme [b]CRÉDIT [/b].'
        lbl_msg = MDLabel(text=msg_text, markup=True, halign='center', theme_text_color='Primary', font_style='Subtitle1', size_hint_y=None, adaptive_height=True)
        content.add_widget(lbl_msg)
        buttons = [MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=(0.5, 0.5, 0.5, 1), on_release=lambda x: self.debt_dialog.dismiss()), MDRaisedButton(text='CONFIRMER', md_bg_color=(0.8, 0, 0, 1), text_color=(1, 1, 1, 1), elevation=2, on_release=lambda x: [self.debt_dialog.dismiss(), self.process_transaction(paid, total, method=method)])]
        self.debt_dialog = MDDialog(title='Attention: Crédit', type='custom', content_cls=content, buttons=buttons)
        self.debt_dialog.open()

    def render_transactions_list(self, transactions, target_rv, is_global_mode=False, reset=True):
        if not target_rv:
            return
        if not transactions:
            if reset:
                empty = {'raw_text': 'Aucune opération', 'raw_sec': '', 'amount_text': '', 'icon': 'information-outline', 'icon_color': [0.5, 0.5, 0.5, 1], 'bg_color': [1, 1, 1, 1], 'is_local': True, 'raw_data': None}
                target_rv.data = [empty]
                target_rv.refresh_from_data()
            if hasattr(target_rv, 'loading_lock'):
                target_rv.loading_lock = False
            if is_global_mode:
                self.is_loading_history = False
            return
        new_data = []
        for item in transactions:
            t_type = str(item.get('transaction_type', '')).upper()
            total = item.get('total_amount', 0)
            date_s = str(item.get('date', ''))
            time_str = date_s.split(' ')[1][:5] if ' ' in date_s else ''
            label = item.get('custom_label', '')
            note = str(item.get('note', '')).strip()
            ent_name = ''
            vis = AppConstants.DOC_VISUALS.get(t_type, {'name': t_type, 'icon': 'file', 'color': (0.2, 0.2, 0.2, 1), 'bg': (1, 1, 1, 1)})
            doc_name = vis['name']
            icon_name = vis['icon']
            icon_color = vis['color']
            bg_col = vis['bg']
            if t_type in ['CLIENT_PAY', 'SUPPLIER_PAY'] and label:
                doc_name = label
                label = ''
                upper_name = doc_name.upper()
                if 'CRÉDIT' in upper_name or 'CREDIT' in upper_name or 'DETTE' in upper_name:
                    icon_name = 'cash-minus'
                    icon_color = (0.85, 0, 0, 1)
                    bg_col = (1, 0.95, 0.95, 1)
                elif 'VERSEMENT' in upper_name or 'RÈGLEMENT' in upper_name or 'REGLEMENT' in upper_name:
                    icon_name = 'cash-check'
                    icon_color = (0, 0.7, 0, 1)
                    bg_col = (0.9, 1, 0.9, 1)
            amount_display = f'{abs(total):.2f} DA'
            transfer_direction_text = ''
            if t_type in ['TR', 'TRANSFER']:
                loc = item.get('location', 'store')
                transfer_direction_text = 'Magasin -> Dépôt' if loc == 'store' else 'Dépôt -> Magasin'
                amount_display = ''
            elif is_global_mode:
                ent_name = item.get('client_name')
                if not ent_name and item.get('entity_id'):
                    cid = item.get('entity_id')
                    is_supp = 'SUPPLIER' in t_type or AppConstants.STOCK_MOVEMENTS.get(t_type) == 1
                    found = self.db.get_entity_by_id(cid, 'supplier' if is_supp else 'account')
                    if found:
                        ent_name = found.get('name')
                if not ent_name or ent_name == 'Inconnu':
                    ent_name = AppConstants.DEFAULT_CLIENT_NAME
                if ent_name == 'COMPTOIR':
                    ent_name = AppConstants.DEFAULT_CLIENT_NAME
            if label:
                doc_name = f'{doc_name} - {label}'
            if is_global_mode and t_type not in ['TR', 'TRANSFER']:
                header = f'{doc_name} / {ent_name}'
            else:
                header = f'{doc_name}'
            if t_type in ['TR', 'TRANSFER']:
                header = f'[color=800080]{header}[/color]'
            sec = f'{time_str}'
            if t_type in ['TR', 'TRANSFER']:
                sec += f' | {transfer_direction_text}'
            if t_type != 'BI':
                if note and note != doc_name:
                    sec += f' | {note[:30]}'
            new_data.append({'raw_text': header, 'raw_sec': sec, 'amount_text': amount_display, 'icon': icon_name, 'icon_color': icon_color, 'bg_color': bg_col, 'is_local': True, 'raw_data': item, 'key': str(item.get('id'))})
        if reset:
            target_rv.data = new_data
            target_rv.scroll_y = 1.0
        else:
            target_rv.data.extend(new_data)
        target_rv.refresh_from_data()
        if hasattr(target_rv, 'loading_lock'):
            target_rv.loading_lock = False
        if is_global_mode:
            self.is_loading_history = False

    def generate_pdf_report(self, trans_id, doc_type):
        try:
            details = self.db.get_transaction_full_details(trans_id)
            if not details:
                self.notify('Transaction introuvable', 'error')
                return
            trans = details['transaction']
            initial_entity = details['entity']
            items = details['items']
            store_info = self.db.get_store_info()
            logging.getLogger('fontTools').setLevel(logging.ERROR)
            pdf = PDF(orientation='P', unit='mm', format='A4')
            pdf.store_info = store_info
            doc_type_raw = str(doc_type).upper().strip()
            stock_factor = AppConstants.STOCK_MOVEMENTS.get(doc_type_raw, 0)
            fin_factor = AppConstants.FINANCIAL_FACTORS.get(doc_type_raw, 0)
            is_transfer = stock_factor == 0 and fin_factor == 0 and (doc_type_raw in ['TR', 'TRANSFER', 'TRANSFERT'])
            is_supplier = (stock_factor == 1 or 'SUPPLIER' in doc_type_raw or 'PURCHASE' in doc_type_raw or ('BA' in doc_type_raw)) and (not is_transfer)
            final_entity_data = initial_entity
            display_name = ''
            ent_id = trans.get('entity_id')
            if ent_id:
                found = None
                if is_supplier:
                    found = self.db.get_entity_by_id(ent_id, 'supplier')
                else:
                    found = self.db.get_entity_by_id(ent_id, 'account')
                if found:
                    final_entity_data = found
                    display_name = found.get('name', '')
            if not display_name:
                display_name = trans.get('client_name') or initial_entity.get('name', '')
            default_check = ['COMPTOIR', 'INCONNU']
            if not display_name or any((d in str(display_name).upper() for d in default_check)):
                display_name = AppConstants.DEFAULT_SUPPLIER_NAME if is_supplier else AppConstants.DEFAULT_CLIENT_NAME
            if is_transfer:
                entity_label = 'Trajet'
                loc = trans.get('purchase_location') or trans.get('location') or 'store'
                display_name = 'Magasin -> Dépôt' if loc == 'store' else 'Dépôt -> Magasin'
                pdf.entity_info = {'label': entity_label, 'name': pdf.smart_text(display_name), 'address': '', 'phone': '', 'email': '', 'nif': '', 'rc': '', 'nis': '', 'nai': ''}
            else:
                entity_label = 'Fournisseur' if is_supplier else 'Client'
                pdf.entity_info = {'label': entity_label, 'name': pdf.smart_text(display_name), 'address': pdf.smart_text(final_entity_data.get('address', '')), 'phone': final_entity_data.get('phone', ''), 'email': final_entity_data.get('email', ''), 'nif': final_entity_data.get('nif', ''), 'rc': final_entity_data.get('rc', ''), 'nis': final_entity_data.get('nis', ''), 'nai': final_entity_data.get('nai', '')}
            visuals = AppConstants.DOC_VISUALS.get(doc_type_raw, {'name': doc_type_raw})
            doc_name_fr = visuals['name'].upper()
            if is_transfer:
                doc_name_fr = 'BON DE TRANSFERT'
            t_date = trans['date']
            try:
                if isinstance(t_date, str):
                    t_date = datetime.strptime(t_date, '%Y-%m-%d %H:%M:%S')
                date_str = t_date.strftime('%d-%m-%Y')
            except:
                date_str = str(t_date)[:10]
            invoice_num = trans.get('custom_label') or f"{doc_type}-{trans['id']:05d}"
            payment_method_val = ''
            timbre_val = Decimal(0)
            paid_val = Decimal(0)
            try:
                if trans.get('payment_details'):
                    pd = json.loads(trans.get('payment_details'))
                    payment_method_val = pd.get('method', '')
                    timbre_val = to_decimal(pd.get('timbre', 0))
                    paid_val = to_decimal(pd.get('amount', 0))
            except:
                pass
            if 'credit' in str(payment_method_val).lower() or 'crédit' in str(payment_method_val).lower():
                payment_method_val = ''
            pdf.doc_info = {'date': date_str, 'doc_name_fr': doc_name_fr, 'doc_number': invoice_num, 'payment_method': pdf.smart_text(payment_method_val) if not is_transfer else '', 'order_number': trans.get('order_number', ''), 'doc_type': doc_type}
            calc_ht_sum = Decimal(0)
            calc_tva_sum = Decimal(0)
            table_data = []
            printable_width = pdf.w - pdf.l_margin - pdf.r_margin
            if is_transfer:
                headers = ['N°', 'Code', 'Désignation', 'Qté']
                col_widths = [15, 35, 110, 30]
            else:
                is_bon = doc_type in ('BV', 'BA')
                if is_bon:
                    headers = ['N°', 'Code', 'Désignation', 'Qté', 'P.U.', 'Remise', 'Total']
                    col_widths = [10, 25, 0, 15, 25, 20, 25]
                else:
                    headers = ['N°', 'Code', 'Désignation', 'Qté', 'P.U. HT', 'Remise', 'TVA', 'Total HT']
                    col_widths = [10, 25, 0, 15, 25, 20, 15, 25]
                fixed_width = sum([w for w in col_widths if w > 0])
                col_widths[2] = max(20, printable_width - fixed_width)
            counter = 1
            for item in items:
                qty = to_decimal(item['qty'])
                qty_disp = str(int(qty)) if qty == qty.to_integral_value() else str(float(qty))
                prod_ref_code = item.get('product_ref', '') or ''
                raw_name = item.get('name', 'Produit') or item.get('product_name', 'Produit')
                ref_text = item.get('reference', '')
                full_name_disp = pdf.smart_text(raw_name)
                if ref_text:
                    full_name_disp += f' ({pdf.smart_text(ref_text)})'
                if is_transfer:
                    row = [str(counter), prod_ref_code, full_name_disp, qty_disp]
                else:
                    price = to_decimal(item['price'])
                    tva_rate = to_decimal(item.get('tva', 0))
                    line_ht = qty * price
                    calc_ht_sum += line_ht
                    line_tva = line_ht * (tva_rate / Decimal(100))
                    calc_tva_sum += line_tva
                    disc_amt = Decimal(0)
                    is_bon = doc_type in ('BV', 'BA')
                    row = [str(counter), prod_ref_code, full_name_disp, qty_disp, format_number_simple(price), format_number_simple(disc_amt)]
                    if not is_bon:
                        row.append(f'{float(tva_rate):g}%')
                    row.append(format_number_simple(line_ht))
                table_data.append(row)
                counter += 1
            if is_transfer:
                pdf.totals = {}
                pdf.amount_in_words = ''
                pdf.payment_info = {}
                pdf.balance_data = None
            else:
                stored_final_total = to_decimal(trans.get('total_amount', 0))
                stored_discount = to_decimal(trans.get('discount', 0))
                if timbre_val == 0 and doc_type == 'FC' and ('Espèce' in str(payment_method_val)):
                    base_calc = calc_ht_sum + calc_tva_sum
                    timbre_val = to_decimal(AppConstants.calculate_stamp_duty(base_calc))
                pdf.totals = {'total_ht': calc_ht_sum, 'total_tva': calc_tva_sum, 'total_discount': stored_discount, 'stamp_duty': timbre_val if timbre_val > 0 else None, 'final_total': stored_final_total}
                try:
                    pdf.amount_in_words = f'{number_to_words_fr(stored_final_total)} dinars algériens.'
                except:
                    pdf.amount_in_words = ''
                pdf.payment_info = {'amount': paid_val}
                balance_data = None
                show_balance_setting = self.db.get_setting('show_balance_in_pdf', 'False') == 'True'
                is_target_doc = doc_type_raw in ['BV', 'BA', 'SALE', 'PURCHASE']
                not_default = display_name != AppConstants.DEFAULT_CLIENT_NAME and display_name != AppConstants.DEFAULT_SUPPLIER_NAME
                if show_balance_setting and is_target_doc and ent_id and not_default and (fin_factor != 0):
                    try:
                        conn = self.db.get_connection()
                        cursor = conn.cursor()
                        target_table = 'suppliers' if is_supplier else 'clients'
                        cursor.execute(f'SELECT balance FROM {target_table} WHERE id = ?', (ent_id,))
                        row = cursor.fetchone()
                        current_balance_db = to_decimal(row[0]) if row else Decimal(0)
                        transaction_effect = stored_final_total * fin_factor - paid_val
                        old_balance = current_balance_db - transaction_effect
                        balance_data = {'old_balance': old_balance, 'transaction_amount': transaction_effect, 'new_balance': current_balance_db}
                        conn.close()
                    except Exception as e:
                        balance_data = None
                        print(f'Error calculating balance for PDF: {e}')
                pdf.balance_data = balance_data
            pdf.add_page()
            pdf.draw_table_with_fill(headers, table_data, col_widths)
            try:
                from android.storage import primary_external_storage_path
                report_dir = os.path.join(primary_external_storage_path(), 'Download')
            except ImportError:
                report_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
            if not os.path.exists(report_dir):
                os.makedirs(report_dir, exist_ok=True)
            safe_client = ''.join([c for c in display_name if c.isalnum() or c in (' ', '-', '_')]).strip().replace(' ', '_') or 'Client'
            safe_number = str(invoice_num).replace('/', '-').replace('\\', '-').replace(':', '')
            readable_type = visuals['name'].replace(' ', '_')
            file_name = f'{readable_type}_N_{safe_number}_{safe_client}.pdf'
            output_path = os.path.join(report_dir, file_name)
            pdf.output(output_path)
            from kivymd.uix.snackbar import MDSnackbar
            from kivymd.uix.label import MDLabel
            msg_text = f'Fichier PDF enregistré :\n{file_name}'
            MDSnackbar(MDLabel(text=msg_text, theme_text_color='Custom', text_color=(1, 1, 1, 1)), md_bg_color=(0, 0.6, 0, 1), duration=4).open()
        except Exception as e:
            self.notify(f'Erreur PDF: {e}', 'error')
            traceback.print_exc()

    def show_pending_dialog(self):
        content = MDBoxLayout(orientation='vertical', size_hint_y=None, height=dp(550))
        tabs_box = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(50), spacing=5)
        self.btn_hist_today = MDRaisedButton(text='AUJ.', size_hint_x=0.33, elevation=0, on_release=lambda x: self.filter_history_list(day_offset=0))
        self.btn_hist_yesterday = MDRaisedButton(text='HIER', size_hint_x=0.33, elevation=0, md_bg_color=(0.5, 0.5, 0.5, 1), on_release=lambda x: self.filter_history_list(day_offset=1))
        self.btn_hist_date = MDRaisedButton(text='CALENDRIER', size_hint_x=0.33, elevation=0, md_bg_color=(0.5, 0.5, 0.5, 1), on_release=self.open_history_date_picker)
        tabs_box.add_widget(self.btn_hist_today)
        tabs_box.add_widget(self.btn_hist_yesterday)
        tabs_box.add_widget(self.btn_hist_date)
        content.add_widget(tabs_box)
        self.rv_history = HistoryRecycleView()
        content.add_widget(self.rv_history)
        self.pending_dialog = MDDialog(title='Historique', type='custom', content_cls=content, size_hint=(0.98, 0.98))
        self.pending_dialog.open()
        self.filter_history_list(day_offset=0)

    def filter_history_list(self, day_offset=None, specific_date=None):
        inactive_color = (0.5, 0.5, 0.5, 1)
        active_color = self.theme_cls.primary_color
        target_date = None
        has_ui_elements = hasattr(self, 'btn_hist_today') and self.btn_hist_today
        if specific_date:
            target_date = specific_date
            if has_ui_elements:
                self.btn_hist_today.md_bg_color = inactive_color
                self.btn_hist_yesterday.md_bg_color = inactive_color
                self.btn_hist_date.md_bg_color = active_color
                self.btn_hist_date.text = str(specific_date)
        else:
            if day_offset is None:
                day_offset = 0
            target_date = datetime.now().date() - timedelta(days=day_offset)
            if has_ui_elements:
                self.btn_hist_today.md_bg_color = active_color if day_offset == 0 else inactive_color
                self.btn_hist_yesterday.md_bg_color = active_color if day_offset == 1 else inactive_color
                self.btn_hist_date.md_bg_color = inactive_color
                self.btn_hist_date.text = 'CALENDRIER'
        self.history_view_date = target_date
        self.history_page_offset = 0
        self.is_loading_history = False
        self.history_rv_data = []
        if hasattr(self, 'rv_history') and self.rv_history:
            self.rv_history.data = []
            self.rv_history.scroll_y = 1.0
            self.rv_history.refresh_from_data()
            self.rv_history.loading_lock = False
        self.load_more_history()

    def load_more_history(self):
        if self.is_loading_history:
            return
        self.is_loading_history = True
        threading.Thread(target=self._history_worker).start()

    def _history_worker(self):
        new_transactions = self.db.get_transactions(target_date=self.history_view_date, limit=self.history_batch_size, offset=self.history_page_offset)
        Clock.schedule_once(lambda dt: self._append_history_data(new_transactions))

    def _append_history_data(self, transactions):
        is_reset = self.history_page_offset == 0
        self.render_transactions_list(transactions, self.rv_history, is_global_mode=True, reset=is_reset)
        if transactions:
            self.history_page_offset += len(transactions)
        self.is_loading_history = False

    def open_pdf_file(self, file_path):
        if platform != 'android':
            try:
                os.startfile(file_path)
            except:
                pass
            return
        try:
            from jnius import autoclass, cast
            from android import activity
            File = autoclass('java.io.File')
            Intent = autoclass('android.content.Intent')
            FileProvider = autoclass('androidx.core.content.FileProvider')
            Context = autoclass('android.content.Context')
            file_obj = File(file_path)
            package_name = activity.getPackageName()
            uri = FileProvider.getUriForFile(Context.getApplicationContext(), f'{package_name}.fileprovider', file_obj)
            intent = Intent(Intent.ACTION_VIEW)
            intent.setDataAndType(uri, 'application/pdf')
            intent.setFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
            current_activity = cast('android.app.Activity', activity)
            current_activity.startActivity(intent)
        except Exception as e:
            print(f'PDF Open Error: {e}')
            self.notify("Impossible d'ouvrir le fichier PDF", 'error')

    def load_transaction_for_edit(self, header_data, items):
        if hasattr(self, 'srv_dialog') and self.srv_dialog:
            self.srv_dialog.dismiss()
        if hasattr(self, 'pending_dialog') and self.pending_dialog:
            self.pending_dialog.dismiss()
        self.editing_transaction_key = int(header_data['id'])
        doc_type = header_data.get('transaction_type', 'BV').upper()
        self.editing_doc_type = doc_type
        self.current_editing_date = header_data.get('date')
        try:
            pd = json.loads(header_data.get('payment_details', '{}'))
            self.editing_payment_amount = float(pd.get('amount', 0))
            self.editing_payment_method = pd.get('method', 'Espèce')
        except:
            self.editing_payment_amount = float(header_data.get('total_amount', 0))
            self.editing_payment_method = 'Espèce'
        target_mode = 'sale'
        if doc_type in ['CLIENT_PAY', 'SUPPLIER_PAY']:
            pass
        elif 'INVOICE' in doc_type or doc_type in ['FC', 'FF']:
            target_mode = 'invoice_purchase' if 'FF' in doc_type or 'PURCHASE' in doc_type else 'invoice_sale'
        elif 'RETURN' in doc_type or doc_type in ['RC', 'RF']:
            target_mode = 'return_purchase' if 'RF' in doc_type or 'PURCHASE' in doc_type else 'return_sale'
        elif 'TRANSFER' in doc_type or doc_type == 'TR':
            target_mode = 'transfer'
        elif 'PROFORMA' in doc_type or doc_type == 'FP':
            target_mode = 'proforma'
        elif 'ORDER' in doc_type or doc_type == 'DP':
            target_mode = 'order_purchase'
        else:
            stock_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
            if stock_f == 1:
                target_mode = 'purchase'
            else:
                target_mode = 'sale'
        if doc_type in ['CLIENT_PAY', 'SUPPLIER_PAY']:
            target_mode = 'supplier_payment' if 'SUPPLIER' in doc_type else 'client_payment'
            ent_id = header_data.get('entity_id')
            if ent_id:
                f = self.db.get_entity_by_id(ent_id, 'supplier' if 'SUPPLIER' in doc_type else 'account')
                self.selected_entity = f
            lbl = header_data.get('custom_label', '')
            self.temp_note = header_data.get('note', '') if header_data.get('note') != lbl else ''
            self.current_mode = target_mode
            self.show_simple_payment_dialog(amount=abs(float(header_data.get('total_amount', 0))))
            return
        self.open_mode(target_mode, skip_dialog=True)
        self.selected_location = header_data.get('location') or header_data.get('purchase_location') or 'store'
        ent_id = header_data.get('entity_id')
        found_entity = None
        is_supp_mode = AppConstants.get_entity_type(target_mode) == 'supplier'
        if ent_id:
            found_entity = self.db.get_entity_by_id(ent_id, 'supplier' if is_supp_mode else 'account')
        if found_entity is None:
            default_name = AppConstants.DEFAULT_SUPPLIER_NAME if is_supp_mode else AppConstants.DEFAULT_CLIENT_NAME
            found_entity = {'id': None, 'name': default_name, 'price_category': 'Gros' if is_supp_mode else 'Détail'}
        self.selected_entity = found_entity
        if self.selected_entity and hasattr(self, 'btn_ent_screen'):
            self.btn_ent_screen.text = self.fix_text(self.selected_entity.get('name', 'Client'))[:15]
        self.cart = []
        for item in items:
            cart_item = {'id': item.get('product_id') or item.get('id'), 'name': item.get('product_name'), 'price': float(item.get('price', 0)), 'qty': float(item.get('qty', 0)), 'tva': float(item.get('tva', 0))}
            if cart_item['id'] == 0 or str(cart_item['name']).startswith('Autre Article'):
                cart_item['id'] = -999
                cart_item['is_virtual'] = True
                cart_item['original_unit_price'] = cart_item['price']
            else:
                prod_db = self.db.get_product_by_id(cart_item['id'])
                if prod_db:
                    cart_item['original_unit_price'] = float(prod_db.get('price', 0))
            self.cart.append(cart_item)
        self.update_cart_button()
        self.notify(f'Modification: {doc_type} #{self.editing_transaction_key}', 'info')
        self.open_cart_screen()

    def confirm_delete_transaction(self, trans_id):
        if not trans_id:
            return

        def do_delete(x):
            try:
                if hasattr(self, 'confirm_del_dialog') and self.confirm_del_dialog:
                    self.confirm_del_dialog.dismiss()
                if hasattr(self, 'srv_dialog') and self.srv_dialog:
                    self.srv_dialog.dismiss()
                trans_type_to_update = 'account'
                try:
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    cursor.execute('SELECT transaction_type FROM transactions WHERE id=?', (trans_id,))
                    row = cursor.fetchone()
                    if row:
                        t_type = row[0]
                        if AppConstants.STOCK_MOVEMENTS.get(t_type) == 1 or 'SUPPLIER' in t_type:
                            trans_type_to_update = 'supplier'
                    conn.close()
                except:
                    pass
                self.db.delete_transaction(trans_id)
                if hasattr(self, 'pending_dialog') and self.pending_dialog:
                    try:
                        self.filter_history_list(0)
                    except:
                        pass
                if hasattr(self, 'entity_hist_dialog') and self.entity_hist_dialog:
                    try:
                        self.filter_entity_history_list(day_offset=0)
                    except:
                        pass
                self.load_local_entities(trans_type_to_update)
                self.check_and_load_stats()
                self.notify('Transaction supprimée', 'success')
            except Exception as e:
                self.notify(f'Erreur suppression: {e}', 'error')
                print(f'Delete Error: {e}')
        self.confirm_del_dialog = MDDialog(title='Confirmation', text='Voulez-vous vraiment supprimer cette opération ?\nCette action est irréversible et ajustera le stock/solde.', buttons=[MDFlatButton(text='NON', on_release=lambda x: self.confirm_del_dialog.dismiss()), MDRaisedButton(text='OUI', md_bg_color=(0.8, 0, 0, 1), text_color=(1, 1, 1, 1), on_release=do_delete)])
        self.confirm_del_dialog.open()

    def view_local_transaction_details(self, transaction_data):
        trans_id = transaction_data.get('id') if isinstance(transaction_data, dict) else transaction_data
        if not trans_id:
            return
        conn = self.db.get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM transactions WHERE id = ?', (trans_id,))
            header_row = cursor.fetchone()
            if not header_row:
                return
            cursor.execute('SELECT * FROM transaction_items WHERE transaction_id = ?', (trans_id,))
            items = [dict(r) for r in cursor.fetchall()]
            header_data = dict(header_row)
        except:
            return
        finally:
            conn.close()
        doc_type = header_data.get('transaction_type', 'BV').upper()
        stock_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
        is_supplier = stock_f == 1 or 'SUPPLIER' in doc_type or 'PURCHASE' in doc_type
        is_transfer = doc_type in ['TR', 'TRANSFER']
        ent_id = header_data.get('entity_id')
        raw_entity_name = header_data.get('client_name', '')
        if ent_id:
            found = self.db.get_entity_by_id(ent_id, 'supplier' if is_supplier else 'account')
            if found:
                raw_entity_name = found.get('name')
        if not raw_entity_name or raw_entity_name == 'COMPTOIR' or raw_entity_name == 'Inconnu':
            raw_entity_name = AppConstants.DEFAULT_SUPPLIER_NAME if is_supplier else AppConstants.DEFAULT_CLIENT_NAME
        vis = AppConstants.DOC_VISUALS.get(doc_type, {'name': doc_type})
        full_doc_name = vis['name']
        if header_data.get('custom_label'):
            full_doc_name = header_data.get('custom_label')
        total_amount = float(header_data.get('total_amount', 0))
        date_display = str(header_data.get('date', ''))
        paid_amount = total_amount
        timbre_val = 0.0
        try:
            pd = json.loads(header_data.get('payment_details', '{}'))
            paid_amount = float(pd.get('amount', total_amount))
            timbre_val = float(pd.get('timbre', 0))
        except:
            pass
        if doc_type == 'BI':
            paid_amount = total_amount
        content = MDBoxLayout(orientation='vertical', spacing=10, size_hint_y=None, height=dp(550))
        header_box = MDCard(orientation='vertical', adaptive_height=True, padding=dp(10), md_bg_color=(0.95, 0.95, 0.95, 1), radius=[10], elevation=0)
        entity_label = 'Fournisseur' if is_supplier else 'Client'
        if is_transfer:
            entity_label = 'Trajet'
            loc = header_data.get('location', 'store')
            raw_entity_name = 'Magasin -> Dépôt' if loc == 'store' else 'Dépôt -> Magasin'
        header_box.add_widget(MDLabel(text=self.fix_text(f'{full_doc_name}'), bold=True, font_style='Subtitle1', adaptive_height=True))
        header_box.add_widget(MDLabel(text=f'{entity_label}: {self.fix_text(raw_entity_name)}', font_style='Subtitle2', adaptive_height=True))
        header_box.add_widget(MDLabel(text=f'Date: {date_display}', font_style='Caption', theme_text_color='Secondary', adaptive_height=True))
        if not is_transfer:
            if timbre_val > 0:
                header_box.add_widget(MDLabel(text=f'Timbre: {timbre_val:.2f} DA', theme_text_color='Custom', text_color=(0.5, 0, 0.5, 1), font_style='Subtitle2', adaptive_height=True))
            header_box.add_widget(MDLabel(text=f'Montant: {total_amount:.2f} DA', theme_text_color='Custom', text_color=(0, 0, 0, 1), bold=True, font_style='H5', adaptive_height=True))
            remaining = total_amount - paid_amount
            if AppConstants.FINANCIAL_FACTORS.get(doc_type) != 0:
                if abs(remaining) < 0.05:
                    pay_row = MDBoxLayout(orientation='horizontal', adaptive_height=True, spacing=dp(5))
                    pay_row.add_widget(MDLabel(text='Payée', theme_text_color='Custom', text_color=(0, 0.7, 0, 1), bold=True, font_style='Subtitle1', adaptive_size=True))
                    pay_row.add_widget(MDIcon(icon='check-circle', theme_text_color='Custom', text_color=(0, 0.7, 0, 1), font_size='20sp', pos_hint={'center_y': 0.5}))
                    header_box.add_widget(pay_row)
                else:
                    label_pay = 'Règlement' if is_supplier else 'Versement'
                    header_box.add_widget(MDLabel(text=f'{label_pay}: {paid_amount:.2f} DA', theme_text_color='Custom', text_color=(0, 0.6, 0, 1), bold=True, font_style='Subtitle1', adaptive_height=True))
                    header_box.add_widget(MDLabel(text=f'Reste: {remaining:.2f} DA', theme_text_color='Custom', text_color=(0.9, 0, 0, 1), bold=True, font_style='Subtitle1', adaptive_height=True))
        content.add_widget(header_box)
        scroll = MDScrollView()
        list_layout = MDList()
        if items:
            for item in items:
                qty = float(item.get('qty', 0))
                qty_display = f'{qty:g}'
                ib = MDBoxLayout(orientation='vertical', adaptive_height=True, padding=[dp(10), dp(8)], spacing=dp(2))
                prod_name = self.fix_text(item.get('product_name', ''))
                ib.add_widget(MDLabel(text=prod_name, theme_text_color='Primary', font_style='Subtitle1', bold=True, adaptive_height=True))
                if is_transfer:
                    ib.add_widget(MDLabel(text=f'Qté: {qty_display}', theme_text_color='Custom', text_color=(0, 0.4, 0.8, 1), font_style='Subtitle2', bold=True, adaptive_height=True))
                else:
                    price = float(item.get('price', 0))
                    tva_rate = float(item.get('tva', 0))
                    line_ht = qty * price
                    if tva_rate > 0:
                        tva_val = line_ht * (tva_rate / 100)
                        line_ttc = line_ht + tva_val
                        ib.add_widget(MDLabel(text=f'{qty_display} x {price:.2f} DA = {line_ht:.2f} DA (HT)', theme_text_color='Secondary', font_style='Body2', adaptive_height=True))
                        tva_text = f'TVA ({int(tva_rate)}%): {tva_val:.2f} DA  |  TTC: {line_ttc:.2f} DA'
                        ib.add_widget(MDLabel(text=tva_text, theme_text_color='Custom', text_color=(0.8, 0, 0, 1), font_style='Caption', bold=True, adaptive_height=True))
                    else:
                        ib.add_widget(MDLabel(text=f'{qty_display} x {price:.2f} DA', theme_text_color='Secondary', font_style='Body2', adaptive_height=True))
                        ib.add_widget(MDLabel(text=f'Total: {line_ht:.2f} DA', theme_text_color='Secondary', font_style='Body2', adaptive_height=True))
                list_layout.add_widget(ib)
                list_layout.add_widget(MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.9, 0.9, 0.9, 1)))
        else:
            list_layout.add_widget(OneLineListItem(text='Opération Financière' if total_amount != 0 else 'Vide'))
        scroll.add_widget(list_layout)
        content.add_widget(scroll)
        actions = MDBoxLayout(orientation='vertical', spacing='10dp', adaptive_height=True, padding=[0, '15dp', 0, 0])
        top_row = MDBoxLayout(orientation='horizontal', spacing='10dp', size_hint_y=None, height='50dp')
        btn_print = MDFillRoundFlatButton(text='IMPRIMER', md_bg_color=(0, 0.5, 0.8, 1), text_color=(1, 1, 1, 1), size_hint_x=0.33)
        p_data = header_data.copy()
        p_data['items'] = items
        btn_print.bind(on_release=lambda x: threading.Thread(target=self.print_ticket_bluetooth, args=(p_data,), daemon=True).start())
        top_row.add_widget(btn_print)
        if doc_type not in ['CLIENT_PAY', 'SUPPLIER_PAY']:
            btn_pdf = MDFillRoundFlatButton(text='PDF', md_bg_color=(0.8, 0.2, 0.2, 1), text_color=(1, 1, 1, 1), size_hint_x=0.33)
            btn_pdf.bind(on_release=lambda x: self.generate_pdf_report(trans_id, doc_type))
            top_row.add_widget(btn_pdf)
        is_today = str(date_display).split(' ')[0] == str(datetime.now().date())
        if not self.is_seller_mode or is_today:
            btn_edit = MDFillRoundFlatButton(text='MODIFIER', md_bg_color=(0, 0.6, 0.4, 1), text_color=(1, 1, 1, 1), size_hint_x=0.33)
            btn_edit.bind(on_release=lambda x: self.load_transaction_for_edit(header_data, items))
            top_row.add_widget(btn_edit)
        actions.add_widget(top_row)
        if not self.is_seller_mode or is_today:
            btn_del = MDFlatButton(text='SUPPRIMER', theme_text_color='Custom', text_color=(0.9, 0, 0, 1), size_hint_x=1)
            btn_del.bind(on_release=lambda x: self.confirm_delete_transaction(trans_id))
            actions.add_widget(btn_del)
        content.add_widget(actions)
        self.srv_dialog = MDDialog(title='Détails', type='custom', content_cls=content, size_hint=(0.95, 0.95), buttons=[MDFlatButton(text='FERMER', on_release=lambda x: self.srv_dialog.dismiss())])
        self.srv_dialog.open()

    def open_barcode_scanner(self, instance):
        self.temp_scanned_cart = []
        self.potential_code = None
        self.consecutive_frames = 0
        self.last_scan_time = 0
        if not hasattr(self, 'target_scan_field') or instance is not None:
            self.target_scan_field = None
        if not decode:
            self.notify('Erreur: Librairie pyzbar manquante', 'error')
            return
        if platform == 'android':
            from android.permissions import request_permissions, Permission

            def on_permission_result(permissions, grants):
                if grants and grants[0]:
                    Clock.schedule_once(lambda dt: self._launch_camera_widget(), 0.1)
                else:
                    self.notify('Permission Caméra Refusée', 'error')
            request_permissions([Permission.CAMERA], on_permission_result)
        else:
            self._launch_camera_widget()

    def _launch_camera_widget(self):
        self.scanner_start_time = time.time()
        try:
            from kivy.uix.camera import Camera
            self.camera_widget = Camera(play=True, index=0, resolution=(640, 480), allow_stretch=True, keep_ratio=False)
            with self.camera_widget.canvas.before:
                PushMatrix()
                self.rotation = Rotate(angle=-90, origin=self.camera_widget.center)
            with self.camera_widget.canvas.after:
                PopMatrix()
            self.camera_widget.bind(center=lambda instance, value: setattr(self.rotation, 'origin', instance.center))
        except Exception as e:
            self.notify('Erreur chargement caméra', 'error')
            return
        is_single_mode = hasattr(self, 'target_scan_field') and self.target_scan_field is not None
        root_layout = MDBoxLayout(orientation='vertical', spacing=0)
        cam_size = 1.0 if is_single_mode else 0.55
        camera_area = MDFloatLayout(size_hint_y=cam_size)
        self.camera_widget.size_hint = (1, 1)
        self.camera_widget.pos_hint = {'center_x': 0.5, 'center_y': 0.5}
        camera_area.add_widget(self.camera_widget)
        close_btn = MDIconButton(icon='close', icon_size='36sp', md_bg_color=(0, 0, 0, 0.5), theme_text_color='Custom', text_color=(1, 1, 1, 1), pos_hint={'top': 0.96, 'right': 0.96}, on_release=self.close_barcode_scanner)
        camera_area.add_widget(close_btn)
        root_layout.add_widget(camera_area)
        if not is_single_mode:
            list_container = MDCard(orientation='vertical', size_hint_y=0.45, radius=[20, 20, 0, 0], md_bg_color=(1, 1, 1, 1), elevation=0)
            header = MDBoxLayout(size_hint_y=None, height=dp(40), padding=[dp(20), 0])
            self.lbl_scan_count = MDLabel(text='Total scannés: 0', bold=True, theme_text_color='Primary')
            header.add_widget(self.lbl_scan_count)
            list_container.add_widget(header)
            scroll = MDScrollView()
            self.scan_list_widget = MDList()
            scroll.add_widget(self.scan_list_widget)
            list_container.add_widget(scroll)
            btn_ok = MDRaisedButton(text='TERMINER', font_size='18sp', size_hint=(1, None), height=dp(55), md_bg_color=(0, 0.7, 0, 1), elevation=0, on_release=self.finish_continuous_scan)
            list_container.add_widget(btn_ok)
            root_layout.add_widget(list_container)
        self.scan_dialog = ModalView(size_hint=(1, 1), auto_dismiss=False, background_color=(0, 0, 0, 1))
        self.scan_dialog.add_widget(root_layout)
        self.scan_dialog.open()
        self.scan_event = Clock.schedule_interval(self.detect_barcode_frame, 1.0 / 20.0)

    def close_barcode_scanner(self, *args):
        if hasattr(self, 'scan_event') and self.scan_event:
            self.scan_event.cancel()
            self.scan_event = None
        if hasattr(self, 'camera_widget') and self.camera_widget:
            self.camera_widget.play = False
        if hasattr(self, 'scan_dialog') and self.scan_dialog:
            self.scan_dialog.dismiss()
            self.scan_dialog = None
        self.temp_scanned_cart = []
        self.target_scan_field = None

    def detect_barcode_frame(self, dt):
        if not hasattr(self, 'scan_dialog') or not self.scan_dialog or (not self.scan_dialog.parent):
            return
        if not hasattr(self, 'camera_widget') or not self.camera_widget.texture:
            return
        if time.time() - getattr(self, 'scanner_start_time', 0) < 1.2:
            return
        try:
            texture = self.camera_widget.texture
            img_data = PILImage.frombytes(mode='RGBA', size=texture.size, data=texture.pixels).convert('L')
            barcodes = decode(img_data)
            if barcodes:
                code = barcodes[0].data.decode('utf-8').strip()
                if code == self.potential_code:
                    self.consecutive_frames += 1
                else:
                    self.potential_code = code
                    self.consecutive_frames = 1
                if self.consecutive_frames >= 2:
                    if time.time() - self.last_scan_time > 1.5:
                        self.last_scan_time = time.time()
                        self.consecutive_frames = 0
                        self.potential_code = None
                        Clock.schedule_once(lambda dt: self.process_continuous_scan(code))
            else:
                self.consecutive_frames = 0
        except:
            pass

    def process_continuous_scan(self, code):
        if not hasattr(self, 'scan_dialog') or not self.scan_dialog.parent:
            return
        if hasattr(self, 'target_scan_field') and self.target_scan_field:
            self.target_scan_field.text = code
            self.play_sound('success')
            self.close_barcode_scanner()
            return
        prod = self.db.get_product_by_barcode(code)
        if prod:
            for item in self.temp_scanned_cart:
                if item['id'] == prod['id']:
                    self.play_sound('duplicate')
                    self.show_duplicate_alert(prod.get('name', 'Article'))
                    return
            self.temp_scanned_cart.append(prod)
            self.update_scan_list_ui()
            self.play_sound('success')
        else:
            self.play_sound('error')
            self.show_not_found_alert(code)

    def update_scan_list_ui(self):
        from kivymd.uix.card import MDCard
        from kivymd.uix.boxlayout import MDBoxLayout
        from kivymd.uix.label import MDLabel
        from kivymd.uix.button import MDIconButton
        from datetime import datetime
        self.scan_list_widget.clear_widgets()
        count = len(self.temp_scanned_cart)
        self.lbl_scan_count.text = f'Articles scannés: {count}'
        if count == 0:
            return
        customer_category = 'Détail'
        if self.selected_entity:
            customer_category = str(self.selected_entity.get('category', 'Détail')).strip()
        doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
        doc_type = doc_type_map.get(self.current_mode, 'BV')
        stock_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
        is_sales_mode = stock_f == -1 or doc_type in ['FP', 'FC', 'RC']
        for prod in reversed(self.temp_scanned_cart):
            raw_name = self.fix_text(prod.get('name', 'Inconnu'))
            raw_ref_text = str(prod.get('reference', '') or '').strip()
            display_text = raw_name
            if raw_ref_text:
                display_text += f' (Ref: {self.fix_text(raw_ref_text)})'
            final_price = 0.0
            if is_sales_mode:
                base_price = float(prod.get('price', 0) or 0)
                final_price = base_price
                if customer_category == 'Gros':
                    val = float(prod.get('price_wholesale', 0) or 0)
                    if val > 0:
                        final_price = val
                elif customer_category == 'Demi-Gros':
                    val = float(prod.get('price_semi', 0) or 0)
                    if val > 0:
                        final_price = val
                if prod.get('is_promo_active', 0) == 1:
                    promo_exp = str(prod.get('promo_expiry', '')).strip()
                    date_valid = True
                    if promo_exp and len(promo_exp) > 5:
                        try:
                            exp_date = datetime.strptime(promo_exp, '%Y-%m-%d').date()
                            if datetime.now().date() > exp_date:
                                date_valid = False
                        except:
                            pass
                    if date_valid:
                        p_type = prod.get('promo_type', 'fixed')
                        try:
                            p_val = float(prod.get('promo_value', 0))
                        except:
                            p_val = 0.0
                        if p_type == 'fixed':
                            if p_val > 0:
                                final_price = p_val
                        else:
                            final_price = base_price * (1 - p_val / 100)
            else:
                final_price = float(prod.get('purchase_price', prod.get('price', 0)) or 0)
            card = MDCard(orientation='horizontal', size_hint_y=None, height=dp(75), padding=[dp(15), 0, 0, 0], radius=[0], elevation=0, md_bg_color=(1, 1, 1, 1))
            text_box = MDBoxLayout(orientation='vertical', pos_hint={'center_y': 0.5}, adaptive_height=True, spacing=dp(4))
            lbl_name = MDLabel(text=display_text, font_style='Subtitle1', theme_text_color='Primary', shorten=False, max_lines=2, halign='left', adaptive_height=True)
            lbl_price = MDLabel(text=f'Prix: {final_price:.2f} DA', font_style='Caption', theme_text_color='Secondary', bold=True, halign='left', adaptive_height=True)
            text_box.add_widget(lbl_name)
            text_box.add_widget(lbl_price)
            del_btn = MDIconButton(icon='delete', theme_text_color='Custom', text_color=(0.9, 0, 0, 1), pos_hint={'center_y': 0.5}, icon_size='24sp', on_release=lambda x, p=prod: self.remove_temp_item(p))
            card.add_widget(text_box)
            card.add_widget(del_btn)
            sep = MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.95, 0.95, 0.95, 1))
            self.scan_list_widget.add_widget(card)
            self.scan_list_widget.add_widget(sep)

    def finish_continuous_scan(self, instance):
        if not hasattr(self, 'temp_scanned_cart') or not self.temp_scanned_cart:
            self.close_barcode_scanner()
            return
        items_to_add = list(self.temp_scanned_cart)
        self.close_barcode_scanner()
        count = 0
        for product in items_to_add:
            self.add_scanned_item_to_cart(product)
            count += 1
        if count > 0:
            self.notify(f'{count} Articles ajoutés au panier', 'success')

    def add_scanned_item_to_cart(self, product):
        try:
            doc_type_map = {'sale': 'BV', 'purchase': 'BA', 'return_sale': 'RC', 'return_purchase': 'RF', 'transfer': 'TR', 'invoice_sale': 'FC', 'invoice_purchase': 'FF', 'proforma': 'FP', 'order_purchase': 'DP'}
            current_mode_key = self.current_mode
            doc_type = doc_type_map.get(current_mode_key, 'BV')
            stock_f = AppConstants.STOCK_MOVEMENTS.get(doc_type, 0)
            final_price = 0.0
            if stock_f == -1 or doc_type in ['FP', 'FC', 'RC']:
                base_price = float(product.get('price', 0) or 0)
                final_price = base_price
                if self.selected_entity:
                    cat = str(self.selected_entity.get('category', 'Détail')).strip()
                    if cat == 'Gros':
                        w_price = float(product.get('price_wholesale', 0) or 0)
                        if w_price > 0:
                            final_price = w_price
                    elif cat == 'Demi-Gros':
                        s_price = float(product.get('price_semi', 0) or 0)
                        if s_price > 0:
                            final_price = s_price
                raw_active = product.get('is_promo_active', 0)
                is_promo = str(raw_active) == '1' or raw_active == 1
                if is_promo:
                    from datetime import datetime
                    promo_exp = str(product.get('promo_expiry', '')).strip()
                    date_valid = True
                    if len(promo_exp) > 5:
                        try:
                            exp_date = datetime.strptime(promo_exp, '%Y-%m-%d').date()
                            if datetime.now().date() > exp_date:
                                date_valid = False
                        except:
                            pass
                    if date_valid:
                        try:
                            p_val = float(product.get('promo_value', 0))
                        except:
                            p_val = 0.0
                        p_type = str(product.get('promo_type', 'fixed'))
                        if p_type == 'fixed':
                            if p_val > 0:
                                final_price = p_val
                        else:
                            final_price = base_price * (1 - p_val / 100)
            else:
                cost_price = float(product.get('purchase_price', 0) or 0)
                if cost_price == 0:
                    cost_price = float(product.get('price', 0) or 0)
                final_price = cost_price
            qty_to_add = 1.0
            found = False
            for item in self.cart:
                if str(item['id']) == str(product['id']):
                    item['qty'] += qty_to_add
                    item['price'] = final_price
                    found = True
                    break
            if not found:
                new_item = {'id': product['id'], 'name': product['name'], 'price': final_price, 'qty': qty_to_add, 'original_unit_price': final_price, 'tva': 0, 'is_virtual': False}
                if product.get('product_ref'):
                    new_item['product_ref'] = product.get('product_ref')
                self.cart.append(new_item)
            self.update_cart_button()
        except Exception as e:
            print(f'Add Cart Error: {e}')

    def remove_temp_item(self, product_to_remove):
        if product_to_remove in self.temp_scanned_cart:
            self.temp_scanned_cart.remove(product_to_remove)
            self.update_scan_list_ui()

    def show_duplicate_alert(self, product_name):
        if hasattr(self, 'is_showing_alert') and self.is_showing_alert:
            return
        self.is_showing_alert = True

        def close_alert(*args):
            self.dup_dialog.dismiss()
            self.is_showing_alert = False
        short_name = self.fix_text(product_name)[:30]
        self.dup_dialog = MDDialog(title='Déjà scanné !', text=f'Le produit:\n[b]{short_name}[/b]\n\nest déjà dans la liste.', buttons=[MDRaisedButton(text='OK', md_bg_color=(0.8, 0, 0, 1), on_release=close_alert)], size_hint=(0.85, None))
        self.dup_dialog.open()

    def show_not_found_alert(self, code):
        if hasattr(self, 'is_showing_alert') and self.is_showing_alert:
            return
        self.is_showing_alert = True

        def close(*args):
            self.not_found_dialog.dismiss()
            self.is_showing_alert = False
        self.not_found_dialog = MDDialog(title='Introuvable !', text=f"Le code-barres:\n[b]{code}[/b]\n\nn'existe pas dans la base de données.", buttons=[MDRaisedButton(text='OK', md_bg_color=(0.2, 0.2, 0.2, 1), on_release=close)], size_hint=(0.85, None))
        self.not_found_dialog.open()

    def get_backup_directory(self):
        if platform == 'android':
            try:
                from jnius import autoclass
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                currentActivity = PythonActivity.mActivity
                file_dir = currentActivity.getExternalFilesDir(None)
                if file_dir:
                    backup_dir = os.path.join(file_dir.getAbsolutePath(), 'MagPro_Backups')
                else:
                    backup_dir = os.path.join(self.user_data_dir, 'Backups')
            except Exception as e:
                print(f'Error getting android dir: {e}')
                backup_dir = os.path.join(self.user_data_dir, 'Backups')
        else:
            home = os.path.expanduser('~')
            backup_dir = os.path.join(home, 'Downloads', 'MagPro_Backups')
        if not os.path.exists(backup_dir):
            try:
                os.makedirs(backup_dir)
            except OSError as e:
                print(f'Error creating directory: {e}')
                return self.user_data_dir
        return backup_dir

    def _rotate_backups(self, backup_dir, limit=30):
        try:
            files = []
            for f in os.listdir(backup_dir):
                if f.startswith('Backup_Auto_') and f.endswith('.db'):
                    full_path = os.path.join(backup_dir, f)
                    files.append(full_path)
            files.sort(key=os.path.getmtime)
            while len(files) > limit:
                file_to_remove = files.pop(0)
                try:
                    os.remove(file_to_remove)
                    if os.path.exists(file_to_remove + '-wal'):
                        os.remove(file_to_remove + '-wal')
                    if os.path.exists(file_to_remove + '-shm'):
                        os.remove(file_to_remove + '-shm')
                    print(f'[INFO] Cleaning up old backup: {file_to_remove}')
                except Exception as e:
                    print(f'[WARN] Failed to delete old backup: {e}')
        except Exception as e:
            print(f'[ERROR] Backup rotation failed: {e}')

    def perform_local_backup(self, auto=False):
        try:
            if auto:
                filename = f"AutoBackup_{datetime.now().strftime('%Y-%m-%d')}.zip"
            else:
                filename = f"MagPro_Backup_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.zip"
            backup_path = self.get_unified_path(filename)
            backup_dir = os.path.dirname(backup_path)
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
            if os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                except:
                    pass
            temp_db_source = os.path.join(self.user_data_dir, 'temp_backup_source.db')
            if os.path.exists(temp_db_source):
                os.remove(temp_db_source)
            if self.db and self.db.conn:
                self.db.conn.execute(f"VACUUM INTO '{temp_db_source}'")
            else:
                self.db.connect()
                self.db.conn.execute(f"VACUUM INTO '{temp_db_source}'")
            import zipfile
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(temp_db_source, arcname='magpro_local.db')
                img_dir = os.path.join(self.user_data_dir, 'product_images')
                if os.path.exists(img_dir):
                    for root, dirs, files in os.walk(img_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.join('product_images', file)
                            zipf.write(file_path, arcname=arcname)
            if os.path.exists(temp_db_source):
                os.remove(temp_db_source)
            if not auto:
                self.notify(f'Sauvegarde réussie (DB+Images):\n{filename}', 'success')
                if platform != 'android':
                    import subprocess
                    subprocess.Popen(f'explorer /select,"{backup_path}"')
        except Exception as e:
            if not auto:
                self.notify(f'Échec de la sauvegarde: {e}', 'error')
            print(f'Backup Error: {e}')

    def show_restore_dialog(self):
        if platform == 'android':
            self.open_android_restore_picker()
            return
        try:
            default_path = os.path.join(os.path.expanduser('~'), 'Downloads')
            content = MDBoxLayout(orientation='vertical', spacing=10, size_hint_y=None, height=dp(550))
            list_background = MDCard(orientation='vertical', size_hint_y=0.8, md_bg_color=(0.15, 0.15, 0.15, 1), radius=[10], padding='5dp', elevation=0)
            self.file_chooser = FileChooserListView(path=default_path, filters=['*.zip', '*.db'], size_hint_y=1)
            list_background.add_widget(self.file_chooser)
            content.add_widget(list_background)
            btn_box = MDBoxLayout(orientation='horizontal', spacing=10, size_hint_y=None, height=dp(50))
            btn_cancel = MDFlatButton(text='ANNULER', on_release=lambda x: self.restore_dialog.dismiss())
            btn_confirm = MDRaisedButton(text='RESTAURER', md_bg_color=(1, 0, 0, 1), on_release=lambda x: self.confirm_restore_action(selected_path=None))
            btn_box.add_widget(btn_cancel)
            btn_box.add_widget(btn_confirm)
            content.add_widget(btn_box)
            self.restore_dialog = MDDialog(title='Choisir une sauvegarde', type='custom', content_cls=content, size_hint=(0.95, 0.9))
            self.restore_dialog.open()
        except Exception as e:
            self.notify(f'Erreur UI: {e}', 'error')

    def open_android_restore_picker(self):
        try:
            from jnius import autoclass, cast
            from android import activity
            Intent = autoclass('android.content.Intent')
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            intent = Intent(Intent.ACTION_GET_CONTENT)
            intent.setType('*/*')
            try:
                mimeTypes = ['application/vnd.sqlite3', 'application/x-sqlite3', 'application/octet-stream', 'application/zip', 'application/x-zip', 'application/x-zip-compressed']
                intent.putExtra(Intent.EXTRA_MIME_TYPES, mimeTypes)
            except:
                pass
            intent.addCategory(Intent.CATEGORY_OPENABLE)
            activity.bind(on_activity_result=self._on_restore_file_chosen)
            currentActivity = cast('android.app.Activity', PythonActivity.mActivity)
            currentActivity.startActivityForResult(intent, 102)
        except Exception as e:
            self.notify(f'Erreur sélecteur: {e}', 'error')

    def _on_restore_file_chosen(self, requestCode, resultCode, intent):
        from android import activity
        activity.unbind(on_activity_result=self._on_restore_file_chosen)
        if requestCode == 102 and resultCode == -1:
            if intent:
                uri = intent.getData()
                if uri:
                    self._copy_restore_uri_to_temp(uri)
                else:
                    self.notify('Aucun fichier sélectionné', 'error')
        else:
            self.notify('Restauration annulée', 'info')

    def _copy_restore_uri_to_temp(self, uri):
        self.notify('Préparation du fichier...', 'info')
        threading.Thread(target=self._background_restore_copy_task, args=(uri,), daemon=True).start()

    def _background_restore_copy_task(self, uri):
        pfd = None
        cursor = None
        try:
            from jnius import autoclass, cast
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            currentActivity = cast('android.app.Activity', PythonActivity.mActivity)
            content_resolver = currentActivity.getContentResolver()
            file_extension = '.db'
            try:
                OpenableColumns = autoclass('android.provider.OpenableColumns')
                cursor = content_resolver.query(uri, None, None, None, None)
                if cursor and cursor.moveToFirst():
                    name_index = cursor.getColumnIndex(OpenableColumns.DISPLAY_NAME)
                    if name_index >= 0:
                        file_name = cursor.getString(name_index)
                        if file_name and str(file_name).lower().endswith('.zip'):
                            file_extension = '.zip'
            except Exception as name_error:
                print(f'Name detection error: {name_error}')
            finally:
                if cursor:
                    cursor.close()
            pfd = content_resolver.openFileDescriptor(uri, 'r')
            fd = pfd.getFd()
            temp_filename = f'temp_restore{file_extension}'
            temp_file_path = os.path.join(self.user_data_dir, temp_filename)
            if os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
            with open(temp_file_path, 'wb') as f_out:
                while True:
                    try:
                        chunk = os.read(fd, 64 * 1024)
                    except OSError:
                        break
                    if not chunk:
                        break
                    f_out.write(chunk)
            try:
                pfd.close()
            except:
                pass
            Clock.schedule_once(lambda dt: self.confirm_restore_action(selected_path=temp_file_path), 0)
        except Exception as e:
            print(f'RESTORE COPY ERROR: {e}')
            if pfd:
                try:
                    pfd.close()
                except:
                    pass
            Clock.schedule_once(lambda dt: self.notify(f'Erreur copie: {str(e)}', 'error'), 0)

    def confirm_restore_action(self, selected_path=None):
        final_path = selected_path
        if not final_path:
            if hasattr(self, 'file_chooser') and self.file_chooser.selection:
                final_path = self.file_chooser.selection[0]
            else:
                self.notify("Veuillez sélectionner un fichier d'abord", 'error')
                return
        if hasattr(self, 'restore_dialog') and self.restore_dialog:
            self.restore_dialog.dismiss()

        def do_restore(x):
            try:
                if hasattr(self, 'final_restore_confirm'):
                    self.final_restore_confirm.dismiss()
                if self.db:
                    self.db.close()
                    time.sleep(0.2)
                if platform == 'android':
                    from jnius import autoclass
                    PythonActivity = autoclass('org.kivy.android.PythonActivity')
                    files_dir = PythonActivity.mActivity.getFilesDir().getAbsolutePath()
                    current_db_path = os.path.join(files_dir, AppConstants.DB_NAME)
                else:
                    app_dir = os.path.dirname(os.path.abspath(__file__))
                    current_db_path = os.path.join(app_dir, AppConstants.DB_NAME)
                try:
                    if os.path.exists(current_db_path + '-wal'):
                        os.remove(current_db_path + '-wal')
                    if os.path.exists(current_db_path + '-shm'):
                        os.remove(current_db_path + '-shm')
                except:
                    pass
                import zipfile
                if final_path.endswith('.zip'):
                    with zipfile.ZipFile(final_path, 'r') as zip_ref:
                        if 'magpro_local.db' in zip_ref.namelist():
                            if os.path.exists(current_db_path):
                                os.remove(current_db_path)
                            zip_ref.extract('magpro_local.db', os.path.dirname(current_db_path))
                        for file in zip_ref.namelist():
                            if file.startswith('product_images/'):
                                zip_ref.extract(file, self.user_data_dir)
                else:
                    if os.path.exists(current_db_path):
                        os.remove(current_db_path)
                    shutil.copyfile(final_path, current_db_path)
                if 'temp_restore' in final_path:
                    try:
                        os.remove(final_path)
                    except:
                        pass
                self.notify('Restauration réussie. Mise à jour...', 'success')
                self.db = DatabaseManager()
                self.load_more_products(reset=True)
                self.check_and_load_stats()
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.notify(f'Erreur critique: {e}', 'error')
                self.db = DatabaseManager()
        self.final_restore_confirm = MDDialog(title='Attention !', text=f'Voulez-vous vraiment restaurer ?\nLes données actuelles seront écrasées.', buttons=[MDFlatButton(text='ANNULER', on_release=lambda x: self.final_restore_confirm.dismiss()), MDRaisedButton(text='OUI, RESTAURER', md_bg_color=(1, 0, 0, 1), on_release=do_restore)])
        self.final_restore_confirm.open()

    def on_stop(self):
        try:
            print("[INFO] Arrêt de l'application...")
            if hasattr(self, 'db') and self.db:
                if self.db.conn:
                    self.db.close()
            if hasattr(self, 'db') and self.db:
                self.db.clean_up_wal()
        except Exception as e:
            print(f'Stop error: {e}')

    def share_database_file(self):
        try:
            # 1. إعداد اسم الملف والمسارات
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            zip_filename = f'MagPro_Backup_{timestamp}.zip'
            zip_path = self.get_unified_path(zip_filename)
            
            # مسار قاعدة البيانات المؤقتة لتجنب مشاكل القفل
            temp_db_path = os.path.join(self.user_data_dir, 'temp_share_source.db')
            if os.path.exists(temp_db_path):
                os.remove(temp_db_path)

            # 2. إنشاء نسخة نظيفة من قاعدة البيانات (Vacuum)
            if self.db and self.db.conn:
                self.db.conn.execute(f"VACUUM INTO '{temp_db_path}'")
            else:
                self.db.connect()
                self.db.conn.execute(f"VACUUM INTO '{temp_db_path}'")

            # 3. إنشاء ملف ZIP (يحتوي قاعدة البيانات + الصور)
            import zipfile
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # إضافة قاعدة البيانات
                zipf.write(temp_db_path, arcname='magpro_local.db')
                
                # إضافة مجلد الصور
                img_dir = os.path.join(self.user_data_dir, 'product_images')
                if os.path.exists(img_dir):
                    for root, dirs, files in os.walk(img_dir):
                        for file in files:
                            full_path = os.path.join(root, file)
                            arcname = os.path.join('product_images', file)
                            zipf.write(full_path, arcname=arcname)

            # حذف قاعدة البيانات المؤقتة بعد الضغط
            if os.path.exists(temp_db_path):
                os.remove(temp_db_path)

            # 4. عملية المشاركة (Android Intent)
            if platform == 'android':
                from jnius import autoclass, cast
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                Intent = autoclass('android.content.Intent')
                Uri = autoclass('android.net.Uri')
                File = autoclass('java.io.File')
                String = autoclass('java.lang.String')
                
                # تجاوز قيود Android لمشاركة الملفات المحلية (FileUriExposed)
                StrictMode = autoclass('android.os.StrictMode')
                Builder = autoclass('android.os.StrictMode$VmPolicy$Builder')
                StrictMode.setVmPolicy(Builder().build())

                # تحضير الملف للمشاركة
                zip_file_obj = File(zip_path)
                uri = Uri.fromFile(zip_file_obj)
                parcelable_uri = cast('android.os.Parcelable', uri)

                # إعداد الـ Intent
                shareIntent = Intent(Intent.ACTION_SEND)
                # استخدام application/zip يضمن ظهور تطبيقات مثل WhatsApp و Gmail
                shareIntent.setType('application/zip') 
                
                # إرفاق الملف
                shareIntent.putExtra(Intent.EXTRA_STREAM, parcelable_uri)
                
                # إضافة عنوان ونص (مفيد لـ Gmail)
                shareIntent.putExtra(Intent.EXTRA_SUBJECT, String(f"Sauvegarde MagPro: {timestamp}"))
                shareIntent.putExtra(Intent.EXTRA_TEXT, String("Veuillez trouver ci-joint le fichier de sauvegarde."))
                
                shareIntent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)

                currentActivity = cast('android.app.Activity', PythonActivity.mActivity)
                chooser_title = String(f'Partager la sauvegarde via...')
                currentActivity.startActivity(Intent.createChooser(shareIntent, chooser_title))

                # وظيفة لحذف الملف المضغوط بعد دقيقة للحفاظ على المساحة
                def delete_later(dt):
                    try:
                        if os.path.exists(zip_path):
                            os.remove(zip_path)
                    except:
                        pass
                Clock.schedule_once(delete_later, 60)
            
            else:
                # للكبيوتر (Windows/Linux)
                import subprocess
                subprocess.Popen(f'explorer /select,"{zip_path}"')
                self.notify(f"Fichier créé: {zip_filename}", "success")

        except Exception as e:
            self.notify(f'Erreur de partage: {e}', 'error')
            print(f"Share Error: {e}")



    def get_storage_path(self):
        if platform == 'android':
            try:
                from android.storage import primary_external_storage_path
                return os.path.join(primary_external_storage_path(), 'Download')
            except:
                return self.user_data_dir
        else:
            return os.path.join(os.path.expanduser('~'), 'Downloads')

    def get_android_documents_path(self, filename):
        if platform == 'android':
            try:
                from jnius import autoclass
                Environment = autoclass('android.os.Environment')
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                public_dir = Environment.getExternalStoragePublicDirectory(Environment.DIRECTORY_DOCUMENTS).getAbsolutePath()
                target_path = os.path.join(public_dir, filename)
                return target_path
            except:
                try:
                    context = PythonActivity.mActivity
                    file_dir = context.getExternalFilesDir(Environment.DIRECTORY_DOCUMENTS)
                    return os.path.join(file_dir.getAbsolutePath(), filename)
                except:
                    pass
        return os.path.join(os.path.expanduser('~'), 'Documents', filename)

    def perform_export(self):
        if hasattr(self, 'settings_menu_dialog') and self.settings_menu_dialog:
            self.settings_menu_dialog.dismiss()
        if openpyxl is None:
            self.notify('Erreur: Librairie openpyxl manquante.', 'error')
            return
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
        filename = f'MagPro_Stock_{timestamp}.xlsx'
        filepath = self.get_unified_path(filename)
        conn = self.db.get_connection()
        cursor = conn.cursor()
        try:
            query = 'SELECT product_ref, barcode, name, reference, stock, purchase_price, price, price_semi, price_wholesale FROM products'
            cursor.execute(query)
            rows = cursor.fetchall()
            if rows:
                headers = ['N° Produit', 'Code-Barres', 'Désignation', 'Référence', 'Stock', 'Prix Achat', 'Prix Détail', 'Prix Demi-Gros', 'Prix Gros']
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Produits'
                ws.append(headers)
                for row in rows:
                    ws.append(list(row))
                wb.save(filepath)
                try:
                    os.chmod(filepath, 511)
                except:
                    pass
                self.notify(f'Excel sauvegardé dans Téléchargements:\n{filename}', 'success')
                if platform != 'android':
                    import subprocess
                    subprocess.Popen(f'explorer /select,"{filepath}"')
            else:
                self.notify('Aucun produit à exporter.', 'warning')
        except Exception as e:
            self.notify(f'Erreur Export: {e}', 'error')
        finally:
            conn.close()

    def import_data_dialog(self):
        if platform == 'android':
            self.open_android_native_picker()
            return
        default_path = os.path.join(os.path.expanduser('~'), 'Downloads')
        if not os.path.exists(default_path):
            default_path = self.user_data_dir
        content = MDBoxLayout(orientation='vertical', spacing=10, size_hint_y=None, height=dp(500))
        content.add_widget(MDLabel(text='Sélectionnez un fichier Excel (.xlsx)', font_style='Caption', theme_text_color='Secondary', size_hint_y=None, height=dp(20)))
        list_bg = MDCard(orientation='vertical', size_hint_y=1, md_bg_color=(0.15, 0.15, 0.15, 1), radius=[10], padding='5dp', elevation=0)
        self.import_file_chooser = FileChooserListView(path=default_path, filters=['*.xlsx'], size_hint_y=1)
        list_bg.add_widget(self.import_file_chooser)
        content.add_widget(list_bg)
        btn_box = MDBoxLayout(spacing=10, size_hint_y=None, height=dp(50))
        btn_cancel = MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=self.theme_cls.primary_color, on_release=lambda x: self.import_diag.dismiss())
        btn_import = MDRaisedButton(text='SUIVANT', md_bg_color=(0.1, 0.4, 0.8, 1), text_color=(1, 1, 1, 1), on_release=lambda x: self.pre_process_import(self.import_file_chooser.selection))
        btn_box.add_widget(btn_cancel)
        btn_box.add_widget(btn_import)
        content.add_widget(btn_box)
        self.import_diag = MDDialog(title='Importer Produits', type='custom', content_cls=content, size_hint=(0.95, 0.9))
        self.import_diag.open()

    def pre_process_import(self, selection):
        if not selection:
            self.notify('Aucun fichier sélectionné.', 'error')
            return
        filepath = selection[0]
        if not filepath.endswith('.xlsx'):
            self.notify('Format invalide. Utilisez .xlsx', 'error')
            return
        if openpyxl is None:
            self.notify('Librairie manquante.', 'error')
            return
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                self.notify('Fichier vide.', 'error')
                return
            excel_headers = [str(h).strip() for h in rows[0]]
            excel_data = rows[1:]
            if hasattr(self, 'import_diag') and self.import_diag:
                self.import_diag.dismiss()
            self.open_mapping_dialog(excel_headers, excel_data)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.notify(f'Erreur lecture: {e}', 'error')

    def open_mapping_dialog(self, excel_headers, excel_data):

        def fix_txt(text):
            if not text:
                return ''
            text = str(text)
            try:
                if any(('\u0600' <= c <= 'ۿ' for c in text)):
                    return get_display(reshaper.reshape(text))
            except:
                pass
            return text
        display_options = ['-- Ignorer --'] + [fix_txt(h) for h in excel_headers]
        display_to_real = dict(zip(display_options, ['-- Ignorer --'] + excel_headers))
        db_fields = {'product_ref': 'N° Produit', 'name': 'Désignation*', 'barcode': 'Code-Barres', 'reference': 'Référence', 'price': 'Prix Détail', 'price_semi': 'Prix Demi-Gros', 'price_wholesale': 'Prix Gros', 'purchase_price': 'Prix Achat (Coût)', 'stock': 'Stock'}
        self.spinners = {}
        self.spinner_map = display_to_real
        content = MDBoxLayout(orientation='vertical', spacing=dp(10), size_hint_y=None, height=dp(550))
        content.add_widget(MDLabel(text='Associez les colonnes Excel aux champs :', font_style='Subtitle2', theme_text_color='Primary', size_hint_y=None, height=dp(30)))
        scroll = MDScrollView(size_hint_y=1)
        grid = MDGridLayout(cols=2, spacing=dp(20), padding=[dp(5), dp(10)], adaptive_height=True)
        for field_key, field_label in db_fields.items():
            lbl = MDLabel(text=field_label, size_hint_x=0.45, theme_text_color='Secondary', bold=True, size_hint_y=None, height=dp(45))
            spinner = Spinner(text='-- Ignorer --', values=display_options, size_hint_x=0.55, size_hint_y=None, height=dp(45), background_normal='', background_color=(0.3, 0.3, 0.3, 1), color=(1, 1, 1, 1), font_name='ArabicFont')
            for h in excel_headers:
                h_lower = str(h).lower().strip()
                matched = False
                if field_key == 'name':
                    if any((x in h_lower for x in ['nom', 'name', 'désignation', 'libelle'])):
                        matched = True
                    if 'article' in h_lower and (not any((c in h_lower for c in ['n°', 'num', 'code']))):
                        matched = True
                elif field_key == 'product_ref':
                    if any((x in h_lower for x in ['n° produit', 'numéro', 'internal_id', 'n° article'])):
                        matched = True
                elif field_key == 'barcode':
                    if any((x in h_lower for x in ['code', 'bar', 'ean', 'gencode'])):
                        matched = True
                elif field_key == 'reference':
                    if any((x in h_lower for x in ['référence', 'reference', 'ref_fr', 'ref'])):
                        matched = True
                elif field_key == 'stock':
                    if any((x in h_lower for x in ['stock', 'qté', 'quantité'])):
                        matched = True
                elif field_key == 'purchase_price':
                    if any((x in h_lower for x in ['achat', 'cost', 'coût', 'revient'])):
                        matched = True
                elif field_key == 'price':
                    if 'détail' in h_lower or 'vente' in h_lower or h_lower == 'prix' or (h_lower == 'prix unitaire'):
                        matched = True
                elif field_key == 'price_semi':
                    if 'demi' in h_lower:
                        matched = True
                elif field_key == 'price_wholesale':
                    if 'gros' in h_lower and 'demi' not in h_lower:
                        matched = True
                if matched:
                    spinner.text = fix_txt(h)
                    spinner.background_color = (0, 0.5, 0.2, 1)
                    break
            self.spinners[field_key] = spinner
            grid.add_widget(lbl)
            grid.add_widget(spinner)
        scroll.add_widget(grid)
        content.add_widget(scroll)
        footer = MDBoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10), padding=[0, dp(10), 0, 0])
        btn_cancel = MDFlatButton(text='ANNULER', theme_text_color='Custom', text_color=(0.6, 0.6, 0.6, 1), on_release=lambda x: self.mapping_diag.dismiss())
        btn_confirm = MDRaisedButton(text='TERMINER', md_bg_color=(0, 0.6, 0.2, 1), text_color=(1, 1, 1, 1), elevation=2, on_release=lambda x: self.finalize_mapping_import(excel_headers, excel_data))
        footer.add_widget(btn_cancel)
        footer.add_widget(btn_confirm)
        content.add_widget(footer)
        self.mapping_diag = MDDialog(title='', type='custom', content_cls=content, size_hint=(0.95, 0.95))
        self.mapping_diag.open()

    def finalize_mapping_import(self, excel_headers, excel_data):
        self.mapping_diag.dismiss()
        mapping_indices = {}
        for field, spinner in self.spinners.items():
            displayed_choice = spinner.text
            if displayed_choice != '-- Ignorer --':
                real_header = self.spinner_map.get(displayed_choice)
                if real_header and real_header in excel_headers:
                    mapping_indices[field] = excel_headers.index(real_header)
        if 'name' not in mapping_indices:
            self.notify("Erreur: Le champ 'Désignation' est obligatoire.", 'error')
            return
        standardized_headers = ['name', 'barcode', 'price', 'purchase_price', 'stock', 'price_semi', 'price_wholesale', 'product_ref', 'reference']
        final_data = []
        for row in excel_data:
            new_row = []
            for field in standardized_headers:
                if field in mapping_indices:
                    idx = mapping_indices[field]
                    val = row[idx] if idx < len(row) else ''
                    new_row.append(val if val is not None else '')
                else:
                    new_row.append('')
            final_data.append(new_row)
        added, skipped = self.bulk_insert_data('products', standardized_headers, final_data)
        msg = f'Terminé: {added} ajoutés + Bon Initial créé.'
        if skipped > 0:
            msg += f' ({skipped} doublons)'
        self.notify(msg, 'success')
        self.load_more_products(reset=True)

    def bulk_insert_data(self, table, headers, data):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        supplier_id = None
        try:
            cursor.execute('SELECT id FROM suppliers WHERE name = ?', (AppConstants.DEFAULT_SUPPLIER_NAME,))
            row = cursor.fetchone()
            if row:
                supplier_id = row[0]
            else:
                cursor.execute("INSERT INTO suppliers (name, price_category, balance) VALUES (?, 'Gros', 0)", (AppConstants.DEFAULT_SUPPLIER_NAME,))
                supplier_id = cursor.lastrowid
        except Exception:
            pass
        existing_names = set()
        existing_barcodes = set()
        try:
            cursor.execute('SELECT name, barcode FROM products')
            for row in cursor.fetchall():
                if row[0]:
                    existing_names.add(str(row[0]).strip().lower())
                if row[1]:
                    existing_barcodes.add(str(row[1]).strip())
        except:
            pass
        skipped_count = 0
        current_date_obj = datetime.now()
        timestamp_base = str(current_date_obj).split('.')[0]
        date_part = current_date_obj.strftime('%d%m')
        try:
            conn.execute('BEGIN TRANSACTION')
            for row in data:
                row_list = list(row)
                val_name = str(row_list[0]).strip().lower()
                if not val_name:
                    continue
                is_dup = False
                if val_name in existing_names:
                    is_dup = True
                if not is_dup:
                    val_bar = str(row_list[1]).strip()
                    if val_bar and val_bar in existing_barcodes:
                        is_dup = True
                if is_dup:
                    skipped_count += 1
                    continue
                for i in [2, 3, 4, 5, 6]:
                    try:
                        txt = str(row_list[i]).replace(',', '.').replace('DA', '').replace(' ', '').strip()
                        row_list[i] = float(txt) if txt else 0.0
                    except:
                        row_list[i] = 0.0
                placeholders = ', '.join(['?'] * len(headers))
                columns = ', '.join(headers)
                sql = f'INSERT INTO {table} ({columns}) VALUES ({placeholders})'
                cursor.execute(sql, row_list)
                new_product_id = cursor.lastrowid
                existing_names.add(val_name)
                if str(row_list[1]).strip():
                    existing_barcodes.add(str(row_list[1]).strip())
                qty = float(row_list[4])
                cost = float(row_list[3])
                if qty > 0 and supplier_id:
                    seq_name = 'SEQ_BI'
                    cursor.execute('UPDATE document_sequences SET current_value = current_value + 1 WHERE name = ?', (seq_name,))
                    if cursor.rowcount == 0:
                        cursor.execute('INSERT INTO document_sequences (name, current_value) VALUES (?, ?)', (seq_name, 1))
                        next_val = 1
                    else:
                        cursor.execute('SELECT current_value FROM document_sequences WHERE name = ?', (seq_name,))
                        next_val = cursor.fetchone()[0]
                    ref_number = f'BI{next_val:05d}/{date_part}'
                    total_amount = qty * cost
                    payment_json = json.dumps({'amount': total_amount, 'method': 'Initial', 'total': total_amount}, ensure_ascii=False)
                    cursor.execute('\n                        INSERT INTO transactions \n                        (transaction_type, total_amount, date, entity_id, custom_label, user_name, note, payment_details, location) \n                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)\n                    ', ('BI', total_amount, timestamp_base, supplier_id, ref_number, self.current_user_name, f'Stock Init: {row_list[0]}', payment_json, 'store'))
                    t_id = cursor.lastrowid
                    cursor.execute('\n                        INSERT INTO transaction_items \n                        (transaction_id, product_id, product_name, qty, price, tva, is_return, cost_price) \n                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)\n                    ', (t_id, new_product_id, row_list[0], qty, cost, 0, 0, cost))
            conn.commit()
            return (len(data) - skipped_count, skipped_count)
        except Exception as e:
            conn.rollback()
            print(f'Import Error: {e}')
            self.notify(f'Erreur Import: {e}', 'error')
            return (0, 0)
        finally:
            conn.close()

def console_excepthook(type, value, tb):
    print('!!! CONSOLE ERROR ENGINE !!!')
    traceback.print_exception(type, value, tb)

sys.excepthook = console_excepthook
if __name__ == '__main__':
    try:
        StockApp().run()
    except Exception as e:
        error_msg = traceback.format_exc()
        print('CRITICAL ERROR:', error_msg)
        try:
            with open('crash_log.txt', 'w', encoding='utf-8') as f:
                f.write(error_msg)
        except:
            pass
        try:
            from jnius import autoclass
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            files_dir = PythonActivity.mActivity.getExternalFilesDir(None).getAbsolutePath()
            log_path = os.path.join(files_dir, 'magpro_crash.txt')
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(error_msg)
        except:
            pass
